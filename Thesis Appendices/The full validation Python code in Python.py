#!/usr/bin/env python
# coding: utf-8

# # Well Metadata Semantic Validation Workflow
# 
# executable notebook containing Cells 1–12. Run every cell in order. 

# In[ ]:


# ============================================================
# CELL 1 — INSTALL REQUIRED PACKAGES
# ============================================================

get_ipython().run_line_magic('pip', 'install --upgrade pip')
get_ipython().run_line_magic('pip', 'install pandas requests openpyxl')


# In[ ]:


# ============================================================
# CELL 2 — IMPORT REQUIRED LIBRARIES
# ============================================================

import json  # Decode the PWLS JSON catalogue files.
import os  # Access environment variables and operating-system paths.
import xml.etree.ElementTree as ET  # Parse the PWLS Property Kind XML.

from datetime import datetime  # Format workbook creation and modification dates.
from io import BytesIO  # Process downloaded repository archives in memory.
from pathlib import Path  # Handle file and folder paths safely.
from zipfile import ZipFile  # Open and read downloaded ZIP archives.

import pandas as pd  # Create and manage metadata, catalogue, and alias DataFrames.
import requests  # Download the PWLS Curve Catalogue repository.

from IPython.display import display  # Display DataFrames clearly in JupyterLab.

from openpyxl import Workbook, load_workbook  # Create and load Excel workbooks.
from openpyxl.formatting.rule import Rule  # Create duplicate-value formatting rules.
from openpyxl.styles import Alignment, Font, PatternFill  # Format workbook cells.
from openpyxl.styles.differential import DifferentialStyle  # Define conditional-format styles.
from openpyxl.utils import get_column_letter  # Convert column numbers to Excel letters.

try:
    import tkinter as tk  # Provide native Windows file-selection windows.
    from tkinter import filedialog  # Provide the Windows file-dialog component.

    Tk = tk.Tk  # Keep the new Cell 7 compatible with the Tk() call.
    TKINTER_AVAILABLE = True

except ImportError:
    tk = None
    Tk = None
    filedialog = None
    TKINTER_AVAILABLE = False

print("Cell 2 imports loaded successfully.")


# In[ ]:


# ============================================================
# CELL 3 — SELECT AND LOAD METADATA CSV
# ============================================================

REQUIRED_METADATA_COLUMNS = ["Service company","mnemonic","unit","description"]  # Define the required metadata columns and canonical order.
METADATA_INPUT = {"path": None,"data": None}  # Store the selected file path and validated metadata.
METADATA_DF = None  # Hold the successfully loaded metadata DataFrame.


def choose_metadata_csv():
    """Open a native file dialog or request the CSV path manually."""

    root = None  # Hold the temporary Tkinter window.
    selected_path = ""  # Hold the selected or manually entered path.

    try:
        if tk is None or filedialog is None:
            raise RuntimeError("Tkinter is unavailable.")

        root = tk.Tk()  # Create a temporary Tkinter window.
        root.withdraw()  # Hide the empty root window.
        root.attributes("-topmost", True)  # Keep the file-selection dialog on top.
        root.update()  # Apply the window settings.

        selected_path = filedialog.askopenfilename(
            parent=root,
            title="Select the metadata CSV file",
            filetypes=[("CSV files","*.csv"),("All files","*.*")]
        )  # Open the native file-selection dialog.

    except Exception:
        print("The Windows file-selection window could not be opened.")
        print("Please paste the complete metadata CSV path below.")

        selected_path = input("Metadata CSV path: ")  # Fall back to manual path entry.

    finally:
        if root is not None:
            try:
                root.destroy()  # Always close the temporary Tkinter window.
            except Exception:
                pass

    raw_path = str(selected_path).strip().strip("\"'").strip()  # Remove surrounding spaces and quotation marks.

    if not raw_path:
        return None  # Represent cancellation or empty input safely.

    return Path(raw_path).expanduser()  # Return a clean Path object and expand "~" when present.


def read_metadata_csv(metadata_path):
    """Find, validate, and load the required metadata columns."""

    if not metadata_path.exists():
        raise FileNotFoundError(f"The selected file does not exist:\n{metadata_path}")

    if not metadata_path.is_file():
        raise ValueError(f"The selected path is not a file:\n{metadata_path}")

    if metadata_path.suffix.lower() != ".csv":
        raise ValueError("The metadata file must have the .csv extension.")

    read_options = {"dtype": str,"keep_default_na": False,"na_filter": False,"sep": None,"engine": "python"}  # Preserve text, empty cells, and detect the delimiter.

    try:
        metadata_df = pd.read_csv(metadata_path, encoding="utf-8-sig", **read_options)  # Read the CSV using UTF-8 first.
    except UnicodeDecodeError:
        metadata_df = pd.read_csv(metadata_path, encoding="cp1252", **read_options)  # Retry using the common Windows encoding.

    actual_columns = metadata_df.columns.tolist()  # Preserve all detected headers for matching and reporting.
    header_lookup = {}  # Connect normalized headers to their original forms.

    for column in actual_columns:
        normalized_header = str(column).strip().casefold()  # Ignore capitalization and surrounding spaces.
        header_lookup.setdefault(normalized_header, []).append(column)  # Store every matching original header.

    missing_columns = [
        column for column in REQUIRED_METADATA_COLUMNS
        if column.casefold() not in header_lookup
    ]  # Find required headers that are not present anywhere in the header row.

    if missing_columns:
        raise ValueError(
            "One or more required metadata columns could not be found.\n"
            "Capitalization, surrounding spaces, extra columns, and column order are accepted.\n\n"
            f"Missing columns:\n{missing_columns}\n\n"
            f"Required columns:\n{REQUIRED_METADATA_COLUMNS}\n\n"
            f"Detected columns:\n{actual_columns}"
        )

    ambiguous_columns = {
        column: header_lookup[column.casefold()]
        for column in REQUIRED_METADATA_COLUMNS
        if len(header_lookup[column.casefold()]) > 1
    }  # Detect multiple input headers matching the same required column.

    if ambiguous_columns:
        raise ValueError(
            "Multiple input columns match the same required metadata column.\n"
            "Remove or rename the duplicate columns before continuing.\n\n"
            f"Ambiguous matches:\n{ambiguous_columns}"
        )

    matched_columns = [
        header_lookup[column.casefold()][0]
        for column in REQUIRED_METADATA_COLUMNS
    ]  # Select each required column regardless of its input position or capitalization.

    ignored_columns = [
        column for column in actual_columns
        if column not in matched_columns
    ]  # Identify extra columns that will not enter the validation workflow.

    header_changes = [
        (matched_header, required_header)
        for matched_header, required_header in zip(matched_columns, REQUIRED_METADATA_COLUMNS)
        if matched_header != required_header
    ]  # Record headers that will receive their canonical names.

    metadata_df = metadata_df.loc[:, matched_columns].copy()  # Keep only the four required metadata columns.
    metadata_df.columns = REQUIRED_METADATA_COLUMNS  # Apply the canonical names and required output order.
    metadata_df.index = pd.RangeIndex(start=1, stop=len(metadata_df) + 1, name="input_order")  # Preserve the original record order.

    return metadata_df, header_changes, ignored_columns  # Return the validated data and import details.


selected_metadata_path = choose_metadata_csv()  # Ask the user to select the metadata CSV.

if selected_metadata_path is None:
    print("No metadata CSV was selected.")
    print("Run Cell 3 again when you are ready to select the file.")

else:
    try:
        METADATA_DF, header_changes, ignored_columns = read_metadata_csv(selected_metadata_path)  # Load and validate the selected CSV.

        METADATA_INPUT["path"] = selected_metadata_path  # Store the validated CSV path.
        METADATA_INPUT["data"] = METADATA_DF  # Store the validated DataFrame.

        print("Cell 3 loaded successfully.")
        print(f"Metadata file: {selected_metadata_path}")
        print(f"Rows loaded: {len(METADATA_DF):,}")

        if header_changes:
            print("\nHeaders normalized:")

            for original_header, canonical_header in header_changes:
                print(f"{original_header!r} -> {canonical_header!r}")  # Report each standardized header.

        if ignored_columns:
            print(f"\nExtra columns ignored: {ignored_columns}")  # Report columns excluded from the workflow.

        print("\nFirst 10 rows:")
        display(METADATA_DF.head(10))  # Preview the first 10 validated records with their input order.

    except Exception as error:
        METADATA_DF = None  # Remove any incomplete DataFrame.
        METADATA_INPUT["path"] = None  # Reset the stored path.
        METADATA_INPUT["data"] = None  # Reset the stored metadata.

        print("Cell 3 failed.")
        print(error)


# In[ ]:


# ============================================================
# CELL 4 — LOAD THE PWLS V4 CURVE CATALOG FILES
# ============================================================

# Requires json, BytesIO, ZipFile, pandas as pd, requests, and display from Cell 2.

PWLS_REPOSITORY_URL = "https://community.opengroup.org/energistics/pwls-curve-catalog"  # Identify the official PWLS Curve Catalog repository.
PWLS_GITLAB_API = "https://community.opengroup.org/api/v4"  # Define the GitLab API base URL.
PWLS_PROJECT_ID = "energistics%2Fpwls-curve-catalog"  # Store the encoded GitLab project path.
PWLS_BRANCH = "main"  # Select the maintained repository branch.
PWLS_REQUEST_TIMEOUT = (15, 180)  # Limit connection and download waiting times.

OBSOLETE_PWLS_NAMES = [
    "PWLS_CURVE_KEY","PWLS_TOOL_KEY","PWLS_TOOL_CURVE_KEY","PWLS_COMPANY_SUMMARY_DF",
    "PWLS_UNRESOLVED_RELATIONSHIPS_DF","PWLS_DUPLICATE_CURVE_KEYS_DF",
    "PWLS_DUPLICATE_TOOL_KEYS_DF","PWLS_DUPLICATE_TOOL_CURVE_KEYS_DF","PWLS_TABLES",
    "PWLS_COMPANY_ROW_INDEX","PWLS_CURVE_EXACT_INDEX","PWLS_GLOBAL_CURVE_EXACT_INDEX",
    "build_company_codes_dataframe","combine_pwls_frames","collect_duplicate_key_rows",
    "validate_pwls_relationships","get_pwls_company_catalog","find_pwls_curves"
]  # List variables and functions removed from the import-only Cell 4 design.

for obsolete_name in OBSOLETE_PWLS_NAMES:
    globals().pop(obsolete_name, None)  # Remove stale relationship or matching objects from an earlier Cell 4 run.

del obsolete_name, OBSOLETE_PWLS_NAMES  # Remove the temporary cleanup names from the notebook namespace.

PWLS_CATALOG = {}  # Hold the complete imported catalogue state.
PWLS_CATALOG_SOURCE = {}  # Store repository-level provenance.
PWLS_IMPORT_STATS = {}  # Store basic counts describing the files and rows loaded by Cell 4.
PWLS_CATALOG_FILES_DF = pd.DataFrame()  # List every imported JSON file and its repository path.
PWLS_FILE_DFS = {}  # Preserve one DataFrame for every JSON file, keyed by its source path.
PWLS_COMPANY_CODES_DF = pd.DataFrame()  # Hold the company-code file without imposing a required schema.
PWLS_CURVES_DF = pd.DataFrame()  # Hold all active curve-mapping records.
PWLS_CURVES_WITHIN_TOOLS_DF = pd.DataFrame()  # Hold all active tool-to-curve records.
PWLS_TOOLS_DF = pd.DataFrame()  # Hold all active tool-list records.
PWLS_ARCHIVE_DF = pd.DataFrame()  # Hold all records loaded from archive paths.
PWLS_OTHER_CATALOG_DFS = {}  # Preserve unclassified files by their source paths.


def classify_pwls_file(source_path):
    """Classify a file by its repository filename for DataFrame organization only."""

    file_name = source_path.rsplit("/", 1)[-1].casefold()  # Read the filename without changing any JSON data.
    file_types = {
        "company_codes_list.json": "company_codes",
        "curve_mappings.json": "curves",
        "curves_within_tools.json": "curves_within_tools",
        "tools_list.json": "tools"
    }  # Map known filenames to convenient DataFrame groups.

    if file_name.startswith("curve_mappings_") and file_name.endswith(".json"):
        return "curves"  # Organize dated curve-mapping files while preserving their source paths.

    return file_types.get(file_name, "other")  # Keep unfamiliar files available without assigning a semantic meaning.


def build_pwls_dataframe(payload, source_path, catalog_type, is_archive):
    """Load one JSON file into a DataFrame and add only source provenance."""

    if isinstance(payload, dict) and isinstance(payload.get("data"), list):
        catalog_df = pd.DataFrame.from_records(payload["data"])  # Load the source records exactly as supplied in the data list.
    elif isinstance(payload, list):
        catalog_df = pd.DataFrame.from_records(payload)  # Load a top-level JSON list without imposing a schema.
    else:
        catalog_df = pd.DataFrame([payload])  # Preserve an unfamiliar JSON structure as one source row.

    path_parts = source_path.split("/")  # Split the repository-relative path into its components.
    company_folder = path_parts[1] if len(path_parts) > 2 else ""  # Preserve the company folder declared by the repository path.
    source_metadata = payload if isinstance(payload, dict) else {}  # Read file-level metadata only when the JSON root is an object.
    provenance = [
        ("Source File", source_path),
        ("Source Record Order", range(1, len(catalog_df) + 1)),
        ("Catalog Company", company_folder),
        ("Catalog Type", catalog_type),
        ("Is Archive", is_archive),
        ("Source Last Updated", source_metadata.get("LastUpdated", "")),
        ("Schema Version", source_metadata.get("schemaVersion", "")),
        ("Schema URI", source_metadata.get("$schema", ""))
    ]  # Define source provenance without validating or interpreting catalogue relationships.

    for position, (column, value) in enumerate(provenance):
        if column not in catalog_df.columns:
            catalog_df.insert(position, column, value)  # Add provenance while leaving every original JSON field unchanged.

    company_position = len(provenance)  # Place file-level company identity after the source-provenance columns.

    if "Company Code" in source_metadata and "Company Code" not in catalog_df.columns:
        catalog_df.insert(company_position, "Company Code", source_metadata["Company Code"])  # Copy the declared file-level company code onto every record.
        company_position += 1

    if "Company Name" in source_metadata and "Company Name" not in catalog_df.columns:
        catalog_df.insert(company_position, "Company Name", source_metadata["Company Name"])  # Copy the declared file-level company name onto every record.

    catalog_df.index = pd.RangeIndex(start=1, stop=len(catalog_df) + 1, name="source_order")  # Preserve a one-based display order for the source file.

    return catalog_df


def combine_pwls_files(source_paths, file_dfs):
    """Combine selected file DataFrames without filtering or altering their records."""

    frames = [file_dfs[source_path] for source_path in source_paths]  # Select complete file DataFrames in repository-path order.

    if not frames:
        return pd.DataFrame()  # Return an empty DataFrame when the repository does not supply this file type.

    combined_df = pd.concat(frames, ignore_index=True, sort=False)  # Append every loaded record and preserve the union of source fields.
    combined_df.index = pd.RangeIndex(start=1, stop=len(combined_df) + 1, name="catalog_order")  # Add a one-based combined display order.

    return combined_df


def load_pwls_curve_catalog():
    """Download one repository snapshot and load every JSON file under catalog."""

    with requests.Session() as session:
        session.headers.update({"User-Agent": "PWLS-thesis-workflow/1.0"})  # Identify the notebook workflow to the source server.
        branch_url = f"{PWLS_GITLAB_API}/projects/{PWLS_PROJECT_ID}/repository/branches/{PWLS_BRANCH}"  # Build the branch-information endpoint.
        branch_response = session.get(branch_url, timeout=PWLS_REQUEST_TIMEOUT)  # Retrieve the current branch information.
        branch_response.raise_for_status()  # Stop only when the repository branch cannot be accessed.
        branch_payload = branch_response.json()  # Decode the branch response.
        source_commit = branch_payload.get("commit", {}).get("id", "")  # Preserve the exact repository commit used for this import.

        if not source_commit:
            raise ValueError("The PWLS source commit could not be determined.")

        archive_url = f"{PWLS_GITLAB_API}/projects/{PWLS_PROJECT_ID}/repository/archive.zip"  # Build the repository-snapshot endpoint.
        archive_response = session.get(archive_url, params={"sha": source_commit}, timeout=PWLS_REQUEST_TIMEOUT)  # Download all repository files in one request.
        archive_response.raise_for_status()  # Stop only when the repository archive cannot be downloaded.

    file_dfs = {}  # Store every loaded JSON DataFrame by its exact repository-relative path.
    file_inventory = []  # Store path and file-level provenance for the imported-file display.
    paths_by_type = {"company_codes": [],"curves": [],"curves_within_tools": [],"tools": [],"archive": [],"other": []}  # Organize paths without checking relationships.

    with ZipFile(BytesIO(archive_response.content)) as archive:
        catalog_members = sorted(
            name for name in archive.namelist()
            if "/catalog/" in name and name.casefold().endswith(".json")
        )  # Locate every JSON file recursively inside the catalog directory.

        if not catalog_members:
            raise FileNotFoundError("No JSON files were found in the PWLS catalog directory.")

        for archive_member in catalog_members:
            source_path = f"catalog/{archive_member.split('/catalog/', 1)[1]}"  # Preserve a stable repository-relative path.

            try:
                payload = json.loads(archive.read(archive_member).decode("utf-8-sig"))  # Decode the complete JSON file.
            except (UnicodeDecodeError, json.JSONDecodeError) as error:
                raise ValueError(f"The JSON file could not be loaded: {source_path}") from error

            catalog_type = classify_pwls_file(source_path)  # Organize the file from its filename only.
            is_archive = "/archive/" in f"/{source_path.casefold()}/"  # Preserve whether the source path belongs to an archive folder.
            path_parts = source_path.split("/")  # Split the source path for provenance.
            company_folder = path_parts[1] if len(path_parts) > 2 else ""  # Preserve the company folder from the repository path.
            source_metadata = payload if isinstance(payload, dict) else {}  # Read available top-level metadata without requiring any field.
            catalog_df = build_pwls_dataframe(payload, source_path, catalog_type, is_archive)  # Load the file without validating its content.

            file_dfs[source_path] = catalog_df  # Keep the complete DataFrame under its exact repository path.
            file_inventory.append({
                "Source File": source_path,
                "Catalog Type": catalog_type,
                "Catalog Company": company_folder,
                "Company Code": source_metadata.get("Company Code", ""),
                "Company Name": source_metadata.get("Company Name", ""),
                "Last Updated": source_metadata.get("LastUpdated", ""),
                "Schema Version": source_metadata.get("schemaVersion", ""),
                "Schema URI": source_metadata.get("$schema", ""),
                "Is Archive": is_archive
            })  # Preserve available file-level metadata without requiring, renaming, or calculating source fields.

            if is_archive:
                paths_by_type["archive"].append(source_path)  # Keep archive-path files separate from active master DataFrames.
            else:
                paths_by_type[catalog_type].append(source_path)  # Organize active files only by their known filenames.

    files_df = pd.DataFrame(file_inventory)  # Build the repository-path inventory.
    files_df.index = pd.RangeIndex(start=1, stop=len(files_df) + 1, name="file_order")  # Add a one-based file display order.
    company_codes_df = combine_pwls_files(paths_by_type["company_codes"], file_dfs)  # Combine the company-code file without schema assumptions.
    curves_df = combine_pwls_files(paths_by_type["curves"], file_dfs)  # Combine all active curve-mapping files.
    tool_curves_df = combine_pwls_files(paths_by_type["curves_within_tools"], file_dfs)  # Combine all active tool-to-curve files.
    tools_df = combine_pwls_files(paths_by_type["tools"], file_dfs)  # Combine all active tool-list files.
    archive_df = combine_pwls_files(paths_by_type["archive"], file_dfs)  # Combine every file loaded from an archive path.
    other_catalog_dfs = {source_path: file_dfs[source_path] for source_path in paths_by_type["other"]}  # Preserve unclassified active files by source path.
    active_company_catalogs = len({
        file_record["Catalog Company"]
        for file_record in file_inventory
        if file_record["Catalog Type"] == "curves" and not file_record["Is Archive"] and file_record["Catalog Company"]
    })  # Count company folders supplying an active curve-mapping file.
    import_stats = {
        "Catalog JSON files imported": len(files_df),
        "Active company catalogs": active_company_catalogs,
        "Active curve rows": len(curves_df),
        "Tool-curve relationship rows": len(tool_curves_df),
        "Tool definition rows": len(tools_df),
        "Archived rows kept separately": len(archive_df)
    }  # Record only basic import counts without analysing catalogue relationships.

    return {
        "source": {"repository": PWLS_REPOSITORY_URL,"branch": PWLS_BRANCH,"commit": source_commit},
        "import_stats": import_stats,
        "files": files_df,
        "file_dataframes": file_dfs,
        "company_codes": company_codes_df,
        "curves": curves_df,
        "curves_within_tools": tool_curves_df,
        "tools": tools_df,
        "archive": archive_df,
        "other": other_catalog_dfs
    }  # Return only imported data and source provenance.


try:
    PWLS_CATALOG = load_pwls_curve_catalog()  # Load every PWLS catalogue JSON file from one repository snapshot.
    PWLS_CATALOG_SOURCE = PWLS_CATALOG["source"]  # Store repository and commit provenance.
    PWLS_IMPORT_STATS = PWLS_CATALOG["import_stats"]  # Store the basic file and row counts produced during loading.
    PWLS_CATALOG_FILES_DF = PWLS_CATALOG["files"]  # Store the complete repository-path inventory.
    PWLS_FILE_DFS = PWLS_CATALOG["file_dataframes"]  # Store each file DataFrame under its exact source path.
    PWLS_COMPANY_CODES_DF = PWLS_CATALOG["company_codes"]  # Store the company-code file without required-column checks.
    PWLS_CURVES_DF = PWLS_CATALOG["curves"]  # Store all active curve-mapping records.
    PWLS_CURVES_WITHIN_TOOLS_DF = PWLS_CATALOG["curves_within_tools"]  # Store all active tool-to-curve records.
    PWLS_TOOLS_DF = PWLS_CATALOG["tools"]  # Store all active tool-list records.
    PWLS_ARCHIVE_DF = PWLS_CATALOG["archive"]  # Store all records from archive paths.
    PWLS_OTHER_CATALOG_DFS = PWLS_CATALOG["other"]  # Store unclassified files by source path.

    print("Cell 4 loaded successfully.")  # Confirm that the PWLS files are available.
    print(f"Source commit: {PWLS_CATALOG_SOURCE['commit']}")  # Display the exact repository version used.

    for statistic_name, statistic_value in PWLS_IMPORT_STATS.items():
        print(f"{statistic_name}: {statistic_value:,}")  # Display each basic import count with readable thousands separators.

    print("\nImported source files:")  # Introduce the repository-path inventory.
    display(PWLS_CATALOG_FILES_DF)  # Display every imported JSON path and its available file metadata.

except Exception as error:
    PWLS_CATALOG = {}  # Clear the combined state after a loading failure.
    PWLS_CATALOG_SOURCE = {}  # Clear repository provenance.
    PWLS_IMPORT_STATS = {}  # Clear the basic import counts.
    PWLS_CATALOG_FILES_DF = pd.DataFrame()  # Clear the source-file inventory.
    PWLS_FILE_DFS = {}  # Clear the per-file DataFrames.
    PWLS_COMPANY_CODES_DF = pd.DataFrame()  # Clear the company-code records.
    PWLS_CURVES_DF = pd.DataFrame()  # Clear the active curve records.
    PWLS_CURVES_WITHIN_TOOLS_DF = pd.DataFrame()  # Clear the tool-to-curve records.
    PWLS_TOOLS_DF = pd.DataFrame()  # Clear the tool-list records.
    PWLS_ARCHIVE_DF = pd.DataFrame()  # Clear archived records.
    PWLS_OTHER_CATALOG_DFS = {}  # Clear unclassified file DataFrames.

    print("Cell 4 failed.")  # Report a concise loading failure.
    print(error)  # Display the specific download, ZIP, or JSON decoding error.


# In[ ]:


# ============================================================
# CELL 5 — LOAD THE PWLS V4 PROPERTY KIND DICTIONARY
# ============================================================

# Requires xml.etree.ElementTree as ET, pandas as pd, requests, and display from Cell 2.

PWLS_PROPERTY_REPOSITORY_URL = "https://github.com/younesbell-cell/uom-for-my-project-"  # Identify the source repository.
PWLS_PROPERTY_GITHUB_API = "https://api.github.com"  # Define the GitHub API base URL.
PWLS_PROPERTY_REPOSITORY = "younesbell-cell/uom-for-my-project-"  # Store the repository owner and name.
PWLS_PROPERTY_BRANCH = "main"  # Select the maintained repository branch.
PWLS_PROPERTY_FILE_PATH = "PWLS 4.0/PWLS_PropertyKindDictionary_v4.0_eml23.xml"  # Preserve the repository-relative XML path.
PWLS_PROPERTY_FILE_PATH_ENCODED = "PWLS%204.0/PWLS_PropertyKindDictionary_v4.0_eml23.xml"  # Encode the path for the raw GitHub URL.
PWLS_PROPERTY_REQUEST_TIMEOUT = (15, 180)  # Limit connection and download waiting times.

PWLS_PROPERTY_DICTIONARY = {}  # Store all loaded Property Kind Dictionary results.
PWLS_PROPERTY_SOURCE = {}  # Store repository, commit, and file provenance.
PWLS_PROPERTY_DICTIONARY_METADATA = {}  # Store dictionary-level attributes and citation fields.
PWLS_PROPERTY_XML_ROOT = None  # Preserve the complete parsed XML root for later cells.
PWLS_PROPERTY_KINDS_DF = pd.DataFrame()  # Hold one row for every PropertyKind element.
PWLS_PROPERTY_ALIASES_DF = pd.DataFrame()  # Hold one row for every PropertyKind alias.
PWLS_PROPERTY_STATS = {}  # Store basic statistics about the loaded XML.


def xml_local_name(xml_tag):
    """Return an XML tag or attribute name without its namespace URI."""

    return xml_tag.rsplit("}", 1)[-1]  # Remove the namespace wrapper while preserving the local name.


def xml_children(xml_element, child_name):
    """Return direct children having the requested local name."""

    if xml_element is None:
        return []  # Return no children when the parent element is absent.

    return [
        child for child in xml_element
        if xml_local_name(child.tag) == child_name
    ]  # Select direct children without depending on a namespace prefix.


def xml_child(xml_element, child_name):
    """Return the first direct child having the requested local name."""

    children = xml_children(xml_element, child_name)  # Locate matching direct children.

    return children[0] if children else None  # Return the first match or None.


def xml_text(xml_element, child_name):
    """Return clean text from a direct child without changing capitalization."""

    child = xml_child(xml_element, child_name)  # Locate the requested child element.

    if child is None or child.text is None:
        return ""  # Preserve missing XML values as empty strings.

    return child.text.strip()  # Remove XML formatting whitespace.


def xml_attribute(xml_element, attribute_name):
    """Return an attribute value by local name."""

    if xml_element is None:
        return ""  # Return an empty value when the element is absent.

    for xml_name, xml_value in xml_element.attrib.items():
        if xml_local_name(xml_name) == attribute_name:
            return xml_value  # Return the matching attribute unchanged.

    return ""  # Return an empty value when the attribute is absent.


def parse_pwls_property_dictionary(xml_content, source_commit):
    """Parse the XML into metadata, Property Kind rows, and alias rows."""

    xml_root = ET.fromstring(xml_content)  # Parse the downloaded XML bytes.

    if xml_local_name(xml_root.tag) != "PropertyKindDictionary":
        raise ValueError(
            f"Unexpected XML root element: {xml_local_name(xml_root.tag)}"
        )

    dictionary_citation = xml_child(xml_root, "Citation")  # Locate the dictionary citation information.

    dictionary_metadata = {
        "Source Repository": PWLS_PROPERTY_REPOSITORY_URL,
        "Source Commit": source_commit,
        "Source File": PWLS_PROPERTY_FILE_PATH,
        "Root Element": xml_local_name(xml_root.tag),
        "Namespace URI": xml_root.tag[1:].split("}", 1)[0] if xml_root.tag.startswith("{") else "",
        "Schema Version": xml_attribute(xml_root, "schemaVersion"),
        "UUID": xml_attribute(xml_root, "uuid"),
        "Schema Location": xml_attribute(xml_root, "schemaLocation"),
        "Title": xml_text(dictionary_citation, "Title"),
        "Originator": xml_text(dictionary_citation, "Originator"),
        "Creation": xml_text(dictionary_citation, "Creation"),
        "Format": xml_text(dictionary_citation, "Format"),
        "Last Update": xml_text(dictionary_citation, "LastUpdate"),
        "Description": xml_text(dictionary_citation, "Description")
    }  # Preserve dictionary-level attributes and citation values.

    property_rows = []  # Collect one flat row for each PropertyKind element.
    alias_rows = []  # Collect each repeating alias in a separate row.
    property_elements = xml_children(xml_root, "PropertyKind")  # Locate direct PropertyKind elements.

    for property_order, property_element in enumerate(property_elements, start=1):
        citation_element = xml_child(property_element, "Citation")  # Locate this Property Kind citation.
        parent_element = xml_child(property_element, "Parent")  # Locate its parent reference.
        alias_elements = xml_children(property_element, "Aliases")  # Locate every attached alias.
        property_uuid = xml_attribute(property_element, "uuid")  # Read the Property Kind UUID.
        property_title = xml_text(citation_element, "Title")  # Read its canonical title.

        property_rows.append({
            "Source File": PWLS_PROPERTY_FILE_PATH,
            "Source Commit": source_commit,
            "Property Record Order": property_order,
            "Schema Version": xml_attribute(property_element, "schemaVersion"),
            "UUID": property_uuid,
            "Title": property_title,
            "Originator": xml_text(citation_element, "Originator"),
            "Creation": xml_text(citation_element, "Creation"),
            "Format": xml_text(citation_element, "Format"),
            "Last Update": xml_text(citation_element, "LastUpdate"),
            "Description": xml_text(citation_element, "Description"),
            "Is Abstract": xml_text(property_element, "IsAbstract"),
            "Deprecation Date": xml_text(property_element, "DeprecationDate"),
            "Quantity Class": xml_text(property_element, "QuantityClass"),
            "Parent UUID": xml_text(parent_element, "Uuid"),
            "Parent Qualified Type": xml_text(parent_element, "QualifiedType"),
            "Parent Title": xml_text(parent_element, "Title"),
            "Alias Count": len(alias_elements)
        })  # Preserve the fields belonging to this Property Kind.

        for alias_order, alias_element in enumerate(alias_elements, start=1):
            alias_rows.append({
                "Source File": PWLS_PROPERTY_FILE_PATH,
                "Source Commit": source_commit,
                "Property Record Order": property_order,
                "Property UUID": property_uuid,
                "Property Title": property_title,
                "Alias Order": alias_order,
                "Authority": xml_attribute(alias_element, "authority"),
                "Identifier": xml_text(alias_element, "Identifier")
            })  # Preserve every alias and its authority separately.

    property_kinds_df = pd.DataFrame(property_rows)  # Build the complete Property Kind table.
    aliases_df = pd.DataFrame(alias_rows)  # Build the complete alias table.

    property_kinds_df.index = pd.RangeIndex(
        start=1,
        stop=len(property_kinds_df) + 1,
        name="property_order"
    )  # Add a one-based Property Kind index.

    aliases_df.index = pd.RangeIndex(
        start=1,
        stop=len(aliases_df) + 1,
        name="alias_order"
    )  # Add a one-based alias index.

    abstract_values = (
        property_kinds_df["Is Abstract"]
        .astype(str)
        .str.strip()
        .str.casefold()
    )  # Standardize only the Boolean values required for counting.

    quantity_classes = [
        value for value in property_kinds_df["Quantity Class"].tolist()
        if str(value).strip()
    ]  # Collect populated Quantity Class values.

    parent_uuids = [
        value for value in property_kinds_df["Parent UUID"].tolist()
        if str(value).strip()
    ]  # Collect populated parent references.

    property_stats = {
        "Property Kind records": len(property_kinds_df),
        "Concrete properties": int((abstract_values == "false").sum()),
        "Abstract Property Kinds": int((abstract_values == "true").sum()),
        "Alias records": len(aliases_df),
        "Unique Quantity Classes": len(set(quantity_classes)),
        "Parent references": len(parent_uuids),
        "Unique parents": len(set(parent_uuids))
    }  # Calculate structural statistics without creating relationships.

    return {
        "source": {
            "repository": PWLS_PROPERTY_REPOSITORY_URL,
            "branch": PWLS_PROPERTY_BRANCH,
            "commit": source_commit,
            "file": PWLS_PROPERTY_FILE_PATH
        },
        "metadata": dictionary_metadata,
        "xml_root": xml_root,
        "property_kinds": property_kinds_df,
        "aliases": aliases_df,
        "stats": property_stats
    }  # Return the complete parsed content and its provenance.


def load_pwls_property_dictionary():
    """Resolve the GitHub commit, download the XML, and parse its contents."""

    with requests.Session() as session:
        session.headers.update({
            "Accept": "application/vnd.github+json",
            "User-Agent": "PWLS-thesis-workflow/1.0"
        })  # Identify the notebook workflow to GitHub.

        commit_url = (
            f"{PWLS_PROPERTY_GITHUB_API}/repos/"
            f"{PWLS_PROPERTY_REPOSITORY}/commits/{PWLS_PROPERTY_BRANCH}"
        )  # Build the branch-commit API endpoint.

        commit_response = session.get(
            commit_url,
            timeout=PWLS_PROPERTY_REQUEST_TIMEOUT
        )  # Resolve the current commit of the selected branch.

        commit_response.raise_for_status()  # Stop if the commit information cannot be retrieved.
        source_commit = commit_response.json().get("sha", "")  # Preserve the exact source commit.

        if not source_commit:
            raise ValueError(
                "The Property Kind Dictionary source commit could not be determined."
            )

        raw_file_url = (
            f"https://raw.githubusercontent.com/"
            f"{PWLS_PROPERTY_REPOSITORY}/"
            f"{source_commit}/"
            f"{PWLS_PROPERTY_FILE_PATH_ENCODED}"
        )  # Build an immutable raw URL using the resolved commit.

        xml_response = session.get(
            raw_file_url,
            timeout=PWLS_PROPERTY_REQUEST_TIMEOUT
        )  # Download the Property Kind Dictionary XML.

        xml_response.raise_for_status()  # Stop if the XML cannot be downloaded.

    return parse_pwls_property_dictionary(
        xml_response.content,
        source_commit
    )  # Parse the downloaded XML bytes.


try:
    PWLS_PROPERTY_DICTIONARY = load_pwls_property_dictionary()  # Download and parse the Property Kind Dictionary.
    PWLS_PROPERTY_SOURCE = PWLS_PROPERTY_DICTIONARY["source"]  # Store source provenance.
    PWLS_PROPERTY_DICTIONARY_METADATA = PWLS_PROPERTY_DICTIONARY["metadata"]  # Store dictionary metadata.
    PWLS_PROPERTY_XML_ROOT = PWLS_PROPERTY_DICTIONARY["xml_root"]  # Preserve the complete XML root.
    PWLS_PROPERTY_KINDS_DF = PWLS_PROPERTY_DICTIONARY["property_kinds"]  # Store Property Kind records.
    PWLS_PROPERTY_ALIASES_DF = PWLS_PROPERTY_DICTIONARY["aliases"]  # Store alias records.
    PWLS_PROPERTY_STATS = PWLS_PROPERTY_DICTIONARY["stats"]  # Store structural statistics.

    print("Cell 5 loaded successfully.")  # Confirm that the dictionary is available.
    print(f"Source commit: {PWLS_PROPERTY_SOURCE['commit']}")  # Display the exact imported version.
    print(f"Source file: {PWLS_PROPERTY_SOURCE['file']}")  # Display the repository-relative source path.
    print(f"Dictionary schema version: {PWLS_PROPERTY_DICTIONARY_METADATA['Schema Version']}")  # Display the schema version.

    for statistic_name, statistic_value in PWLS_PROPERTY_STATS.items():
        print(f"{statistic_name}: {statistic_value:,}")  # Display each structural statistic.

    print("\nFirst Property Kind records:")  # Introduce a short data preview.
    display(PWLS_PROPERTY_KINDS_DF.head(10))  # Display the first ten Property Kind records.

except Exception as error:
    PWLS_PROPERTY_DICTIONARY = {}  # Clear the combined result after failure.
    PWLS_PROPERTY_SOURCE = {}  # Clear source provenance.
    PWLS_PROPERTY_DICTIONARY_METADATA = {}  # Clear dictionary metadata.
    PWLS_PROPERTY_XML_ROOT = None  # Clear the XML root.
    PWLS_PROPERTY_KINDS_DF = pd.DataFrame()  # Clear Property Kind records.
    PWLS_PROPERTY_ALIASES_DF = pd.DataFrame()  # Clear alias records.
    PWLS_PROPERTY_STATS = {}  # Clear structural statistics.

    print("Cell 5 failed.")  # Report a concise cell-level failure.
    print(error)  # Display the specific download or XML parsing error.


# In[ ]:


# ============================================================
# CELL 6 — LOAD THE ENERGISTICS UOM DICTIONARY AND MAPPINGS
# ============================================================

# Requires xml.etree.ElementTree as ET, pandas as pd, requests, and display from Cell 2.

UOM_REPOSITORY_URL = "https://github.com/younesbell-cell/My-Grad-project"  # Identify the GitHub repository containing the UOM reference files.
UOM_GITHUB_API = "https://api.github.com"  # Define the GitHub API base URL.
UOM_REPOSITORY = "younesbell-cell/My-Grad-project"  # Store the GitHub owner and repository name.
UOM_BRANCH = "main"  # Select the maintained repository branch.
UOM_FOLDER = "UOM dictionary V1.01/"  # Limit the import to the complete UOM dictionary folder.
UOM_REQUEST_TIMEOUT = (15, 180)  # Limit connection and download waiting times.

UOM_DICTIONARY = {}  # Store all loaded UOM dictionary results.
UOM_SOURCE = {}  # Store repository, branch, commit, folder, and source-file provenance.
UOM_DICTIONARY_METADATA = {}  # Store metadata from the main UOM dictionary root.
UOM_XML_ROOTS = {}  # Preserve every parsed XML root under its repository path.
UOM_SOURCE_FILES_DF = pd.DataFrame()  # Describe every XML source file loaded from GitHub.
UOM_DIMENSIONS_DF = pd.DataFrame()  # Hold the unit-dimension definitions.
UOM_QUANTITY_CLASSES_DF = pd.DataFrame()  # Hold the Quantity Class definitions.
UOM_QUANTITY_CLASS_MEMBERS_DF = pd.DataFrame()  # Hold one row for every Quantity Class and member-unit relationship.
UOM_UNITS_DF = pd.DataFrame()  # Hold the canonical Energistics unit definitions.
UOM_REFERENCES_DF = pd.DataFrame()  # Hold the conversion-reference definitions.
UOM_PREFIXES_DF = pd.DataFrame()  # Hold the supported unit prefixes.
UOM_CLASS_CODES_DF = pd.DataFrame()  # Hold the Quantity Class integer-code assignments.
UOM_UNIT_CODES_DF = pd.DataFrame()  # Hold the unit integer-code assignments.
UOM_CLASS_MAPPINGS_DF = pd.DataFrame()  # Hold all external-class-to-Energistics-class mappings.
UOM_UNIT_MAPPINGS_DF = pd.DataFrame()  # Hold all external-unit-to-Energistics-unit mappings.
UOM_STATS = {}  # Store structural counts describing the complete import.
UOM_UNIT_MAPPING_STATE_STATS = {}  # Store the number of unit mappings in each declared state.


def uom_xml_local_name(xml_tag):
    """Return an XML tag or attribute name without its namespace URI."""

    return xml_tag.rsplit("}", 1)[-1]  # Remove the namespace wrapper while preserving the local name.


def uom_xml_children(xml_element, child_name):
    """Return direct child elements having the requested local name."""

    if xml_element is None:
        return []  # Return no children when the requested parent is absent.

    return [child for child in xml_element if uom_xml_local_name(child.tag) == child_name]  # Select direct children without hardcoding a namespace prefix.


def uom_xml_child(xml_element, child_name):
    """Return the first direct child having the requested local name."""

    children = uom_xml_children(xml_element, child_name)  # Locate matching direct children.

    return children[0] if children else None  # Return the first occurrence or None.


def uom_xml_value(xml_element):
    """Return clean element text without changing its capitalization or internal content."""

    if xml_element is None or xml_element.text is None:
        return ""  # Preserve an absent or empty XML value as an empty string.

    return xml_element.text.strip()  # Remove only surrounding XML formatting whitespace.


def uom_xml_text(xml_element, child_name):
    """Return clean text from a requested direct child."""

    return uom_xml_value(uom_xml_child(xml_element, child_name))  # Read the child value while preserving its original case.


def uom_xml_attribute(xml_element, attribute_name):
    """Return an attribute value by local name, including namespaced attributes."""

    if xml_element is None:
        return ""  # Return an empty value when the XML element is absent.

    for xml_name, xml_value in xml_element.attrib.items():
        if uom_xml_local_name(xml_name) == attribute_name:
            return xml_value  # Return the matching attribute unchanged.

    return ""  # Return an empty value when the requested attribute is absent.


def uom_mapping_source(source_file):
    """Derive a readable mapping-source label from its repository filename."""

    file_name = source_file.rsplit("/", 1)[-1]  # Isolate the XML filename from its repository path.
    source_label = file_name.split("_from_", 1)[-1].rsplit(".xml", 1)[0]  # Preserve the filename section identifying the source dictionary.

    return source_label.replace("_", " ")  # Convert filename separators into readable spaces.


def uom_build_dataframe(rows, columns, index_name):
    """Build a DataFrame with stable columns and a one-based record index."""

    dataframe = pd.DataFrame(rows, columns=columns)  # Preserve the intended columns even when no records are present.
    dataframe.index = pd.RangeIndex(start=1, stop=len(dataframe) + 1, name=index_name)  # Add a stable one-based notebook index.

    return dataframe  # Return the completed table without changing or deduplicating its records.


def parse_uom_xml_files(downloaded_files, source_commit):
    """Parse every UOM XML file into separate structural DataFrames."""

    xml_roots = {}  # Preserve every complete XML tree by repository path.
    source_file_rows = []  # Collect file-level metadata and provenance.
    dimension_rows = []  # Collect unit-dimension records.
    quantity_class_rows = []  # Collect Quantity Class records.
    quantity_class_member_rows = []  # Collect Quantity Class member-unit relationships.
    unit_rows = []  # Collect canonical unit definitions.
    reference_rows = []  # Collect conversion-reference definitions.
    prefix_rows = []  # Collect supported prefix definitions.
    class_code_rows = []  # Collect Quantity Class integer codes.
    unit_code_rows = []  # Collect unit integer codes.
    class_mapping_rows = []  # Collect external Quantity Class mappings.
    unit_mapping_rows = []  # Collect external unit mappings.

    root_categories = {
        "uomDictionary": "UOM Dictionary",
        "integerCodeSet": "Integer Code Dictionary",
        "classMappingSet": "Quantity Class Mapping",
        "unitMappingSet": "Unit Mapping"
    }  # Describe the supported XML root structures without changing their content.

    for downloaded_file in downloaded_files:
        source_file = downloaded_file["path"]  # Preserve the exact repository-relative path.
        xml_root = ET.fromstring(downloaded_file["content"])  # Parse the downloaded XML bytes.
        root_name = uom_xml_local_name(xml_root.tag)  # Read the XML root name without its namespace URI.

        if root_name not in root_categories:
            raise ValueError(f"Unsupported UOM XML root element in {source_file}: {root_name}")

        xml_roots[source_file] = xml_root  # Keep the complete parsed XML tree for later cells.
        source_file_rows.append({
            "Source File": source_file,
            "Source Commit": source_commit,
            "Blob SHA": downloaded_file["sha"],
            "File Size Bytes": downloaded_file["size"],
            "File Category": root_categories[root_name],
            "Root Element": root_name,
            "Version": uom_xml_attribute(xml_root, "version"),
            "Schema Location": uom_xml_attribute(xml_root, "schemaLocation"),
            "Title": uom_xml_text(xml_root, "title"),
            "Originator": uom_xml_text(xml_root, "originator"),
            "Description": uom_xml_text(xml_root, "description")
        })  # Preserve file metadata, paths, and immutable GitHub provenance.

        if root_name == "uomDictionary":
            dimension_set = uom_xml_child(xml_root, "unitDimensionSet")  # Locate the unit-dimension set.
            quantity_class_set = uom_xml_child(xml_root, "quantityClassSet")  # Locate the Quantity Class set.
            unit_set = uom_xml_child(xml_root, "unitSet")  # Locate the canonical unit set.
            reference_set = uom_xml_child(xml_root, "referenceSet")  # Locate the conversion-reference set.
            prefix_set = uom_xml_child(xml_root, "prefixSet")  # Locate the prefix set.

            for dimension_order, dimension_element in enumerate(uom_xml_children(dimension_set, "unitDimension"), start=1):
                dimension_rows.append({
                    "Source File": source_file,
                    "Source Commit": source_commit,
                    "Set Version": uom_xml_attribute(dimension_set, "version"),
                    "Dimension Record Order": dimension_order,
                    "Name": uom_xml_text(dimension_element, "name"),
                    "Dimension": uom_xml_text(dimension_element, "dimension"),
                    "Base For Conversion": uom_xml_text(dimension_element, "baseForConversion"),
                    "Canonical Unit": uom_xml_text(dimension_element, "canonicalUnit"),
                    "Description": uom_xml_text(dimension_element, "description")
                })  # Preserve every field belonging to this unit dimension.

            for class_order, class_element in enumerate(uom_xml_children(quantity_class_set, "quantityClass"), start=1):
                class_name = uom_xml_text(class_element, "name")  # Read the Quantity Class name exactly as supplied.
                class_dimension = uom_xml_text(class_element, "dimension")  # Read the declared dimensional expression.
                class_base = uom_xml_text(class_element, "baseForConversion")  # Read the declared conversion base.
                member_elements = uom_xml_children(class_element, "memberUnit")  # Preserve every member-unit relationship.

                quantity_class_rows.append({
                    "Source File": source_file,
                    "Source Commit": source_commit,
                    "Set Version": uom_xml_attribute(quantity_class_set, "version"),
                    "Quantity Class Record Order": class_order,
                    "Name": class_name,
                    "Dimension": class_dimension,
                    "Base For Conversion": class_base,
                    "Alternative Base": uom_xml_text(class_element, "alternativeBase"),
                    "Description": uom_xml_text(class_element, "description"),
                    "Member Unit Count": len(member_elements)
                })  # Preserve every Quantity Class without calculating relationships.

                for member_order, member_element in enumerate(member_elements, start=1):
                    quantity_class_member_rows.append({
                        "Source File": source_file,
                        "Source Commit": source_commit,
                        "Quantity Class Record Order": class_order,
                        "Quantity Class": class_name,
                        "Quantity Class Dimension": class_dimension,
                        "Base For Conversion": class_base,
                        "Member Unit Order": member_order,
                        "Member Unit": uom_xml_value(member_element)
                    })  # Preserve each explicit Quantity Class and member-unit link as its own row.

            for unit_order, unit_element in enumerate(uom_xml_children(unit_set, "unit"), start=1):
                unit_rows.append({
                    "Source File": source_file,
                    "Source Commit": source_commit,
                    "Set Version": uom_xml_attribute(unit_set, "version"),
                    "Unit Record Order": unit_order,
                    "Symbol": uom_xml_text(unit_element, "symbol"),
                    "Name": uom_xml_text(unit_element, "name"),
                    "Dimension": uom_xml_text(unit_element, "dimension"),
                    "Is SI": uom_xml_text(unit_element, "isSI"),
                    "Category": uom_xml_text(unit_element, "category"),
                    "Base Unit": uom_xml_text(unit_element, "baseUnit"),
                    "Conversion Reference": uom_xml_text(unit_element, "conversionRef"),
                    "Is Exact": uom_xml_text(unit_element, "isExact"),
                    "A": uom_xml_text(unit_element, "A"),
                    "B": uom_xml_text(unit_element, "B"),
                    "C": uom_xml_text(unit_element, "C"),
                    "D": uom_xml_text(unit_element, "D"),
                    "Underlying Definition": uom_xml_text(unit_element, "underlyingDef"),
                    "Description": uom_xml_text(unit_element, "description"),
                    "Is Base": uom_xml_text(unit_element, "isBase")
                })  # Preserve all supplied unit-definition fields without evaluating conversions.

            for reference_order, reference_element in enumerate(uom_xml_children(reference_set, "reference"), start=1):
                reference_rows.append({
                    "Source File": source_file,
                    "Source Commit": source_commit,
                    "Set Version": uom_xml_attribute(reference_set, "version"),
                    "Reference Record Order": reference_order,
                    "ID": uom_xml_text(reference_element, "ID"),
                    "Description": uom_xml_text(reference_element, "description")
                })  # Preserve each conversion-reference definition.

            for prefix_order, prefix_element in enumerate(uom_xml_children(prefix_set, "prefix"), start=1):
                prefix_rows.append({
                    "Source File": source_file,
                    "Source Commit": source_commit,
                    "Set Version": uom_xml_attribute(prefix_set, "version"),
                    "Prefix Record Order": prefix_order,
                    "Symbol": uom_xml_text(prefix_element, "symbol"),
                    "Name": uom_xml_text(prefix_element, "name"),
                    "Multiplier": uom_xml_text(prefix_element, "multiplier"),
                    "Common Name": uom_xml_text(prefix_element, "commonName")
                })  # Preserve every prefix and its supplied multiplier.

        elif root_name == "integerCodeSet":
            class_code_set = uom_xml_child(xml_root, "classCodeSet")  # Locate Quantity Class integer codes.
            unit_code_set = uom_xml_child(xml_root, "unitCodeSet")  # Locate unit integer codes.

            for class_code_order, class_code_element in enumerate(uom_xml_children(class_code_set, "classCode"), start=1):
                class_unit_element = uom_xml_child(class_code_element, "unit")  # Locate the class conversion-base unit reference.
                class_code_rows.append({
                    "Source File": source_file,
                    "Source Commit": source_commit,
                    "Class Code Record Order": class_code_order,
                    "Term": uom_xml_text(class_code_element, "term"),
                    "Code": uom_xml_text(class_code_element, "code"),
                    "Unit Symbol": uom_xml_value(class_unit_element),
                    "Unit Code": uom_xml_attribute(class_unit_element, "code"),
                    "Deprecated": uom_xml_text(class_code_element, "deprecated")
                })  # Preserve every Quantity Class code and its linked unit code.

            for unit_code_order, unit_code_element in enumerate(uom_xml_children(unit_code_set, "unitCode"), start=1):
                unit_code_rows.append({
                    "Source File": source_file,
                    "Source Commit": source_commit,
                    "Unit Code Record Order": unit_code_order,
                    "Term": uom_xml_text(unit_code_element, "term"),
                    "Code": uom_xml_text(unit_code_element, "code"),
                    "Deprecated": uom_xml_text(unit_code_element, "deprecated")
                })  # Preserve every unit integer code and any deprecation statement.

        elif root_name == "classMappingSet":
            mapping_source = uom_mapping_source(source_file)  # Record the external dictionary represented by this mapping file.

            for mapping_order, mapping_element in enumerate(uom_xml_children(xml_root, "classMap"), start=1):
                class_mapping_rows.append({
                    "Source File": source_file,
                    "Source Commit": source_commit,
                    "Mapping Source": mapping_source,
                    "Mapping Record Order": mapping_order,
                    "Maps From": uom_xml_text(mapping_element, "mapsFrom"),
                    "Maps To": uom_xml_text(mapping_element, "mapsTo")
                })  # Preserve every supplied class mapping, including empty destinations.

        elif root_name == "unitMappingSet":
            mapping_source = uom_mapping_source(source_file)  # Record the external dictionary represented by this mapping file.

            for mapping_order, mapping_element in enumerate(uom_xml_children(xml_root, "unitMap"), start=1):
                unit_mapping_rows.append({
                    "Source File": source_file,
                    "Source Commit": source_commit,
                    "Mapping Source": mapping_source,
                    "Mapping Record Order": mapping_order,
                    "Maps From": uom_xml_text(mapping_element, "mapsFrom"),
                    "Maps To": uom_xml_text(mapping_element, "mapsTo"),
                    "State": uom_xml_text(mapping_element, "state"),
                    "Note": uom_xml_text(mapping_element, "note")
                })  # Preserve every supplied unit mapping, state, note, and empty destination.

    source_files_df = uom_build_dataframe(source_file_rows, ["Source File","Source Commit","Blob SHA","File Size Bytes","File Category","Root Element","Version","Schema Location","Title","Originator","Description"], "source_file_order")
    dimensions_df = uom_build_dataframe(dimension_rows, ["Source File","Source Commit","Set Version","Dimension Record Order","Name","Dimension","Base For Conversion","Canonical Unit","Description"], "dimension_order")
    quantity_classes_df = uom_build_dataframe(quantity_class_rows, ["Source File","Source Commit","Set Version","Quantity Class Record Order","Name","Dimension","Base For Conversion","Alternative Base","Description","Member Unit Count"], "quantity_class_order")
    quantity_class_members_df = uom_build_dataframe(quantity_class_member_rows, ["Source File","Source Commit","Quantity Class Record Order","Quantity Class","Quantity Class Dimension","Base For Conversion","Member Unit Order","Member Unit"], "membership_order")
    units_df = uom_build_dataframe(unit_rows, ["Source File","Source Commit","Set Version","Unit Record Order","Symbol","Name","Dimension","Is SI","Category","Base Unit","Conversion Reference","Is Exact","A","B","C","D","Underlying Definition","Description","Is Base"], "unit_order")
    references_df = uom_build_dataframe(reference_rows, ["Source File","Source Commit","Set Version","Reference Record Order","ID","Description"], "reference_order")
    prefixes_df = uom_build_dataframe(prefix_rows, ["Source File","Source Commit","Set Version","Prefix Record Order","Symbol","Name","Multiplier","Common Name"], "prefix_order")
    class_codes_df = uom_build_dataframe(class_code_rows, ["Source File","Source Commit","Class Code Record Order","Term","Code","Unit Symbol","Unit Code","Deprecated"], "class_code_order")
    unit_codes_df = uom_build_dataframe(unit_code_rows, ["Source File","Source Commit","Unit Code Record Order","Term","Code","Deprecated"], "unit_code_order")
    class_mappings_df = uom_build_dataframe(class_mapping_rows, ["Source File","Source Commit","Mapping Source","Mapping Record Order","Maps From","Maps To"], "class_mapping_order")
    unit_mappings_df = uom_build_dataframe(unit_mapping_rows, ["Source File","Source Commit","Mapping Source","Mapping Record Order","Maps From","Maps To","State","Note"], "unit_mapping_order")

    dictionary_metadata_rows = source_files_df.loc[source_files_df["Root Element"] == "uomDictionary"]  # Locate metadata belonging to the main UOM dictionary.

    if len(dictionary_metadata_rows) != 1:
        raise ValueError(f"Expected one main UOM dictionary but detected {len(dictionary_metadata_rows)}.")

    dictionary_metadata = dictionary_metadata_rows.iloc[0].to_dict()  # Store the main dictionary metadata as a convenient dictionary.
    class_mappings_with_targets = int(class_mappings_df["Maps To"].ne("").sum())  # Count supplied class mappings having a destination.
    unit_mappings_with_targets = int(unit_mappings_df["Maps To"].ne("").sum())  # Count supplied unit mappings having a destination.
    mapping_state_counts = unit_mappings_df["State"].value_counts().sort_index()  # Count each mapping state without changing mapping rows.
    unit_mapping_state_stats = {state: int(count) for state, count in mapping_state_counts.items()}  # Store the state counts in a compact dictionary.

    stats = {
        "XML files imported": len(source_files_df),
        "UOM dictionary files": int((source_files_df["Root Element"] == "uomDictionary").sum()),
        "Integer-code files": int((source_files_df["Root Element"] == "integerCodeSet").sum()),
        "Quantity Class mapping files": int((source_files_df["Root Element"] == "classMappingSet").sum()),
        "Unit mapping files": int((source_files_df["Root Element"] == "unitMappingSet").sum()),
        "Unit dimensions": len(dimensions_df),
        "Quantity Classes": len(quantity_classes_df),
        "Quantity Class member rows": len(quantity_class_members_df),
        "Unit definitions": len(units_df),
        "Reference definitions": len(references_df),
        "Prefix definitions": len(prefixes_df),
        "Quantity Class integer codes": len(class_codes_df),
        "Unit integer codes": len(unit_codes_df),
        "Quantity Class mapping rows": len(class_mappings_df),
        "Quantity Class mappings with targets": class_mappings_with_targets,
        "Quantity Class mappings without targets": len(class_mappings_df) - class_mappings_with_targets,
        "Unit mapping rows": len(unit_mappings_df),
        "Unit mappings with targets": unit_mappings_with_targets,
        "Unit mappings without targets": len(unit_mappings_df) - unit_mappings_with_targets
    }  # Calculate only structural import statistics without performing validation or matching.

    return {
        "source": {
            "repository": UOM_REPOSITORY_URL,
            "branch": UOM_BRANCH,
            "commit": source_commit,
            "folder": UOM_FOLDER,
            "files": source_files_df["Source File"].tolist()
        },
        "metadata": dictionary_metadata,
        "xml_roots": xml_roots,
        "source_files": source_files_df,
        "dimensions": dimensions_df,
        "quantity_classes": quantity_classes_df,
        "quantity_class_members": quantity_class_members_df,
        "units": units_df,
        "references": references_df,
        "prefixes": prefixes_df,
        "class_codes": class_codes_df,
        "unit_codes": unit_codes_df,
        "class_mappings": class_mappings_df,
        "unit_mappings": unit_mappings_df,
        "stats": stats,
        "unit_mapping_state_stats": unit_mapping_state_stats
    }  # Return all XML roots, source paths, tables, and loading statistics.


def load_uom_dictionary():
    """Resolve the GitHub commit, download every UOM XML file, and parse all structures."""

    with requests.Session() as session:
        session.headers.update({"Accept": "application/vnd.github+json", "User-Agent": "PWLS-thesis-workflow/1.0"})  # Identify the notebook workflow and request GitHub JSON responses.
        commit_url = f"{UOM_GITHUB_API}/repos/{UOM_REPOSITORY}/commits/{UOM_BRANCH}"  # Build the branch-commit endpoint.
        commit_response = session.get(commit_url, timeout=UOM_REQUEST_TIMEOUT)  # Resolve the current commit of the maintained branch.
        commit_response.raise_for_status()  # Stop when the branch information cannot be retrieved.
        source_commit = commit_response.json().get("sha", "")  # Preserve the immutable commit used for this import.

        if not source_commit:
            raise ValueError("The UOM dictionary source commit could not be determined.")

        tree_url = f"{UOM_GITHUB_API}/repos/{UOM_REPOSITORY}/git/trees/{source_commit}?recursive=1"  # Build the commit-pinned repository-tree endpoint.
        tree_response = session.get(tree_url, timeout=UOM_REQUEST_TIMEOUT)  # Retrieve all paths under the selected repository commit.
        tree_response.raise_for_status()  # Stop when the repository tree cannot be retrieved.
        tree_data = tree_response.json()  # Read the GitHub tree response.

        if tree_data.get("truncated"):
            raise ValueError("The GitHub repository tree response was truncated.")

        xml_files = sorted([
            item for item in tree_data.get("tree", [])
            if item.get("type") == "blob" and item.get("path", "").startswith(UOM_FOLDER) and item.get("path", "").lower().endswith(".xml")
        ], key=lambda item: item["path"])  # Select every XML file inside the UOM folder and its mapping subfolders.

        if not xml_files:
            raise FileNotFoundError(f"No XML files were found under repository folder: {UOM_FOLDER}")

        downloaded_files = []  # Collect the bytes and GitHub provenance for every selected XML file.

        for xml_file in xml_files:
            encoded_path = requests.utils.quote(xml_file["path"], safe="/")  # Encode spaces while preserving repository path separators.
            raw_file_url = f"https://raw.githubusercontent.com/{UOM_REPOSITORY}/{source_commit}/{encoded_path}"  # Build an immutable raw-file URL.
            xml_response = session.get(raw_file_url, timeout=UOM_REQUEST_TIMEOUT)  # Download this XML file.
            xml_response.raise_for_status()  # Stop when any required XML file cannot be downloaded.
            downloaded_files.append({
                "path": xml_file["path"],
                "sha": xml_file.get("sha", ""),
                "size": xml_file.get("size", len(xml_response.content)),
                "content": xml_response.content
            })  # Preserve the exact path, blob identifier, size, and XML bytes.

    return parse_uom_xml_files(downloaded_files, source_commit)  # Parse all downloaded dictionary and mapping files.


try:
    UOM_DICTIONARY = load_uom_dictionary()  # Load every XML file in the UOM dictionary folder.
    UOM_SOURCE = UOM_DICTIONARY["source"]  # Store repository and source-file provenance.
    UOM_DICTIONARY_METADATA = UOM_DICTIONARY["metadata"]  # Store metadata from the main dictionary.
    UOM_XML_ROOTS = UOM_DICTIONARY["xml_roots"]  # Preserve every complete XML root.
    UOM_SOURCE_FILES_DF = UOM_DICTIONARY["source_files"]  # Store file-level metadata and paths.
    UOM_DIMENSIONS_DF = UOM_DICTIONARY["dimensions"]  # Store unit dimensions.
    UOM_QUANTITY_CLASSES_DF = UOM_DICTIONARY["quantity_classes"]  # Store Quantity Class definitions.
    UOM_QUANTITY_CLASS_MEMBERS_DF = UOM_DICTIONARY["quantity_class_members"]  # Store Quantity Class memberships.
    UOM_UNITS_DF = UOM_DICTIONARY["units"]  # Store canonical unit definitions.
    UOM_REFERENCES_DF = UOM_DICTIONARY["references"]  # Store conversion-reference definitions.
    UOM_PREFIXES_DF = UOM_DICTIONARY["prefixes"]  # Store supported prefixes.
    UOM_CLASS_CODES_DF = UOM_DICTIONARY["class_codes"]  # Store Quantity Class integer codes.
    UOM_UNIT_CODES_DF = UOM_DICTIONARY["unit_codes"]  # Store unit integer codes.
    UOM_CLASS_MAPPINGS_DF = UOM_DICTIONARY["class_mappings"]  # Store all Quantity Class mappings.
    UOM_UNIT_MAPPINGS_DF = UOM_DICTIONARY["unit_mappings"]  # Store all unit mappings.
    UOM_STATS = UOM_DICTIONARY["stats"]  # Store structural import statistics.
    UOM_UNIT_MAPPING_STATE_STATS = UOM_DICTIONARY["unit_mapping_state_stats"]  # Store unit-mapping state counts.

    print("Cell 6 loaded successfully.")  # Confirm that all UOM reference files are available.
    print(f"Source commit: {UOM_SOURCE['commit']}")  # Display the exact repository version used.
    print(f"Source folder: {UOM_SOURCE['folder']}")  # Display the repository-relative source folder.
    print(f"Dictionary version: {UOM_DICTIONARY_METADATA['Version']}")  # Display the main UOM dictionary version.

    for statistic_name, statistic_value in UOM_STATS.items():
        print(f"{statistic_name}: {statistic_value:,}")  # Display each structural loading statistic.

    print("\nUnit mapping states:")  # Introduce the declared mapping-state statistics.

    for state_name, state_count in UOM_UNIT_MAPPING_STATE_STATS.items():
        print(f"{state_name}: {state_count:,}")  # Display each state exactly as declared by the mapping files.

    print("\nImported XML files:")  # Introduce the source-path preview.
    display(UOM_SOURCE_FILES_DF[["Source File","File Category","Root Element","File Size Bytes"]])  # Display every imported XML path and file type.

    print("\nFirst unit definitions:")  # Introduce a short canonical-unit preview.
    display(UOM_UNITS_DF.head(10))  # Display the first ten unit definitions.

    print("\nFirst unit mapping records:")  # Introduce a short unit-mapping preview.
    display(UOM_UNIT_MAPPINGS_DF.head(10))  # Display the first ten external unit mappings.

except Exception as error:
    UOM_DICTIONARY = {}  # Clear the combined result after a loading or parsing failure.
    UOM_SOURCE = {}  # Clear repository provenance.
    UOM_DICTIONARY_METADATA = {}  # Clear dictionary metadata.
    UOM_XML_ROOTS = {}  # Clear parsed XML roots.
    UOM_SOURCE_FILES_DF = pd.DataFrame()  # Clear source-file metadata.
    UOM_DIMENSIONS_DF = pd.DataFrame()  # Clear unit dimensions.
    UOM_QUANTITY_CLASSES_DF = pd.DataFrame()  # Clear Quantity Class definitions.
    UOM_QUANTITY_CLASS_MEMBERS_DF = pd.DataFrame()  # Clear Quantity Class memberships.
    UOM_UNITS_DF = pd.DataFrame()  # Clear canonical unit definitions.
    UOM_REFERENCES_DF = pd.DataFrame()  # Clear conversion references.
    UOM_PREFIXES_DF = pd.DataFrame()  # Clear prefixes.
    UOM_CLASS_CODES_DF = pd.DataFrame()  # Clear Quantity Class integer codes.
    UOM_UNIT_CODES_DF = pd.DataFrame()  # Clear unit integer codes.
    UOM_CLASS_MAPPINGS_DF = pd.DataFrame()  # Clear Quantity Class mappings.
    UOM_UNIT_MAPPINGS_DF = pd.DataFrame()  # Clear unit mappings.
    UOM_STATS = {}  # Clear loading statistics.
    UOM_UNIT_MAPPING_STATE_STATS = {}  # Clear mapping-state statistics.

    print("Cell 6 failed.")  # Report a concise cell-level failure.
    print(error)  # Display the specific GitHub, download, or XML parsing error.


# In[ ]:


# ============================================================
# CELL 7 — CREATE OR LOAD ALIAS WORKBOOK
# ============================================================

_PREDEFINED_ALIASES_WORKBOOK_PATH = globals().get("ALIASES_WORKBOOK_PATH")

EXPERT_REVIEW_WORKBOOK_FILENAME = "Expert review workbook.xlsx"
ALIASES_HEADER_FILL = "FFFF773D"
ALIASES_DUPLICATE_FILL = "FFFFC7CE"
ALIASES_DUPLICATE_FONT = "FF000000"

ALIASES_WORKBOOK_SCHEMA = {
    "Mnemonic Aliases": [
        "Curve Mnemonic",
        "Property Kind",
        "Quantity Class",
        "LIS Curve Mnemonic",
        "Curve Description",
        "Note",
        "Reviewer Name",
        "Date (dd-mm-yyyy)",
    ],
    "Unit Aliases": [
        "Maps From",
        "Maps To",
        "Note",
        "Reviewer Name",
        "Date (dd-mm-yyyy)",
    ],
    "Company Aliases": [
        "Common Company Name",
        "Canonical Company Name",
        "Note",
        "Reviewer Name",
        "Date (dd-mm-yyyy)",
    ],
}

MNEMONIC_STARTUP_ALIASES_DATA = [
    {"Curve Mnemonic": "VERS", "Property Kind": "digital file format version", "Quantity Class": "unitless", "LIS Curve Mnemonic": "", "Curve Description": "LAS version", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
    {"Curve Mnemonic": "WRAP", "Property Kind": "flag", "Quantity Class": "unitless", "LIS Curve Mnemonic": "", "Curve Description": "Wrapped data flag", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
    {"Curve Mnemonic": "DLM", "Property Kind": "", "Quantity Class": "", "LIS Curve Mnemonic": "", "Curve Description": "Delimiter", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
    {"Curve Mnemonic": "WELL", "Property Kind": "well name", "Quantity Class": "unitless", "LIS Curve Mnemonic": "", "Curve Description": "Well name", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
    {"Curve Mnemonic": "COMP", "Property Kind": "organization name", "Quantity Class": "unitless", "LIS Curve Mnemonic": "", "Curve Description": "Company", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
    {"Curve Mnemonic": "FLD", "Property Kind": "field name", "Quantity Class": "unitless", "LIS Curve Mnemonic": "", "Curve Description": "Field", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
    {"Curve Mnemonic": "LOC", "Property Kind": "geographic location", "Quantity Class": "unitless", "LIS Curve Mnemonic": "", "Curve Description": "Location", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
    {"Curve Mnemonic": "PROV", "Property Kind": "geopolitical region code", "Quantity Class": "unitless", "LIS Curve Mnemonic": "", "Curve Description": "Province/state", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
    {"Curve Mnemonic": "CNTY", "Property Kind": "county name", "Quantity Class": "unitless", "LIS Curve Mnemonic": "", "Curve Description": "County", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
    {"Curve Mnemonic": "STAT", "Property Kind": "state name", "Quantity Class": "unitless", "LIS Curve Mnemonic": "", "Curve Description": "State", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
    {"Curve Mnemonic": "CTRY", "Property Kind": "country name", "Quantity Class": "unitless", "LIS Curve Mnemonic": "", "Curve Description": "Country", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
    {"Curve Mnemonic": "SRVC", "Property Kind": "organization name", "Quantity Class": "unitless", "LIS Curve Mnemonic": "", "Curve Description": "Service company", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
    {"Curve Mnemonic": "UWI", "Property Kind": "uwi", "Quantity Class": "unitless", "LIS Curve Mnemonic": "", "Curve Description": "Unique well identifier", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
    {"Curve Mnemonic": "API", "Property Kind": "api well number", "Quantity Class": "unitless", "LIS Curve Mnemonic": "", "Curve Description": "API number or API unit depending context", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
    {"Curve Mnemonic": "STRT", "Property Kind": "start depth", "Quantity Class": "length", "LIS Curve Mnemonic": "", "Curve Description": "Start depth", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
    {"Curve Mnemonic": "STOP", "Property Kind": "stop depth", "Quantity Class": "length", "LIS Curve Mnemonic": "", "Curve Description": "Stop depth", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
    {"Curve Mnemonic": "STEP", "Property Kind": "depth increment", "Quantity Class": "length", "LIS Curve Mnemonic": "", "Curve Description": "Step interval", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
    {"Curve Mnemonic": "NULL", "Property Kind": "", "Quantity Class": "", "LIS Curve Mnemonic": "", "Curve Description": "Null value", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
    {"Curve Mnemonic": "RUN", "Property Kind": "run number", "Quantity Class": "unitless", "LIS Curve Mnemonic": "", "Curve Description": "Run number/name", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
    {"Curve Mnemonic": "DATE", "Property Kind": "date", "Quantity Class": "time", "LIS Curve Mnemonic": "", "Curve Description": "Acquisition date", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
    {"Curve Mnemonic": "DEPT", "Property Kind": "measured depth", "Quantity Class": "length", "LIS Curve Mnemonic": "", "Curve Description": "Depth curve", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
    {"Curve Mnemonic": "DEPTH", "Property Kind": "measured depth", "Quantity Class": "length", "LIS Curve Mnemonic": "", "Curve Description": "Depth curve", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
    {"Curve Mnemonic": "TIME", "Property Kind": "time", "Quantity Class": "time", "LIS Curve Mnemonic": "", "Curve Description": "Time curve", "Note": "", "Reviewer Name": "", "Date (dd-mm-yyyy)": ""},
]

ALIASES_WORKBOOK_INFO = {}
ALIASES_WORKBOOK_STATS = {}
ALIASES_SCHEMA_ISSUES_DF = pd.DataFrame(columns=["Sheet", "Issue", "Expected", "Actual"])
MNEMONIC_ALIASES_DF = pd.DataFrame(columns=ALIASES_WORKBOOK_SCHEMA["Mnemonic Aliases"])
UNIT_ALIASES_DF = pd.DataFrame(columns=ALIASES_WORKBOOK_SCHEMA["Unit Aliases"])
COMPANY_NAME_ALIASES_DF = pd.DataFrame(columns=ALIASES_WORKBOOK_SCHEMA["Company Aliases"])


def clean_user_path(path_value):
    path_text = str(path_value).strip().strip('"').strip("'")

    if not path_text:
        raise ValueError("The expert-review-workbook path is empty.")

    path_text = os.path.expandvars(os.path.expanduser(path_text))
    return Path(path_text)


def choose_expert_review_folder():
    selected_folder = ""
    dialog_error = None
    root = None

    if TKINTER_AVAILABLE:
        try:
            root = Tk()
            root.withdraw()
            root.attributes("-topmost", True)

            selected_folder = filedialog.askdirectory(
                title="Choose the folder for Expert review workbook.xlsx"
            )

        except Exception as exc:
            dialog_error = exc

        finally:
            if root is not None:
                root.destroy()

    if not selected_folder:
        if dialog_error is not None:
            print(f"Windows folder selection was unavailable: {dialog_error}")

        selected_folder = input(
            "Enter the existing folder in which Expert review workbook.xlsx "
            "should be created or found: "
        ).strip().strip('"').strip("'")

    if not selected_folder:
        raise RuntimeError("No expert-review-workbook folder was selected.")

    return clean_user_path(selected_folder)


def resolve_expert_review_workbook_path(predefined_path=None):
    if predefined_path is None or not str(predefined_path).strip():
        folder_path = choose_expert_review_folder()
        location_source = "Folder selected during Alias Workbook Preparation"
        workbook_path = folder_path / EXPERT_REVIEW_WORKBOOK_FILENAME

    else:
        supplied_path = clean_user_path(predefined_path)
        location_source = "Previously defined ALIASES_WORKBOOK_PATH"

        if supplied_path.suffix.lower() == ".xlsx":
            if supplied_path.name != EXPERT_REVIEW_WORKBOOK_FILENAME:
                raise ValueError(
                    "ALIASES_WORKBOOK_PATH must use the fixed filename "
                    f"'{EXPERT_REVIEW_WORKBOOK_FILENAME}'."
                )

            folder_path = supplied_path.parent
            workbook_path = supplied_path

        else:
            folder_path = supplied_path
            workbook_path = folder_path / EXPERT_REVIEW_WORKBOOK_FILENAME

    if not folder_path.exists():
        raise FileNotFoundError(
            f"The selected folder does not exist: {folder_path}"
        )

    if not folder_path.is_dir():
        raise NotADirectoryError(
            f"The selected location is not a folder: {folder_path}"
        )

    return workbook_path.resolve(), location_source


def autosize_alias_columns(worksheet, headers):
    for column_number, header in enumerate(headers, start=1):
        values = [header]

        values.extend(
            worksheet.cell(row=row_number, column=column_number).value
            for row_number in range(2, worksheet.max_row + 1)
        )

        visible_lengths = [
            len(str(value))
            for value in values
            if value not in (None, "")
        ]

        column_width = max(visible_lengths) + 2
        column_letter = get_column_letter(column_number)
        worksheet.column_dimensions[column_letter].width = column_width


def format_new_alias_sheet(worksheet, headers):
    header_fill = PatternFill(
        fill_type="solid",
        fgColor=ALIASES_HEADER_FILL,
    )

    for cell in worksheet[1]:
        cell.font = Font(
            bold=True,
            color="FF000000",
        )
        cell.fill = header_fill
        cell.alignment = Alignment(
            wrap_text=False,
        )

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=len(headers),
    ):
        for cell in row:
            cell.alignment = Alignment(
                wrap_text=False,
            )

    date_column = headers.index("Date (dd-mm-yyyy)") + 1

    for row_number in range(2, worksheet.max_row + 1):
        worksheet.cell(
            row=row_number,
            column=date_column,
        ).number_format = "dd-mm-yyyy"

    duplicate_style = DifferentialStyle(
        font=Font(
            color=ALIASES_DUPLICATE_FONT,
        ),
        fill=PatternFill(
            fill_type="solid",
            fgColor=ALIASES_DUPLICATE_FILL,
        ),
    )

    worksheet.conditional_formatting.add(
        "A2:A1048576",
        Rule(
            type="duplicateValues",
            dxf=duplicate_style,
        ),
    )

    autosize_alias_columns(
        worksheet,
        headers,
    )


def create_expert_review_workbook(workbook_path):
    if workbook_path.exists():
        raise FileExistsError(
            "The workbook appeared before creation and will not be overwritten: "
            f"{workbook_path}"
        )

    workbook = Workbook()

    mnemonic_sheet = workbook.active
    mnemonic_sheet.title = "Mnemonic Aliases"

    workbook.create_sheet("Unit Aliases")
    workbook.create_sheet("Company Aliases")

    for sheet_name, headers in ALIASES_WORKBOOK_SCHEMA.items():
        worksheet = workbook[sheet_name]
        worksheet.append(headers)

        if sheet_name == "Mnemonic Aliases":
            for record in MNEMONIC_STARTUP_ALIASES_DATA:
                worksheet.append(
                    [record[header] for header in headers]
                )

        format_new_alias_sheet(
            worksheet,
            headers,
        )

    workbook.save(workbook_path)
    workbook.close()


def validate_existing_alias_workbook(workbook_path):
    issues = []

    try:
        workbook = load_workbook(
            workbook_path,
            read_only=True,
            data_only=False,
        )

    except Exception as exc:
        raise RuntimeError(
            "The existing Expert review workbook could not be opened. "
            "It was not overwritten."
        ) from exc

    try:
        for sheet_name, expected_headers in ALIASES_WORKBOOK_SCHEMA.items():
            if sheet_name not in workbook.sheetnames:
                issues.append(
                    {
                        "Sheet": sheet_name,
                        "Issue": "Required sheet missing",
                        "Expected": sheet_name,
                        "Actual": "",
                    }
                )
                continue

            worksheet = workbook[sheet_name]

            actual_headers = [
                worksheet.cell(
                    row=1,
                    column=column_number,
                ).value
                for column_number in range(
                    1,
                    worksheet.max_column + 1,
                )
            ]

            while actual_headers and actual_headers[-1] is None:
                actual_headers.pop()

            if actual_headers != expected_headers:
                issues.append(
                    {
                        "Sheet": sheet_name,
                        "Issue": "Header names or order changed",
                        "Expected": expected_headers,
                        "Actual": actual_headers,
                    }
                )

    finally:
        workbook.close()

    return issues


def load_alias_sheet(workbook_path, sheet_name):
    expected_headers = ALIASES_WORKBOOK_SCHEMA[sheet_name]

    dataframe = pd.read_excel(
        workbook_path,
        sheet_name=sheet_name,
        dtype=object,
        keep_default_na=False,
        engine="openpyxl",
    )

    return dataframe.loc[:, expected_headers].copy()


def get_filesystem_dates(workbook_path):
    file_statistics = workbook_path.stat()

    creation_timestamp = getattr(
        file_statistics,
        "st_birthtime",
        file_statistics.st_ctime,
    )

    date_format = "%d-%m-%Y %H:%M:%S"

    return {
        "Creation Date": datetime.fromtimestamp(
            creation_timestamp
        ).strftime(date_format),
        "Last Modified Date": datetime.fromtimestamp(
            file_statistics.st_mtime
        ).strftime(date_format),
    }


try:
    _resolved_workbook_path, _location_source = (
        resolve_expert_review_workbook_path(
            _PREDEFINED_ALIASES_WORKBOOK_PATH
        )
    )

    if _resolved_workbook_path.exists():
        if not _resolved_workbook_path.is_file():
            raise FileExistsError(
                "The target exists but is not a file: "
                f"{_resolved_workbook_path}"
            )

        _workbook_action = "Reused"

    else:
        create_expert_review_workbook(
            _resolved_workbook_path
        )

        _workbook_action = "Created"

    _schema_issues = validate_existing_alias_workbook(
        _resolved_workbook_path
    )

    ALIASES_SCHEMA_ISSUES_DF = pd.DataFrame(
        _schema_issues,
        columns=[
            "Sheet",
            "Issue",
            "Expected",
            "Actual",
        ],
    )

    if not ALIASES_SCHEMA_ISSUES_DF.empty:
        raise ValueError(
            "The existing workbook does not match the fixed "
            "expert-review-workbook schema. No existing content was changed. "
            "Review ALIASES_SCHEMA_ISSUES_DF."
        )

    MNEMONIC_ALIASES_DF = load_alias_sheet(
        _resolved_workbook_path,
        "Mnemonic Aliases",
    )

    UNIT_ALIASES_DF = load_alias_sheet(
        _resolved_workbook_path,
        "Unit Aliases",
    )

    COMPANY_NAME_ALIASES_DF = load_alias_sheet(
        _resolved_workbook_path,
        "Company Aliases",
    )

    ALIASES_WORKBOOK_PATH = _resolved_workbook_path
    _file_dates = get_filesystem_dates(
        ALIASES_WORKBOOK_PATH
    )

    ALIASES_WORKBOOK_INFO = {
        "Workbook Action": _workbook_action,
        "Location Source": _location_source,
        "Workbook Path": str(ALIASES_WORKBOOK_PATH),
        **_file_dates,
    }

    ALIASES_WORKBOOK_STATS = {
        "Mnemonic Alias Rows Loaded": len(
            MNEMONIC_ALIASES_DF
        ),
        "Unit Alias Rows Loaded": len(
            UNIT_ALIASES_DF
        ),
        "Company Alias Rows Loaded": len(
            COMPANY_NAME_ALIASES_DF
        ),
        "Schema Issues": len(
            ALIASES_SCHEMA_ISSUES_DF
        ),
    }

    print("Cell 7 completed successfully.")
    print(f"Location source: {_location_source}")
    print(f"Workbook action: {_workbook_action}")
    print(f"Workbook path: {ALIASES_WORKBOOK_PATH}")
    print(f"Creation date: {_file_dates['Creation Date']}")
    print(f"Last modified date: {_file_dates['Last Modified Date']}")

    if _workbook_action == "Reused":
        print(
            "The existing Expert review workbook was loaded "
            "without modification."
        )
    else:
        print(
            "No existing workbook was found. "
            "A new Expert review workbook was created."
        )

    for statistic_name, statistic_value in ALIASES_WORKBOOK_STATS.items():
        print(f"{statistic_name}: {statistic_value}")

except Exception as exc:
    ALIASES_WORKBOOK_PATH = None
    ALIASES_WORKBOOK_INFO = {}
    ALIASES_WORKBOOK_STATS = {}

    MNEMONIC_ALIASES_DF = pd.DataFrame(
        columns=ALIASES_WORKBOOK_SCHEMA["Mnemonic Aliases"]
    )

    UNIT_ALIASES_DF = pd.DataFrame(
        columns=ALIASES_WORKBOOK_SCHEMA["Unit Aliases"]
    )

    COMPANY_NAME_ALIASES_DF = pd.DataFrame(
        columns=ALIASES_WORKBOOK_SCHEMA["Company Aliases"]
    )

    print("Cell 7 failed.")
    print(f"Error: {exc}")

    raise


# In[ ]:


# ============================================================
# CELL 8 — INPUT PREPARATION
# ============================================================

# Requires pandas as pd and display from Cell 2.
# Uses METADATA_DF from Cell 3.
# Cell 7 separately creates or loads the Expert Review Workbook.

METADATA_INPUT_PREPROCESSING_COLUMNS = ["Service company", "mnemonic", "unit", "description"]

# Characters that may be removed only from the beginning or end of a value.
BOUNDARY_INVISIBLE_CHARACTERS = " \t\n\r\v\f\u00a0\u1680\u2000\u2001\u2002\u2003\u2004\u2005\u2006\u2007\u2008\u2009\u200a\u2028\u2029\u202f\u205f\u3000\u200b\u2060\ufeff"

# Convert supported Unicode superscript powers to the ASCII notation used by the UOM references.
UNIT_POWER_TRANSLATION = str.maketrans({"²": "2", "³": "3", "⁴": "4", "⁵": "5", "⁶": "6", "⁷": "7", "⁸": "8", "⁹": "9"})

# Convert a leading Unicode micro sign or Greek mu to the ASCII micro prefix.
UNIT_MICRO_PREFIX_VARIANTS = {"µ": "u", "μ": "u"}

# Standard descriptions used consistently in the audit log.
BOUNDARY_RULE = "Boundary whitespace/invisible-character trim"
POWER_RULE = "Unicode powers 2-9 converted to ASCII digits"
MICRO_RULE = "Leading Unicode micro prefix converted to ASCII u"

INPUT_PREPROCESSING_LOG_COLUMNS = [
    "Source Row Order",
    "Source Index",
    "Column",
    "Original Value",
    "Prepared Value",
    "Original Representation",
    "Prepared Representation",
    "Applied Rules",
]

# Initialize the output objects so the cell is safe to rerun.
METADATA_ORIGINAL_DF = pd.DataFrame()
METADATA_INPUT_PREPROCESSING_DF = pd.DataFrame()
METADATA_INPUT_PREPROCESSING_LOG_DF = pd.DataFrame(columns=INPUT_PREPROCESSING_LOG_COLUMNS)
METADATA_INPUT_PREPROCESSING_STATS = {}


def strip_boundary_invisible_characters(value):
    """Remove supported whitespace and invisible characters only from value boundaries."""

    if value is None or pd.isna(value):
        return ""

    # str.strip() removes only boundary characters and leaves internal characters unchanged.
    return str(value).strip(BOUNDARY_INVISIBLE_CHARACTERS)


def normalize_unit_notation(unit_value):
    """Prepare supported Unicode powers and a leading Unicode micro prefix."""

    # Convert superscript powers 2–9 wherever they occur in the unit expression.
    normalized_unit = unit_value.translate(UNIT_POWER_TRANSLATION)
    power_changed = normalized_unit != unit_value
    micro_changed = False

    # Convert µ or μ only when it is the first character and another character follows it.
    if len(normalized_unit) > 1 and normalized_unit[0] in UNIT_MICRO_PREFIX_VARIANTS:
        normalized_unit = UNIT_MICRO_PREFIX_VARIANTS[normalized_unit[0]] + normalized_unit[1:]
        micro_changed = True

    return normalized_unit, power_changed, micro_changed


def prepare_metadata_value(value, column_name):
    """Apply the permitted Input Data Preprocessing transformations to one metadata value."""

    # Preserve a stable textual representation while representing missing values as empty strings.
    original_value = "" if value is None or pd.isna(value) else str(value)

    # Apply boundary preparation to every metadata column.
    prepared_value = strip_boundary_invisible_characters(original_value)
    applied_rules = []

    if prepared_value != original_value:
        applied_rules.append(BOUNDARY_RULE)

    # Power and micro-prefix preparation applies only to the unit column.
    if column_name == "unit":
        prepared_value, power_changed, micro_changed = normalize_unit_notation(prepared_value)

        if power_changed:
            applied_rules.append(POWER_RULE)

        if micro_changed:
            applied_rules.append(MICRO_RULE)

    return original_value, prepared_value, applied_rules


def validate_input_preprocessing_input(metadata_df):
    """Verify that Cell 3 produced the expected metadata structure."""

    if not isinstance(metadata_df, pd.DataFrame):
        raise TypeError("METADATA_DF must be a pandas DataFrame.")

    actual_columns = list(metadata_df.columns)

    # Require the exact column names and order expected by the later workflow.
    if actual_columns != METADATA_INPUT_PREPROCESSING_COLUMNS:
        raise ValueError(
            f"Cell 8 requires these columns in this exact order: {METADATA_INPUT_PREPROCESSING_COLUMNS}. "
            f"Received: {actual_columns}"
        )


def apply_metadata_input_preprocessing(metadata_df):
    """Create a prepared metadata copy, audit log, and Input Data Preprocessing statistics."""

    validate_input_preprocessing_input(metadata_df)

    # Work on a deep copy so the original imported metadata remains unchanged.
    input_preprocessing_df = metadata_df.copy(deep=True)
    change_rows = []

    rule_counts = {
        "Boundary-trimmed cells": 0,
        "Power-notation cells converted": 0,
        "Micro-prefix cells converted": 0,
    }

    # Iterate by position so duplicate DataFrame indices are processed independently.
    for source_position in range(len(metadata_df)):
        source_row_order = source_position + 1
        source_index = metadata_df.index[source_position]

        for column_position, column_name in enumerate(METADATA_INPUT_PREPROCESSING_COLUMNS):
            source_value = metadata_df.iat[source_position, column_position]
            original_value, prepared_value, applied_rules = prepare_metadata_value(source_value, column_name)

            # Write only to the Input Data Preprocessing copy, never to METADATA_DF.
            input_preprocessing_df.iat[source_position, column_position] = prepared_value

            if not applied_rules:
                continue

            rule_counts["Boundary-trimmed cells"] += int(BOUNDARY_RULE in applied_rules)
            rule_counts["Power-notation cells converted"] += int(POWER_RULE in applied_rules)
            rule_counts["Micro-prefix cells converted"] += int(MICRO_RULE in applied_rules)

            # Record one audit row for every source cell that changed.
            change_rows.append({
                "Source Row Order": source_row_order,
                "Source Index": source_index,
                "Column": column_name,
                "Original Value": original_value,
                "Prepared Value": prepared_value,
                "Original Representation": repr(original_value),
                "Prepared Representation": repr(prepared_value),
                "Applied Rules": "; ".join(applied_rules),
            })

    # Preserve a stable schema even when no cells required transformation.
    input_preprocessing_log_df = pd.DataFrame(change_rows, columns=INPUT_PREPROCESSING_LOG_COLUMNS)
    input_preprocessing_log_df.index = pd.RangeIndex(start=1, stop=len(input_preprocessing_log_df) + 1, name="change_order")

    changed_row_count = input_preprocessing_log_df["Source Row Order"].nunique() if not input_preprocessing_log_df.empty else 0

    # A standalone u is preserved without interpretation for later expert review.
    standalone_u_count = int(input_preprocessing_df["unit"].eq("u").sum())

    input_preprocessing_stats = {
        "Input rows": len(metadata_df),
        "Rows containing changes": changed_row_count,
        "Cells changed": len(input_preprocessing_log_df),
        "Boundary-trimmed cells": rule_counts["Boundary-trimmed cells"],
        "Power-notation cells converted": rule_counts["Power-notation cells converted"],
        "Micro-prefix cells converted": rule_counts["Micro-prefix cells converted"],
        "Standalone u units preserved": standalone_u_count,
    }

    return input_preprocessing_df, input_preprocessing_log_df, input_preprocessing_stats


try:
    # Cell 8 depends only on the metadata imported by Cell 3.
    if "METADATA_DF" not in globals() or METADATA_DF is None:
        raise NameError("METADATA_DF is unavailable. Run Cell 3 before running Cell 8.")

    # Preserve an untouched snapshot of the imported metadata.
    METADATA_ORIGINAL_DF = METADATA_DF.copy(deep=True)

    # Create the prepared working DataFrame, transformation log, and statistics.
    METADATA_INPUT_PREPROCESSING_DF, METADATA_INPUT_PREPROCESSING_LOG_DF, METADATA_INPUT_PREPROCESSING_STATS = apply_metadata_input_preprocessing(METADATA_ORIGINAL_DF)

    # Confirm that Input Data Preprocessing did not modify the original imported DataFrame.
    pd.testing.assert_frame_equal(METADATA_DF, METADATA_ORIGINAL_DF, check_dtype=True, check_exact=True)

    print("Cell 8 input preparation completed successfully.")

    for statistic_name, statistic_value in METADATA_INPUT_PREPROCESSING_STATS.items():
        print(f"{statistic_name}: {statistic_value:,}")

    print("\nPrepared metadata preview:")
    display(METADATA_INPUT_PREPROCESSING_DF.head(10))

    if METADATA_INPUT_PREPROCESSING_LOG_DF.empty:
        print("\nNo metadata cells required Input Data Preprocessing transformations.")
    else:
        print("\nInput Data Preprocessing transformation log:")
        display(METADATA_INPUT_PREPROCESSING_LOG_DF.head(20))

except Exception as error:
    # Reset every Cell 8 output so later cells cannot use partial results.
    METADATA_ORIGINAL_DF = pd.DataFrame()
    METADATA_INPUT_PREPROCESSING_DF = pd.DataFrame()
    METADATA_INPUT_PREPROCESSING_LOG_DF = pd.DataFrame(columns=INPUT_PREPROCESSING_LOG_COLUMNS)
    METADATA_INPUT_PREPROCESSING_STATS = {}

    print("Cell 8 failed.")
    print(f"Error: {error}")


# In[ ]:


# ============================================================
# CELL 9 — COMPANY AND MNEMONIC VALIDATION
# ============================================================

# Requires pandas as pd and display from Cell 2.
# Uses the prepared metadata from Cell 8 and governed references from Cells 4, 5, and 7.

METADATA_VALIDATION_INPUT_COLUMNS = ["Service company", "mnemonic", "unit", "description"]

PWLS_COMPANY_REQUIRED_COLUMNS = ["Source File", "Source Record Order", "Company Code", "Organization Name"]
PWLS_CURVE_REQUIRED_COLUMNS = ["Source File", "Source Record Order", "Catalog Company", "Company Code", "Company Name", "Curve Mnemonic", "Property Kind", "Quantity Class", "LIS Curve Mnemonic", "Curve Description"]
PWLS_ARCHIVE_REQUIRED_COLUMNS = PWLS_CURVE_REQUIRED_COLUMNS + ["Catalog Type", "Is Archive"]
PWLS_PROPERTY_REQUIRED_COLUMNS = ["Source File", "Source Commit", "Property Record Order", "UUID", "Title", "Description", "Is Abstract", "Quantity Class", "Parent UUID", "Parent Title"]
PWLS_PROPERTY_ALIAS_REQUIRED_COLUMNS = ["Source File", "Source Commit", "Property Record Order", "Property UUID", "Property Title", "Alias Order", "Authority", "Identifier"]
MNEMONIC_ALIAS_REQUIRED_COLUMNS = ["Curve Mnemonic", "Property Kind", "Quantity Class", "LIS Curve Mnemonic", "Curve Description", "Note", "Reviewer Name", "Date (dd-mm-yyyy)"]
COMPANY_ALIAS_REQUIRED_COLUMNS = ["Common Company Name", "Canonical Company Name", "Note", "Reviewer Name", "Date (dd-mm-yyyy)"]

COMPANY_CANDIDATE_COLUMNS = [
    "Source Row Order",
    "Source Index",
    "Prepared Service Company",
    "Company Match Method",
    "Company Alias Workbook Row",
    "Alias Common Company Name",
    "Alias Canonical Company Name",
    "Alias Note",
    "Alias Reviewer Name",
    "Alias Review Date",
    "PWLS Company Reference Position",
    "Company Code",
    "Company Name",
    "Company Reference Source File",
    "Company Reference Source Record Order",
    "Candidate Resolution Status",
]

MNEMONIC_CANDIDATE_COLUMNS = [
    "Source Row Order",
    "Source Index",
    "Source Candidate Order",
    "Prepared Service Company",
    "Prepared Mnemonic",
    "Prepared Description",
    "Mnemonic Search Scope",
    "Mnemonic Match Method",
    "Mnemonic Validation Flag",
    "Candidate Source",
    "Candidate Source Position",
    "PWLS Curve Catalogue Status",
    "Candidate Company Code",
    "Candidate Company Name",
    "Catalog Company",
    "Reference Curve Mnemonic",
    "Candidate Property Kind",
    "Declared Quantity Class",
    "LIS Curve Mnemonic",
    "Reference Curve Description",
    "Property Kind Match Method",
    "Property Kind Resolution Status",
    "Property Kind Alias Position",
    "Property Kind Alias Identifier",
    "Property Kind Alias Authority",
    "Property Kind Alias Source File",
    "Property Kind Alias Source Commit",
    "Property Kind UUID",
    "Property Kind Title",
    "Property Kind Description",
    "Property Kind Is Abstract",
    "Dictionary Quantity Class",
    "Quantity Class Agreement",
    "Curve Source File",
    "Curve Source Record Order",
    "Curve Source Commit",
    "Property Kind Source File",
    "Property Kind Source Commit",
    "Mnemonic Alias Workbook Row",
    "Mnemonic Alias Note",
    "Mnemonic Alias Reviewer Name",
    "Mnemonic Alias Review Date",
]

MNEMONIC_VALIDATION_COLUMNS = [
    "Source Row Order",
    "Source Index",
    "Original Service Company",
    "Original Mnemonic",
    "Original Unit",
    "Original Description",
    "Prepared Service Company",
    "Prepared Mnemonic",
    "Prepared Unit",
    "Prepared Description",
    "Company Resolution Status",
    "Company Match Method",
    "Company Raw Match Count",
    "Company Candidate Count",
    "Resolved Company Code",
    "Resolved Company Name",
    "Company Candidate Codes",
    "Company Candidate Names",
    "Mnemonic Search Scope",
    "Mnemonic Match Method",
    "Mnemonic Validation Flag",
    "Mnemonic Raw Candidate Count",
    "Property Candidate Row Count",
    "Distinct Semantic Mapping Count",
    "Mnemonic Resolution Status",
    "Reference Curve Mnemonic",
    "Candidate Property Kind",
    "Property Kind UUID",
    "Property Kind Title",
    "Property Kind Match Method",
    "Property Kind Is Abstract",
    "Declared Quantity Class",
    "Expected Quantity Class",
    "Quantity Class Agreement",
    "LIS Curve Mnemonic",
    "Reference Curve Description",
    "Candidate Company Codes",
    "Candidate Company Names",
    "Candidate Property Kind UUIDs",
    "Candidate Property Kind Titles",
    "Candidate Quantity Classes",
    "Candidate Curve Descriptions",
    "Mnemonic Validation Status",
    "Review Required",
    "Cell 9 Status",
    "Review Reason",
    "PWLS Curve Catalogue Status",
    "PWLS Curve Source Files",
    "PWLS Curve Source Commit",
    "Property Kind Source Files",
    "Property Kind Source Commit",
    "Alias Workbook Path",
]

COMPANY_METHOD_EXACT = "Exact PWLS company name"
COMPANY_METHOD_ALIAS = "Expert-approved company alias"
COMPANY_METHOD_NONE = "No company match"
COMPANY_METHOD_MISSING = "No company supplied"

MNEMONIC_METHOD_COMPANY_EXACT = "Company-specific exact PWLS match"
MNEMONIC_METHOD_GLOBAL_EXACT = "Global exact PWLS match"
MNEMONIC_METHOD_CASE_CANDIDATE = "Global case-insensitive PWLS candidate"
MNEMONIC_METHOD_EXPERT_ALIAS = "Expert-approved mnemonic alias"
MNEMONIC_METHOD_NONE = "No mnemonic match"
MNEMONIC_METHOD_MISSING = "Missing mnemonic"

# These are the seven reporting flags defined by the validation design. Archive
# status is deliberately not a separate flag: it is provenance attached to the
# same company-specific, global, case-insensitive, alias, or no-match outcome.
MNEMONIC_FLAG_COMPANY_SINGLE = "Company Specific PWLS match [Single]"
MNEMONIC_FLAG_COMPANY_MULTIPLE = "Company Specific PWLS matches [Multiples]"
MNEMONIC_FLAG_GLOBAL_SINGLE = "Global PWLS match [Single]"
MNEMONIC_FLAG_GLOBAL_MULTIPLE = "Global PWLS match [multiples]"
MNEMONIC_FLAG_CASE_INSENSITIVE = "Case insensitive PWLS match"
MNEMONIC_FLAG_EXPERT_ALIAS = "Expert mapped alias PWLS match"
MNEMONIC_FLAG_NO_MATCH = "No PWLS match could be found"

PROPERTY_METHOD_TITLE = "Exact canonical Property Kind title"
PROPERTY_METHOD_ALIAS = "Exact Property Kind dictionary alias"
PROPERTY_METHOD_NONE = "No Property Kind match"
PROPERTY_METHOD_MISSING = "Missing Property Kind"

# Initialize every output before processing so rerunning the cell cannot expose stale results.
COMPANY_CANDIDATES_DF = pd.DataFrame(columns=COMPANY_CANDIDATE_COLUMNS)
MNEMONIC_CANDIDATES_DF = pd.DataFrame(columns=MNEMONIC_CANDIDATE_COLUMNS)
MNEMONIC_VALIDATION_DF = pd.DataFrame(columns=MNEMONIC_VALIDATION_COLUMNS)
MNEMONIC_REVIEW_DF = pd.DataFrame(columns=MNEMONIC_VALIDATION_COLUMNS)
MNEMONIC_VALIDATION_STATS = {}


def cell9_stable_text(value):
    """Return a stable comparison string without changing capitalization or spacing."""

    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass

    return str(value)


def cell9_ordered_unique(values, include_empty=False):
    """Return distinct textual values in first-seen order."""

    ordered_values = []
    seen_values = set()

    for value in values:
        text_value = cell9_stable_text(value)

        if not include_empty and not text_value:
            continue

        if text_value not in seen_values:
            seen_values.add(text_value)
            ordered_values.append(text_value)

    return tuple(ordered_values)


def cell9_append_reason(reasons, reason):
    """Append one review reason without creating duplicates."""

    if reason and reason not in reasons:
        reasons.append(reason)


def cell9_require_dataframe(object_name, required_columns, allow_empty):
    """Validate one required DataFrame and report its actual schema clearly."""

    if object_name not in globals() or globals()[object_name] is None:
        raise NameError(f"{object_name} is unavailable.")

    dataframe = globals()[object_name]

    if not isinstance(dataframe, pd.DataFrame):
        raise TypeError(f"{object_name} must be a pandas DataFrame.")

    actual_columns = list(dataframe.columns)
    missing_columns = [column for column in required_columns if column not in actual_columns]

    if missing_columns:
        raise ValueError(
            f"{object_name} is missing required columns. "
            f"Expected at least: {required_columns}. Received: {actual_columns}"
        )

    if not allow_empty and dataframe.empty:
        raise ValueError(f"{object_name} is empty.")


def cell9_validate_inputs():
    """Verify that the authoritative outputs from Cells 3, 4, 5, 7, and 8 are available."""

    dependency_messages = {
        "METADATA_DF": "METADATA_DF is unavailable. Run Cell 3 before running Cell 9.",
        "METADATA_ORIGINAL_DF": "METADATA_ORIGINAL_DF is unavailable. Run Cell 8 before running Cell 9.",
        "METADATA_INPUT_PREPROCESSING_DF": "METADATA_INPUT_PREPROCESSING_DF is unavailable. Run Cell 8 before running Cell 9.",
        "PWLS_COMPANY_CODES_DF": "PWLS_COMPANY_CODES_DF is unavailable. Run Cell 4 before running Cell 9.",
        "PWLS_CURVES_DF": "PWLS_CURVES_DF is unavailable. Run Cell 4 before running Cell 9.",
        "PWLS_ARCHIVE_DF": "PWLS_ARCHIVE_DF is unavailable. Run Cell 4 before running Cell 9.",
        "PWLS_PROPERTY_KINDS_DF": "PWLS_PROPERTY_KINDS_DF is unavailable. Run Cell 5 before running Cell 9.",
        "PWLS_PROPERTY_ALIASES_DF": "PWLS_PROPERTY_ALIASES_DF is unavailable. Run Cell 5 before running Cell 9.",
        "MNEMONIC_ALIASES_DF": "The mnemonic-alias workbook object is unavailable. Run Cell 7 before running Cell 9.",
        "COMPANY_NAME_ALIASES_DF": "The company-alias workbook object is unavailable. Run Cell 7 before running Cell 9.",
    }

    for object_name, error_message in dependency_messages.items():
        if object_name not in globals() or globals()[object_name] is None:
            raise NameError(error_message)

    for object_name in ["METADATA_DF", "METADATA_ORIGINAL_DF", "METADATA_INPUT_PREPROCESSING_DF"]:
        dataframe = globals()[object_name]

        if not isinstance(dataframe, pd.DataFrame):
            raise TypeError(f"{object_name} must be a pandas DataFrame.")

        actual_columns = list(dataframe.columns)

        if actual_columns != METADATA_VALIDATION_INPUT_COLUMNS:
            raise ValueError(
                f"{object_name} requires these columns in this exact order: "
                f"{METADATA_VALIDATION_INPUT_COLUMNS}. Received: {actual_columns}"
            )

    if len(METADATA_ORIGINAL_DF) != len(METADATA_INPUT_PREPROCESSING_DF):
        raise ValueError("Cell 8 original and prepared metadata row counts do not agree.")

    if not METADATA_ORIGINAL_DF.index.equals(METADATA_INPUT_PREPROCESSING_DF.index):
        raise ValueError("Cell 8 original and prepared metadata indices do not agree.")

    pd.testing.assert_frame_equal(METADATA_DF, METADATA_ORIGINAL_DF, check_dtype=True, check_exact=True)

    if "METADATA_INPUT_PREPROCESSING_STATS" not in globals() or not isinstance(METADATA_INPUT_PREPROCESSING_STATS, dict):
        raise NameError("METADATA_INPUT_PREPROCESSING_STATS is unavailable. Run Cell 8 before running Cell 9.")

    if METADATA_INPUT_PREPROCESSING_STATS.get("Input rows") != len(METADATA_INPUT_PREPROCESSING_DF):
        raise ValueError("Cell 8 Input Data Preprocessing statistics do not agree with METADATA_INPUT_PREPROCESSING_DF.")

    cell9_require_dataframe("PWLS_COMPANY_CODES_DF", PWLS_COMPANY_REQUIRED_COLUMNS, allow_empty=False)
    cell9_require_dataframe("PWLS_CURVES_DF", PWLS_CURVE_REQUIRED_COLUMNS, allow_empty=False)

    if not isinstance(PWLS_ARCHIVE_DF, pd.DataFrame):
        raise TypeError("PWLS_ARCHIVE_DF must be a pandas DataFrame.")

    if not PWLS_ARCHIVE_DF.empty:
        cell9_require_dataframe("PWLS_ARCHIVE_DF", PWLS_ARCHIVE_REQUIRED_COLUMNS, allow_empty=False)
    cell9_require_dataframe("PWLS_PROPERTY_KINDS_DF", PWLS_PROPERTY_REQUIRED_COLUMNS, allow_empty=False)
    cell9_require_dataframe("PWLS_PROPERTY_ALIASES_DF", PWLS_PROPERTY_ALIAS_REQUIRED_COLUMNS, allow_empty=False)
    cell9_require_dataframe("MNEMONIC_ALIASES_DF", MNEMONIC_ALIAS_REQUIRED_COLUMNS, allow_empty=True)
    cell9_require_dataframe("COMPANY_NAME_ALIASES_DF", COMPANY_ALIAS_REQUIRED_COLUMNS, allow_empty=True)

    if "PWLS_CATALOG_SOURCE" not in globals() or not isinstance(PWLS_CATALOG_SOURCE, dict) or not PWLS_CATALOG_SOURCE.get("commit"):
        raise NameError("PWLS_CATALOG_SOURCE is unavailable. Run Cell 4 before running Cell 9.")

    if "PWLS_PROPERTY_SOURCE" not in globals() or not isinstance(PWLS_PROPERTY_SOURCE, dict) or not PWLS_PROPERTY_SOURCE.get("commit"):
        raise NameError("PWLS_PROPERTY_SOURCE is unavailable. Run Cell 5 before running Cell 9.")

    if "ALIASES_WORKBOOK_PATH" not in globals() or ALIASES_WORKBOOK_PATH is None:
        raise NameError("The alias workbook objects are unavailable. Run Cell 7 before running Cell 9.")

    if "ALIASES_WORKBOOK_INFO" not in globals() or not isinstance(ALIASES_WORKBOOK_INFO, dict):
        raise NameError("ALIASES_WORKBOOK_INFO is unavailable. Run Cell 7 before running Cell 9.")


def cell9_build_indexes():
    """Build reusable positional indexes so metadata rows do not repeatedly scan reference tables."""

    company_name_index = {}
    company_alias_index = {}
    company_curve_index = {}
    global_curve_index = {}
    casefold_curve_index = {}
    archive_company_curve_index = {}
    archive_global_curve_index = {}
    archive_casefold_curve_index = {}
    mnemonic_alias_index = {}
    property_title_index = {}
    property_alias_index = {}
    property_uuid_index = {}

    for position in range(len(PWLS_COMPANY_CODES_DF)):
        company_name = cell9_stable_text(PWLS_COMPANY_CODES_DF.iloc[position]["Organization Name"])

        if company_name:
            company_name_index.setdefault(company_name, []).append(position)

    for position in range(len(COMPANY_NAME_ALIASES_DF)):
        common_name = cell9_stable_text(COMPANY_NAME_ALIASES_DF.iloc[position]["Common Company Name"])

        if common_name:
            company_alias_index.setdefault(common_name, []).append(position)

    for position in range(len(PWLS_CURVES_DF)):
        curve_row = PWLS_CURVES_DF.iloc[position]
        company_code = cell9_stable_text(curve_row["Company Code"])
        curve_mnemonic = cell9_stable_text(curve_row["Curve Mnemonic"])

        if curve_mnemonic:
            global_curve_index.setdefault(curve_mnemonic, []).append(position)
            casefold_curve_index.setdefault(curve_mnemonic.casefold(), []).append(position)

            if company_code:
                company_curve_index.setdefault((company_code, curve_mnemonic), []).append(position)

    # Cell 4 keeps every archive-path JSON row in one table. Only archived curve
    # mappings are indexed here; archived tools and other catalogue objects are
    # preserved by Cell 4 but cannot participate in mnemonic validation.
    for position in range(len(PWLS_ARCHIVE_DF)):
        curve_row = PWLS_ARCHIVE_DF.iloc[position]
        catalog_type = cell9_stable_text(curve_row["Catalog Type"])
        is_archive = cell9_stable_text(curve_row["Is Archive"]).casefold() == "true"

        if catalog_type != "curves" or not is_archive:
            continue

        company_code = cell9_stable_text(curve_row["Company Code"])
        curve_mnemonic = cell9_stable_text(curve_row["Curve Mnemonic"])

        if curve_mnemonic:
            archive_global_curve_index.setdefault(curve_mnemonic, []).append(position)
            archive_casefold_curve_index.setdefault(curve_mnemonic.casefold(), []).append(position)

            if company_code:
                archive_company_curve_index.setdefault((company_code, curve_mnemonic), []).append(position)

    for position in range(len(MNEMONIC_ALIASES_DF)):
        alias_mnemonic = cell9_stable_text(MNEMONIC_ALIASES_DF.iloc[position]["Curve Mnemonic"])

        if alias_mnemonic:
            mnemonic_alias_index.setdefault(alias_mnemonic, []).append(position)

    for position in range(len(PWLS_PROPERTY_KINDS_DF)):
        property_row = PWLS_PROPERTY_KINDS_DF.iloc[position]
        property_title = cell9_stable_text(property_row["Title"])
        property_uuid = cell9_stable_text(property_row["UUID"])

        if property_title:
            property_title_index.setdefault(property_title, []).append(position)

        if property_uuid:
            property_uuid_index.setdefault(property_uuid, []).append(position)

    for position in range(len(PWLS_PROPERTY_ALIASES_DF)):
        alias_identifier = cell9_stable_text(PWLS_PROPERTY_ALIASES_DF.iloc[position]["Identifier"])

        if alias_identifier:
            property_alias_index.setdefault(alias_identifier, []).append(position)

    return {
        "company_name": company_name_index,
        "company_alias": company_alias_index,
        "company_curve": company_curve_index,
        "global_curve": global_curve_index,
        "casefold_curve": casefold_curve_index,
        "archive_company_curve": archive_company_curve_index,
        "archive_global_curve": archive_global_curve_index,
        "archive_casefold_curve": archive_casefold_curve_index,
        "mnemonic_alias": mnemonic_alias_index,
        "property_title": property_title_index,
        "property_alias": property_alias_index,
        "property_uuid": property_uuid_index,
    }


def cell9_empty_company_candidate():
    """Create one schema-complete company candidate record."""

    return {column: "" for column in COMPANY_CANDIDATE_COLUMNS}


def cell9_company_semantics(candidate_rows):
    """Return distinct usable company identities without deleting their raw provenance rows."""

    semantic_candidates = []
    seen_keys = set()

    for candidate in candidate_rows:
        company_code = cell9_stable_text(candidate["Company Code"])
        company_name = cell9_stable_text(candidate["Company Name"])

        if candidate["Candidate Resolution Status"] != "Resolved" or not company_code or not company_name:
            continue

        semantic_key = (company_code, company_name)

        if semantic_key not in seen_keys:
            seen_keys.add(semantic_key)
            semantic_candidates.append({
                "code_text": company_code,
                "name_text": company_name,
                "code_value": candidate["Company Code"],
                "name_value": candidate["Company Name"],
            })

    return semantic_candidates


def cell9_resolve_company(source_row_order, source_index, prepared_company, indexes):
    """Resolve one prepared company through exact PWLS names and then governed aliases."""

    candidate_rows = []
    alias_problem = False

    if not prepared_company:
        return {
            "status": "Not provided",
            "method": COMPANY_METHOD_MISSING,
            "raw_count": 0,
            "semantic_candidates": [],
            "resolved_code": "",
            "resolved_name": "",
            "candidate_codes": (),
            "candidate_names": (),
            "alias_problem": False,
        }, candidate_rows

    direct_positions = indexes["company_name"].get(prepared_company, [])

    if direct_positions:
        for company_position in direct_positions:
            company_row = PWLS_COMPANY_CODES_DF.iloc[company_position]
            company_code = company_row["Company Code"]
            company_name = company_row["Organization Name"]
            candidate_status = "Resolved" if cell9_stable_text(company_code) and cell9_stable_text(company_name) else "Company code or name missing"
            candidate = cell9_empty_company_candidate()
            candidate.update({
                "Source Row Order": source_row_order,
                "Source Index": source_index,
                "Prepared Service Company": prepared_company,
                "Company Match Method": COMPANY_METHOD_EXACT,
                "PWLS Company Reference Position": company_position + 1,
                "Company Code": company_code,
                "Company Name": company_name,
                "Company Reference Source File": company_row["Source File"],
                "Company Reference Source Record Order": company_row["Source Record Order"],
                "Candidate Resolution Status": candidate_status,
            })
            candidate_rows.append(candidate)

        method = COMPANY_METHOD_EXACT

    else:
        alias_positions = indexes["company_alias"].get(prepared_company, [])

        if not alias_positions:
            return {
                "status": "Not found",
                "method": COMPANY_METHOD_NONE,
                "raw_count": 0,
                "semantic_candidates": [],
                "resolved_code": "",
                "resolved_name": "",
                "candidate_codes": (),
                "candidate_names": (),
                "alias_problem": False,
            }, candidate_rows

        method = COMPANY_METHOD_ALIAS

        for alias_position in alias_positions:
            alias_row = COMPANY_NAME_ALIASES_DF.iloc[alias_position]
            canonical_name = cell9_stable_text(alias_row["Canonical Company Name"])
            target_positions = indexes["company_name"].get(canonical_name, []) if canonical_name else []

            if not target_positions:
                alias_problem = True
                candidate = cell9_empty_company_candidate()
                candidate.update({
                    "Source Row Order": source_row_order,
                    "Source Index": source_index,
                    "Prepared Service Company": prepared_company,
                    "Company Match Method": COMPANY_METHOD_ALIAS,
                    "Company Alias Workbook Row": alias_position + 2,
                    "Alias Common Company Name": alias_row["Common Company Name"],
                    "Alias Canonical Company Name": alias_row["Canonical Company Name"],
                    "Alias Note": alias_row["Note"],
                    "Alias Reviewer Name": alias_row["Reviewer Name"],
                    "Alias Review Date": alias_row["Date (dd-mm-yyyy)"],
                    "Candidate Resolution Status": "Canonical company name missing" if not canonical_name else "Canonical company name not found",
                })
                candidate_rows.append(candidate)
                continue

            for company_position in target_positions:
                company_row = PWLS_COMPANY_CODES_DF.iloc[company_position]
                company_code = company_row["Company Code"]
                company_name = company_row["Organization Name"]
                candidate_status = "Resolved" if cell9_stable_text(company_code) and cell9_stable_text(company_name) else "Company code or name missing"

                if candidate_status != "Resolved":
                    alias_problem = True

                candidate = cell9_empty_company_candidate()
                candidate.update({
                    "Source Row Order": source_row_order,
                    "Source Index": source_index,
                    "Prepared Service Company": prepared_company,
                    "Company Match Method": COMPANY_METHOD_ALIAS,
                    "Company Alias Workbook Row": alias_position + 2,
                    "Alias Common Company Name": alias_row["Common Company Name"],
                    "Alias Canonical Company Name": alias_row["Canonical Company Name"],
                    "Alias Note": alias_row["Note"],
                    "Alias Reviewer Name": alias_row["Reviewer Name"],
                    "Alias Review Date": alias_row["Date (dd-mm-yyyy)"],
                    "PWLS Company Reference Position": company_position + 1,
                    "Company Code": company_code,
                    "Company Name": company_name,
                    "Company Reference Source File": company_row["Source File"],
                    "Company Reference Source Record Order": company_row["Source Record Order"],
                    "Candidate Resolution Status": candidate_status,
                })
                candidate_rows.append(candidate)

    semantic_candidates = cell9_company_semantics(candidate_rows)
    candidate_count = len(semantic_candidates)

    if candidate_count == 1:
        status = "Resolved"
        resolved_code = semantic_candidates[0]["code_value"]
        resolved_name = semantic_candidates[0]["name_value"]
    elif candidate_count > 1:
        status = "Ambiguous"
        resolved_code = ""
        resolved_name = ""
    else:
        status = "Not found"
        resolved_code = ""
        resolved_name = ""

    return {
        "status": status,
        "method": method,
        "raw_count": len(candidate_rows),
        "semantic_candidates": semantic_candidates,
        "resolved_code": resolved_code,
        "resolved_name": resolved_name,
        "candidate_codes": tuple(candidate["code_text"] for candidate in semantic_candidates),
        "candidate_names": tuple(candidate["name_text"] for candidate in semantic_candidates),
        "alias_problem": alias_problem,
    }, candidate_rows


def cell9_pwls_source_candidate(curve_position, source_candidate_order, catalogue_status):
    """Preserve one active or archived PWLS curve row with explicit provenance."""

    if catalogue_status == "Active":
        curve_row = PWLS_CURVES_DF.iloc[curve_position]
    elif catalogue_status == "Archived":
        curve_row = PWLS_ARCHIVE_DF.iloc[curve_position]
    else:
        raise ValueError(f"Unsupported PWLS curve catalogue status: {catalogue_status}")

    return {
        "source_candidate_order": source_candidate_order,
        "candidate_source": "PWLS Curve Catalog",
        "candidate_source_position": curve_position + 1,
        "pwls_curve_catalogue_status": catalogue_status,
        "company_code": curve_row["Company Code"],
        "company_name": curve_row["Company Name"],
        "catalog_company": curve_row["Catalog Company"],
        "curve_mnemonic": curve_row["Curve Mnemonic"],
        "property_kind": curve_row["Property Kind"],
        "declared_quantity_class": curve_row["Quantity Class"],
        "lis_curve_mnemonic": curve_row["LIS Curve Mnemonic"],
        "curve_description": curve_row["Curve Description"],
        "curve_source_file": curve_row["Source File"],
        "curve_source_record_order": curve_row["Source Record Order"],
        "curve_source_commit": PWLS_CATALOG_SOURCE["commit"],
        "mnemonic_alias_workbook_row": "",
        "mnemonic_alias_note": "",
        "mnemonic_alias_reviewer": "",
        "mnemonic_alias_date": "",
    }


def cell9_expert_alias_source_candidate(alias_position, source_candidate_order):
    """Preserve one governed workbook mnemonic alias as a raw candidate."""

    alias_row = MNEMONIC_ALIASES_DF.iloc[alias_position]
    workbook_row = alias_position + 2

    return {
        "source_candidate_order": source_candidate_order,
        "candidate_source": "Expert Review Workbook",
        "candidate_source_position": workbook_row,
        "pwls_curve_catalogue_status": "",
        "company_code": "",
        "company_name": "",
        "catalog_company": "",
        "curve_mnemonic": alias_row["Curve Mnemonic"],
        "property_kind": alias_row["Property Kind"],
        "declared_quantity_class": alias_row["Quantity Class"],
        "lis_curve_mnemonic": alias_row["LIS Curve Mnemonic"],
        "curve_description": alias_row["Curve Description"],
        "curve_source_file": "",
        "curve_source_record_order": "",
        "curve_source_commit": "",
        "mnemonic_alias_workbook_row": workbook_row,
        "mnemonic_alias_note": alias_row["Note"],
        "mnemonic_alias_reviewer": alias_row["Reviewer Name"],
        "mnemonic_alias_date": alias_row["Date (dd-mm-yyyy)"],
    }


def cell9_prefer_active_positions(active_positions, archive_positions):
    """Return active matches when present; otherwise return archived matches."""

    if active_positions:
        return sorted(set(active_positions)), "Active"

    if archive_positions:
        return sorted(set(archive_positions)), "Archived"

    return [], ""


def cell9_match_flag(match_method, candidate_count):
    """Translate one technical match method into the user's exact reporting flag."""

    if match_method == MNEMONIC_METHOD_COMPANY_EXACT:
        return MNEMONIC_FLAG_COMPANY_SINGLE if candidate_count == 1 else MNEMONIC_FLAG_COMPANY_MULTIPLE

    if match_method == MNEMONIC_METHOD_GLOBAL_EXACT:
        return MNEMONIC_FLAG_GLOBAL_SINGLE if candidate_count == 1 else MNEMONIC_FLAG_GLOBAL_MULTIPLE

    if match_method == MNEMONIC_METHOD_CASE_CANDIDATE:
        return MNEMONIC_FLAG_CASE_INSENSITIVE

    if match_method == MNEMONIC_METHOD_EXPERT_ALIAS:
        return MNEMONIC_FLAG_EXPERT_ALIAS

    return MNEMONIC_FLAG_NO_MATCH


def cell9_select_mnemonic_candidates(prepared_mnemonic, company_info, indexes):
    """Run the existing search hierarchy, preferring active data inside every PWLS route."""

    if not prepared_mnemonic:
        return "Not searched", MNEMONIC_METHOD_MISSING, MNEMONIC_FLAG_NO_MATCH, []

    active_company_positions = []
    archive_company_positions = []

    if company_info["semantic_candidates"]:
        reached_active_positions = set()
        reached_archive_positions = set()

        for company_candidate in company_info["semantic_candidates"]:
            company_key = (company_candidate["code_text"], prepared_mnemonic)

            for curve_position in indexes["company_curve"].get(company_key, []):
                reached_active_positions.add(curve_position)

            for curve_position in indexes["archive_company_curve"].get(company_key, []):
                reached_archive_positions.add(curve_position)

        active_company_positions = sorted(reached_active_positions)
        archive_company_positions = sorted(reached_archive_positions)

    if active_company_positions:
        source_candidates = [
            cell9_pwls_source_candidate(curve_position, candidate_order, "Active")
            for candidate_order, curve_position in enumerate(active_company_positions, start=1)
        ]
        match_flag = cell9_match_flag(MNEMONIC_METHOD_COMPANY_EXACT, len(source_candidates))
        return "Company context", MNEMONIC_METHOD_COMPANY_EXACT, match_flag, source_candidates

    active_global_positions = indexes["global_curve"].get(prepared_mnemonic, [])

    if active_global_positions:
        source_candidates = [
            cell9_pwls_source_candidate(curve_position, candidate_order, "Active")
            for candidate_order, curve_position in enumerate(active_global_positions, start=1)
        ]
        match_flag = cell9_match_flag(MNEMONIC_METHOD_GLOBAL_EXACT, len(source_candidates))
        return "Global", MNEMONIC_METHOD_GLOBAL_EXACT, match_flag, source_candidates

    if archive_company_positions:
        source_candidates = [
            cell9_pwls_source_candidate(curve_position, candidate_order, "Archived")
            for candidate_order, curve_position in enumerate(archive_company_positions, start=1)
        ]
        match_flag = cell9_match_flag(MNEMONIC_METHOD_COMPANY_EXACT, len(source_candidates))
        return "Company context", MNEMONIC_METHOD_COMPANY_EXACT, match_flag, source_candidates

    archive_global_positions = indexes["archive_global_curve"].get(prepared_mnemonic, [])

    if archive_global_positions:
        source_candidates = [
            cell9_pwls_source_candidate(curve_position, candidate_order, "Archived")
            for candidate_order, curve_position in enumerate(archive_global_positions, start=1)
        ]
        match_flag = cell9_match_flag(MNEMONIC_METHOD_GLOBAL_EXACT, len(source_candidates))
        return "Global", MNEMONIC_METHOD_GLOBAL_EXACT, match_flag, source_candidates

    casefold_positions, catalogue_status = cell9_prefer_active_positions(
        indexes["casefold_curve"].get(prepared_mnemonic.casefold(), []),
        indexes["archive_casefold_curve"].get(prepared_mnemonic.casefold(), []),
    )

    if casefold_positions:
        source_candidates = [
            cell9_pwls_source_candidate(curve_position, candidate_order, catalogue_status)
            for candidate_order, curve_position in enumerate(casefold_positions, start=1)
        ]
        return "Global", MNEMONIC_METHOD_CASE_CANDIDATE, MNEMONIC_FLAG_CASE_INSENSITIVE, source_candidates

    alias_positions = indexes["mnemonic_alias"].get(prepared_mnemonic, [])

    if alias_positions:
        source_candidates = [
            cell9_expert_alias_source_candidate(alias_position, candidate_order)
            for candidate_order, alias_position in enumerate(alias_positions, start=1)
        ]
        return "Expert Review Workbook", MNEMONIC_METHOD_EXPERT_ALIAS, MNEMONIC_FLAG_EXPERT_ALIAS, source_candidates

    return "Global and Expert Review Workbook", MNEMONIC_METHOD_NONE, MNEMONIC_FLAG_NO_MATCH, []


def cell9_property_record(property_position, match_method, resolution_status, alias_position=None, broken_alias=False):
    """Create one canonical Property Kind candidate with optional dictionary-alias provenance."""

    property_row = PWLS_PROPERTY_KINDS_DF.iloc[property_position]
    alias_row = PWLS_PROPERTY_ALIASES_DF.iloc[alias_position] if alias_position is not None else None

    return {
        "match_method": match_method,
        "resolution_status": resolution_status,
        "alias_position": alias_position + 1 if alias_position is not None else "",
        "alias_identifier": alias_row["Identifier"] if alias_row is not None else "",
        "alias_authority": alias_row["Authority"] if alias_row is not None else "",
        "alias_source_file": alias_row["Source File"] if alias_row is not None else "",
        "alias_source_commit": alias_row["Source Commit"] if alias_row is not None else "",
        "uuid": property_row["UUID"],
        "title": property_row["Title"],
        "description": property_row["Description"],
        "is_abstract": property_row["Is Abstract"],
        "dictionary_quantity_class": property_row["Quantity Class"],
        "property_source_file": property_row["Source File"],
        "property_source_commit": property_row["Source Commit"],
        "broken_alias": broken_alias,
    }


def cell9_unresolved_property_record(match_method, resolution_status, alias_position=None, broken_alias=False):
    """Preserve an unresolved Property Kind relationship without inventing a canonical record."""

    alias_row = PWLS_PROPERTY_ALIASES_DF.iloc[alias_position] if alias_position is not None else None

    return {
        "match_method": match_method,
        "resolution_status": resolution_status,
        "alias_position": alias_position + 1 if alias_position is not None else "",
        "alias_identifier": alias_row["Identifier"] if alias_row is not None else "",
        "alias_authority": alias_row["Authority"] if alias_row is not None else "",
        "alias_source_file": alias_row["Source File"] if alias_row is not None else "",
        "alias_source_commit": alias_row["Source Commit"] if alias_row is not None else "",
        "uuid": "",
        "title": "",
        "description": "",
        "is_abstract": "",
        "dictionary_quantity_class": "",
        "property_source_file": "",
        "property_source_commit": "",
        "broken_alias": broken_alias,
    }


def cell9_resolve_property_kind(property_kind, indexes):
    """Resolve a candidate Property Kind by canonical title and then exact dictionary alias."""

    property_text = cell9_stable_text(property_kind)

    if not property_text:
        return [cell9_unresolved_property_record(PROPERTY_METHOD_MISSING, "Missing")]

    title_positions = indexes["property_title"].get(property_text, [])

    if title_positions:
        uuid_values = cell9_ordered_unique(PWLS_PROPERTY_KINDS_DF.iloc[position]["UUID"] for position in title_positions)
        resolution_status = "Resolved" if len(uuid_values) == 1 else "Ambiguous" if len(uuid_values) > 1 else "Not found"
        records = []

        for property_position in title_positions:
            property_uuid = cell9_stable_text(PWLS_PROPERTY_KINDS_DF.iloc[property_position]["UUID"])
            row_status = resolution_status if property_uuid else "Not found"
            records.append(cell9_property_record(property_position, PROPERTY_METHOD_TITLE, row_status))

        return records

    alias_positions = indexes["property_alias"].get(property_text, [])

    if not alias_positions:
        return [cell9_unresolved_property_record(PROPERTY_METHOD_NONE, "Not found")]

    valid_relationships = []

    for alias_position in alias_positions:
        property_uuid = cell9_stable_text(PWLS_PROPERTY_ALIASES_DF.iloc[alias_position]["Property UUID"])

        for property_position in indexes["property_uuid"].get(property_uuid, []):
            valid_relationships.append((alias_position, property_position, property_uuid))

    unique_uuids = cell9_ordered_unique(relationship[2] for relationship in valid_relationships)
    valid_status = "Resolved" if len(unique_uuids) == 1 else "Ambiguous" if len(unique_uuids) > 1 else "Not found"
    records = []

    for alias_position in alias_positions:
        property_uuid = cell9_stable_text(PWLS_PROPERTY_ALIASES_DF.iloc[alias_position]["Property UUID"])
        property_positions = indexes["property_uuid"].get(property_uuid, [])

        if not property_uuid or not property_positions:
            records.append(cell9_unresolved_property_record(PROPERTY_METHOD_ALIAS, "Not found", alias_position=alias_position, broken_alias=True))
            continue

        for property_position in property_positions:
            records.append(cell9_property_record(property_position, PROPERTY_METHOD_ALIAS, valid_status, alias_position=alias_position))

    return records


def cell9_quantity_class_agreement(declared_quantity_class, property_candidate):
    """Compare the declared class with the canonical dictionary class without consulting UOM data."""

    if property_candidate["resolution_status"] not in {"Resolved", "Ambiguous"} or not cell9_stable_text(property_candidate["uuid"]):
        return "Not assessed"

    declared_text = cell9_stable_text(declared_quantity_class)
    dictionary_text = cell9_stable_text(property_candidate["dictionary_quantity_class"])

    if declared_text and dictionary_text:
        return "Agree" if declared_text == dictionary_text else "Conflict"

    if dictionary_text:
        return "Declared value missing"

    if declared_text:
        return "Dictionary value missing"

    return "Not assessed"


def cell9_expand_mnemonic_candidates(source_row_order, source_index, prepared_company, prepared_mnemonic, prepared_description, search_scope, match_method, match_flag, source_candidates, indexes):
    """Expand raw mnemonic candidates through UUID-based Property Kind resolution."""

    expanded_rows = []
    internal_rows = []

    for source_candidate in source_candidates:
        property_candidates = cell9_resolve_property_kind(source_candidate["property_kind"], indexes)

        for property_candidate in property_candidates:
            agreement = cell9_quantity_class_agreement(source_candidate["declared_quantity_class"], property_candidate)
            expanded_row = {column: "" for column in MNEMONIC_CANDIDATE_COLUMNS}
            expanded_row.update({
                "Source Row Order": source_row_order,
                "Source Index": source_index,
                "Source Candidate Order": source_candidate["source_candidate_order"],
                "Prepared Service Company": prepared_company,
                "Prepared Mnemonic": prepared_mnemonic,
                "Prepared Description": prepared_description,
                "Mnemonic Search Scope": search_scope,
                "Mnemonic Match Method": match_method,
                "Mnemonic Validation Flag": match_flag,
                "Candidate Source": source_candidate["candidate_source"],
                "Candidate Source Position": source_candidate["candidate_source_position"],
                "PWLS Curve Catalogue Status": source_candidate["pwls_curve_catalogue_status"],
                "Candidate Company Code": source_candidate["company_code"],
                "Candidate Company Name": source_candidate["company_name"],
                "Catalog Company": source_candidate["catalog_company"],
                "Reference Curve Mnemonic": source_candidate["curve_mnemonic"],
                "Candidate Property Kind": source_candidate["property_kind"],
                "Declared Quantity Class": source_candidate["declared_quantity_class"],
                "LIS Curve Mnemonic": source_candidate["lis_curve_mnemonic"],
                "Reference Curve Description": source_candidate["curve_description"],
                "Property Kind Match Method": property_candidate["match_method"],
                "Property Kind Resolution Status": property_candidate["resolution_status"],
                "Property Kind Alias Position": property_candidate["alias_position"],
                "Property Kind Alias Identifier": property_candidate["alias_identifier"],
                "Property Kind Alias Authority": property_candidate["alias_authority"],
                "Property Kind Alias Source File": property_candidate["alias_source_file"],
                "Property Kind Alias Source Commit": property_candidate["alias_source_commit"],
                "Property Kind UUID": property_candidate["uuid"],
                "Property Kind Title": property_candidate["title"],
                "Property Kind Description": property_candidate["description"],
                "Property Kind Is Abstract": property_candidate["is_abstract"],
                "Dictionary Quantity Class": property_candidate["dictionary_quantity_class"],
                "Quantity Class Agreement": agreement,
                "Curve Source File": source_candidate["curve_source_file"],
                "Curve Source Record Order": source_candidate["curve_source_record_order"],
                "Curve Source Commit": source_candidate["curve_source_commit"],
                "Property Kind Source File": property_candidate["property_source_file"],
                "Property Kind Source Commit": property_candidate["property_source_commit"],
                "Mnemonic Alias Workbook Row": source_candidate["mnemonic_alias_workbook_row"],
                "Mnemonic Alias Note": source_candidate["mnemonic_alias_note"],
                "Mnemonic Alias Reviewer Name": source_candidate["mnemonic_alias_reviewer"],
                "Mnemonic Alias Review Date": source_candidate["mnemonic_alias_date"],
            })
            expanded_rows.append(expanded_row)
            internal_rows.append({"output": expanded_row, "broken_alias": property_candidate["broken_alias"]})

    return expanded_rows, internal_rows


def cell9_aggregate_agreement(candidate_rows):
    """Aggregate candidate agreement states using a deterministic severity order."""

    agreements = cell9_ordered_unique(row["Quantity Class Agreement"] for row in candidate_rows)

    for priority in ["Conflict", "Dictionary value missing", "Not assessed", "Declared value missing", "Agree"]:
        if priority in agreements:
            return priority

    return "Not assessed"


def cell9_apply_validation(metadata_original_df, metadata_input_preprocessing_df, indexes):
    """Apply company resolution, mnemonic matching, and Property Kind enrichment to every input row."""

    company_candidate_rows = []
    mnemonic_candidate_rows = []
    validation_rows = []
    row_metrics = []
    accepted_mnemonic_methods = {MNEMONIC_METHOD_COMPANY_EXACT, MNEMONIC_METHOD_GLOBAL_EXACT, MNEMONIC_METHOD_EXPERT_ALIAS}

    for source_position in range(len(metadata_input_preprocessing_df)):
        source_row_order = source_position + 1
        source_index = metadata_input_preprocessing_df.index[source_position]
        original_values = [metadata_original_df.iat[source_position, column_position] for column_position in range(len(METADATA_VALIDATION_INPUT_COLUMNS))]
        prepared_values = [metadata_input_preprocessing_df.iat[source_position, column_position] for column_position in range(len(METADATA_VALIDATION_INPUT_COLUMNS))]
        prepared_company = cell9_stable_text(prepared_values[0])
        prepared_mnemonic = cell9_stable_text(prepared_values[1])
        prepared_description = cell9_stable_text(prepared_values[3])

        # Company identity restricts the strongest mnemonic search but is never inferred from the mnemonic itself.
        company_info, row_company_candidates = cell9_resolve_company(source_row_order, source_index, prepared_company, indexes)
        company_candidate_rows.extend(row_company_candidates)

        search_scope, mnemonic_method, mnemonic_flag, source_candidates = cell9_select_mnemonic_candidates(prepared_mnemonic, company_info, indexes)
        expanded_rows, internal_rows = cell9_expand_mnemonic_candidates(
            source_row_order,
            source_index,
            prepared_company,
            prepared_mnemonic,
            prepared_description,
            search_scope,
            mnemonic_method,
            mnemonic_flag,
            source_candidates,
            indexes,
        )
        mnemonic_candidate_rows.extend(expanded_rows)

        semantic_keys = []
        seen_semantic_keys = set()

        for candidate in expanded_rows:
            property_uuid = cell9_stable_text(candidate["Property Kind UUID"])
            dictionary_class = cell9_stable_text(candidate["Dictionary Quantity Class"])

            if property_uuid and candidate["Property Kind Resolution Status"] in {"Resolved", "Ambiguous"}:
                semantic_key = (property_uuid, dictionary_class)

                if semantic_key not in seen_semantic_keys:
                    seen_semantic_keys.add(semantic_key)
                    semantic_keys.append(semantic_key)

        semantic_mapping_count = len(semantic_keys)
        property_unresolved = not expanded_rows or any(
            row["Property Kind Resolution Status"] != "Resolved" or not cell9_stable_text(row["Property Kind UUID"])
            for row in expanded_rows
        )
        property_missing = any(row["Property Kind Resolution Status"] == "Missing" for row in expanded_rows)
        property_not_found = any(row["Property Kind Resolution Status"] == "Not found" for row in expanded_rows)
        property_ambiguous = any(row["Property Kind Resolution Status"] == "Ambiguous" for row in expanded_rows)
        broken_property_alias = any(row["broken_alias"] for row in internal_rows)
        abstract_property = any(cell9_stable_text(row["Property Kind Is Abstract"]).casefold() == "true" for row in expanded_rows)
        quantity_conflict = any(row["Quantity Class Agreement"] == "Conflict" for row in expanded_rows)
        expected_class_missing = any(
            row["Property Kind Resolution Status"] == "Resolved" and not cell9_stable_text(row["Dictionary Quantity Class"])
            for row in expanded_rows
        )

        if not prepared_mnemonic:
            mnemonic_resolution_status = "Not provided"
        elif mnemonic_method == MNEMONIC_METHOD_CASE_CANDIDATE:
            mnemonic_resolution_status = "Candidate only"
        elif not source_candidates:
            mnemonic_resolution_status = "Unresolved"
        elif semantic_mapping_count > 1:
            mnemonic_resolution_status = "Ambiguous"
        elif semantic_mapping_count == 1 and not property_unresolved:
            mnemonic_resolution_status = "Resolved"
        else:
            mnemonic_resolution_status = "Unresolved"

        mnemonic_pass = (
            mnemonic_method in accepted_mnemonic_methods
            and bool(source_candidates)
            and semantic_mapping_count == 1
            and not property_unresolved
            and bool(semantic_keys[0][1])
            and not quantity_conflict
            and not abstract_property
        )
        mnemonic_validation_status = "Pass" if mnemonic_pass else "Review required"

        review_reasons = []

        if prepared_company and company_info["alias_problem"]:
            cell9_append_reason(review_reasons, "Company alias target was missing or unresolved")

        if prepared_company and company_info["status"] == "Ambiguous":
            cell9_append_reason(review_reasons, "Service company resolved to multiple company codes")
        elif prepared_company and company_info["status"] == "Not found":
            cell9_append_reason(review_reasons, "Supplied service company was not found")

        if not prepared_mnemonic:
            cell9_append_reason(review_reasons, "Mnemonic was not provided")
        elif mnemonic_method == MNEMONIC_METHOD_NONE:
            cell9_append_reason(review_reasons, "No PWLS or approved mnemonic-alias match was found")
        elif mnemonic_method == MNEMONIC_METHOD_CASE_CANDIDATE:
            cell9_append_reason(review_reasons, "Only case-insensitive mnemonic candidates were found")

        if semantic_mapping_count > 1:
            cell9_append_reason(review_reasons, "Multiple distinct semantic mnemonic mappings were found")

        if property_missing:
            cell9_append_reason(review_reasons, "Candidate Property Kind was missing")

        if property_not_found:
            cell9_append_reason(review_reasons, "Candidate Property Kind was not found in the dictionary")

        if property_ambiguous:
            cell9_append_reason(review_reasons, "Candidate Property Kind resolved to multiple UUIDs")

        if broken_property_alias:
            cell9_append_reason(review_reasons, "Property Kind alias points to a missing UUID")

        if abstract_property:
            cell9_append_reason(review_reasons, "Resolved Property Kind is abstract")

        if expected_class_missing:
            cell9_append_reason(review_reasons, "Expected Quantity Class is missing")

        if quantity_conflict:
            cell9_append_reason(review_reasons, "Declared and dictionary Quantity Classes conflict")

        company_requires_review = bool(prepared_company) and (company_info["status"] != "Resolved" or company_info["alias_problem"])
        review_required = mnemonic_validation_status != "Pass" or company_requires_review
        cell9_status = "Review required" if review_required else "Pass"
        mapping_resolved = mnemonic_method in accepted_mnemonic_methods and semantic_mapping_count == 1 and not property_unresolved

        reference_mnemonics = cell9_ordered_unique(row["Reference Curve Mnemonic"] for row in expanded_rows)
        candidate_property_kinds = cell9_ordered_unique(row["Candidate Property Kind"] for row in expanded_rows)
        property_uuids = cell9_ordered_unique(row["Property Kind UUID"] for row in expanded_rows)
        property_titles = cell9_ordered_unique(row["Property Kind Title"] for row in expanded_rows)
        property_methods = cell9_ordered_unique(row["Property Kind Match Method"] for row in expanded_rows)
        abstract_values = cell9_ordered_unique(row["Property Kind Is Abstract"] for row in expanded_rows)
        declared_classes = cell9_ordered_unique(row["Declared Quantity Class"] for row in expanded_rows)
        dictionary_classes = cell9_ordered_unique(row["Dictionary Quantity Class"] for row in expanded_rows)
        lis_mnemonics = cell9_ordered_unique(row["LIS Curve Mnemonic"] for row in expanded_rows)
        reference_descriptions = cell9_ordered_unique(row["Reference Curve Description"] for row in expanded_rows)
        candidate_company_codes = cell9_ordered_unique(row["Candidate Company Code"] for row in expanded_rows)
        candidate_company_names = cell9_ordered_unique(row["Candidate Company Name"] for row in expanded_rows)
        curve_catalogue_statuses = cell9_ordered_unique(row["PWLS Curve Catalogue Status"] for row in expanded_rows)
        curve_source_files = cell9_ordered_unique(row["Curve Source File"] for row in expanded_rows)
        property_source_files = cell9_ordered_unique(row["Property Kind Source File"] for row in expanded_rows)
        candidate_quantity_values = []

        for row in expanded_rows:
            if cell9_stable_text(row["Declared Quantity Class"]):
                candidate_quantity_values.append(row["Declared Quantity Class"])

            if cell9_stable_text(row["Dictionary Quantity Class"]):
                candidate_quantity_values.append(row["Dictionary Quantity Class"])

        candidate_quantity_classes = cell9_ordered_unique(candidate_quantity_values)

        validation_row = {
            "Source Row Order": source_row_order,
            "Source Index": source_index,
            "Original Service Company": original_values[0],
            "Original Mnemonic": original_values[1],
            "Original Unit": original_values[2],
            "Original Description": original_values[3],
            "Prepared Service Company": prepared_values[0],
            "Prepared Mnemonic": prepared_values[1],
            "Prepared Unit": prepared_values[2],
            "Prepared Description": prepared_values[3],
            "Company Resolution Status": company_info["status"],
            "Company Match Method": company_info["method"],
            "Company Raw Match Count": company_info["raw_count"],
            "Company Candidate Count": len(company_info["semantic_candidates"]),
            "Resolved Company Code": company_info["resolved_code"],
            "Resolved Company Name": company_info["resolved_name"],
            "Company Candidate Codes": company_info["candidate_codes"],
            "Company Candidate Names": company_info["candidate_names"],
            "Mnemonic Search Scope": search_scope,
            "Mnemonic Match Method": mnemonic_method,
            "Mnemonic Validation Flag": mnemonic_flag,
            "Mnemonic Raw Candidate Count": len(source_candidates),
            "Property Candidate Row Count": len(expanded_rows),
            "Distinct Semantic Mapping Count": semantic_mapping_count,
            "Mnemonic Resolution Status": mnemonic_resolution_status,
            "Reference Curve Mnemonic": reference_mnemonics[0] if mapping_resolved and len(reference_mnemonics) == 1 else "",
            "Candidate Property Kind": candidate_property_kinds[0] if mapping_resolved and len(candidate_property_kinds) == 1 else "",
            "Property Kind UUID": property_uuids[0] if mapping_resolved and len(property_uuids) == 1 else "",
            "Property Kind Title": property_titles[0] if mapping_resolved and len(property_titles) == 1 else "",
            "Property Kind Match Method": property_methods[0] if mapping_resolved and len(property_methods) == 1 else "Multiple methods" if mapping_resolved and len(property_methods) > 1 else "",
            "Property Kind Is Abstract": abstract_values[0] if mapping_resolved and len(abstract_values) == 1 else "",
            "Declared Quantity Class": declared_classes[0] if mapping_resolved and len(declared_classes) == 1 else "",
            "Expected Quantity Class": dictionary_classes[0] if mapping_resolved and len(dictionary_classes) == 1 else "",
            "Quantity Class Agreement": cell9_aggregate_agreement(expanded_rows),
            "LIS Curve Mnemonic": lis_mnemonics[0] if mapping_resolved and len(lis_mnemonics) == 1 else "",
            "Reference Curve Description": reference_descriptions[0] if mapping_resolved and len(reference_descriptions) == 1 else "",
            "Candidate Company Codes": candidate_company_codes,
            "Candidate Company Names": candidate_company_names,
            "Candidate Property Kind UUIDs": property_uuids,
            "Candidate Property Kind Titles": property_titles,
            "Candidate Quantity Classes": candidate_quantity_classes,
            "Candidate Curve Descriptions": reference_descriptions,
            "Mnemonic Validation Status": mnemonic_validation_status,
            "Review Required": review_required,
            "Cell 9 Status": cell9_status,
            "Review Reason": "; ".join(review_reasons),
            "PWLS Curve Catalogue Status": curve_catalogue_statuses[0] if len(curve_catalogue_statuses) == 1 else "Mixed" if curve_catalogue_statuses else "",
            "PWLS Curve Source Files": curve_source_files,
            "PWLS Curve Source Commit": PWLS_CATALOG_SOURCE["commit"] if curve_source_files else "",
            "Property Kind Source Files": property_source_files,
            "Property Kind Source Commit": PWLS_PROPERTY_SOURCE["commit"] if property_source_files else "",
            "Alias Workbook Path": str(ALIASES_WORKBOOK_PATH),
        }
        validation_rows.append(validation_row)
        row_metrics.append({
            "property_unresolved": property_unresolved,
            "quantity_conflict": quantity_conflict,
            "title_resolution": any(row["Property Kind Match Method"] == PROPERTY_METHOD_TITLE and row["Property Kind Resolution Status"] == "Resolved" for row in expanded_rows),
            "alias_resolution": any(row["Property Kind Match Method"] == PROPERTY_METHOD_ALIAS and row["Property Kind Resolution Status"] == "Resolved" for row in expanded_rows),
        })

    company_candidates_df = pd.DataFrame(company_candidate_rows, columns=COMPANY_CANDIDATE_COLUMNS)
    company_candidates_df.index = pd.RangeIndex(start=1, stop=len(company_candidates_df) + 1, name="company_candidate_order")
    mnemonic_candidates_df = pd.DataFrame(mnemonic_candidate_rows, columns=MNEMONIC_CANDIDATE_COLUMNS)
    mnemonic_candidates_df.index = pd.RangeIndex(start=1, stop=len(mnemonic_candidates_df) + 1, name="mnemonic_candidate_order")
    validation_df = pd.DataFrame(validation_rows, columns=MNEMONIC_VALIDATION_COLUMNS)
    validation_df.index = metadata_input_preprocessing_df.index.copy()
    review_df = validation_df.loc[validation_df["Review Required"].eq(True)].copy()
    review_df.index = pd.RangeIndex(start=1, stop=len(review_df) + 1, name="review_order")

    stats = {
        "Input rows": len(validation_df),
        "Company resolved by exact PWLS name": int(((validation_df["Company Resolution Status"] == "Resolved") & (validation_df["Company Match Method"] == COMPANY_METHOD_EXACT)).sum()),
        "Company resolved by approved alias": int(((validation_df["Company Resolution Status"] == "Resolved") & (validation_df["Company Match Method"] == COMPANY_METHOD_ALIAS)).sum()),
        "Company inputs not provided": int((validation_df["Company Resolution Status"] == "Not provided").sum()),
        "Company inputs not found": int((validation_df["Company Resolution Status"] == "Not found").sum()),
        "Company resolutions ambiguous": int((validation_df["Company Resolution Status"] == "Ambiguous").sum()),
        MNEMONIC_FLAG_COMPANY_SINGLE: int((validation_df["Mnemonic Validation Flag"] == MNEMONIC_FLAG_COMPANY_SINGLE).sum()),
        MNEMONIC_FLAG_COMPANY_MULTIPLE: int((validation_df["Mnemonic Validation Flag"] == MNEMONIC_FLAG_COMPANY_MULTIPLE).sum()),
        MNEMONIC_FLAG_GLOBAL_SINGLE: int((validation_df["Mnemonic Validation Flag"] == MNEMONIC_FLAG_GLOBAL_SINGLE).sum()),
        MNEMONIC_FLAG_GLOBAL_MULTIPLE: int((validation_df["Mnemonic Validation Flag"] == MNEMONIC_FLAG_GLOBAL_MULTIPLE).sum()),
        MNEMONIC_FLAG_CASE_INSENSITIVE: int((validation_df["Mnemonic Validation Flag"] == MNEMONIC_FLAG_CASE_INSENSITIVE).sum()),
        MNEMONIC_FLAG_EXPERT_ALIAS: int((validation_df["Mnemonic Validation Flag"] == MNEMONIC_FLAG_EXPERT_ALIAS).sum()),
        MNEMONIC_FLAG_NO_MATCH: int((validation_df["Mnemonic Validation Flag"] == MNEMONIC_FLAG_NO_MATCH).sum()),
        "Company-specific exact mnemonic matches": int((validation_df["Mnemonic Match Method"] == MNEMONIC_METHOD_COMPANY_EXACT).sum()),
        "Global exact mnemonic matches": int((validation_df["Mnemonic Match Method"] == MNEMONIC_METHOD_GLOBAL_EXACT).sum()),
        "Case-insensitive mnemonic candidate rows": int((validation_df["Mnemonic Match Method"] == MNEMONIC_METHOD_CASE_CANDIDATE).sum()),
        "Expert-approved mnemonic alias matches": int((validation_df["Mnemonic Match Method"] == MNEMONIC_METHOD_EXPERT_ALIAS).sum()),
        "Missing mnemonics": int((validation_df["Mnemonic Match Method"] == MNEMONIC_METHOD_MISSING).sum()),
        "No mnemonic matches": int((validation_df["Mnemonic Match Method"] == MNEMONIC_METHOD_NONE).sum()),
        "PWLS matches from active catalogue": int((validation_df["PWLS Curve Catalogue Status"] == "Active").sum()),
        "PWLS matches from archived catalogue": int((validation_df["PWLS Curve Catalogue Status"] == "Archived").sum()),
        "Mnemonic passes": int((validation_df["Mnemonic Validation Status"] == "Pass").sum()),
        "Mnemonic reviews required": int((validation_df["Mnemonic Validation Status"] == "Review required").sum()),
        "Rows with one semantic mapping": int((validation_df["Distinct Semantic Mapping Count"] == 1).sum()),
        "Rows with multiple semantic mappings": int((validation_df["Distinct Semantic Mapping Count"] > 1).sum()),
        "Rows with canonical-title Property Kind resolution": sum(int(metric["title_resolution"]) for metric in row_metrics),
        "Rows with dictionary-alias Property Kind resolution": sum(int(metric["alias_resolution"]) for metric in row_metrics),
        "Rows with unresolved Property Kinds": sum(int(metric["property_unresolved"]) for metric in row_metrics),
        "Rows with Quantity Class conflicts": sum(int(metric["quantity_conflict"]) for metric in row_metrics),
        "Cell 9 passes": int((validation_df["Cell 9 Status"] == "Pass").sum()),
        "Cell 9 reviews required": int((validation_df["Cell 9 Status"] == "Review required").sum()),
    }

    return company_candidates_df, mnemonic_candidates_df, validation_df, review_df, stats


def cell9_verify_outputs(upstream_snapshots):
    """Confirm that Cell 9 preserved every upstream object and every source row."""

    for object_name, snapshot in upstream_snapshots.items():
        pd.testing.assert_frame_equal(globals()[object_name], snapshot, check_dtype=True, check_exact=True)

    if len(MNEMONIC_VALIDATION_DF) != len(METADATA_INPUT_PREPROCESSING_DF):
        raise AssertionError("MNEMONIC_VALIDATION_DF does not contain exactly one row per input row.")

    if not MNEMONIC_VALIDATION_DF.index.equals(METADATA_INPUT_PREPROCESSING_DF.index):
        raise AssertionError("MNEMONIC_VALIDATION_DF did not preserve the Input Data Preprocessing input index.")

    expected_source_orders = list(range(1, len(METADATA_INPUT_PREPROCESSING_DF) + 1))

    if MNEMONIC_VALIDATION_DF["Source Row Order"].tolist() != expected_source_orders:
        raise AssertionError("Source Row Order is not the complete one-based positional input sequence.")

    valid_source_orders = set(expected_source_orders)

    if not COMPANY_CANDIDATES_DF.empty and not set(COMPANY_CANDIDATES_DF["Source Row Order"]).issubset(valid_source_orders):
        raise AssertionError("COMPANY_CANDIDATES_DF contains an invalid Source Row Order.")

    if not MNEMONIC_CANDIDATES_DF.empty and not set(MNEMONIC_CANDIDATES_DF["Source Row Order"]).issubset(valid_source_orders):
        raise AssertionError("MNEMONIC_CANDIDATES_DF contains an invalid Source Row Order.")

    allowed_flags = {
        MNEMONIC_FLAG_COMPANY_SINGLE,
        MNEMONIC_FLAG_COMPANY_MULTIPLE,
        MNEMONIC_FLAG_GLOBAL_SINGLE,
        MNEMONIC_FLAG_GLOBAL_MULTIPLE,
        MNEMONIC_FLAG_CASE_INSENSITIVE,
        MNEMONIC_FLAG_EXPERT_ALIAS,
        MNEMONIC_FLAG_NO_MATCH,
    }

    if not set(MNEMONIC_VALIDATION_DF["Mnemonic Validation Flag"]).issubset(allowed_flags):
        raise AssertionError("Mnemonic Validation Flag contains a value outside the seven approved flags.")

    if not MNEMONIC_CANDIDATES_DF.empty:
        pwls_candidates = MNEMONIC_CANDIDATES_DF.loc[
            MNEMONIC_CANDIDATES_DF["Candidate Source"].eq("PWLS Curve Catalog")
        ]

        if not set(pwls_candidates["PWLS Curve Catalogue Status"]).issubset({"Active", "Archived"}):
            raise AssertionError("Every PWLS candidate must be marked Active or Archived.")

        selected_status_counts = pwls_candidates.groupby("Source Row Order", sort=False)["PWLS Curve Catalogue Status"].nunique()

        if not selected_status_counts.empty and int(selected_status_counts.max()) > 1:
            raise AssertionError("One source row cannot combine active and archived PWLS candidates.")

        archived_candidates = pwls_candidates.loc[pwls_candidates["PWLS Curve Catalogue Status"].eq("Archived")]

        if not archived_candidates.empty:
            invalid_archive_paths = [
                source_path
                for source_path in archived_candidates["Curve Source File"]
                if "/archive/" not in f"/{cell9_stable_text(source_path).casefold()}/"
            ]

            if invalid_archive_paths:
                raise AssertionError("An archived PWLS candidate does not preserve an archive source path.")

    company_counts = COMPANY_CANDIDATES_DF.groupby("Source Row Order", sort=False).size().to_dict() if not COMPANY_CANDIDATES_DF.empty else {}
    property_counts = MNEMONIC_CANDIDATES_DF.groupby("Source Row Order", sort=False).size().to_dict() if not MNEMONIC_CANDIDATES_DF.empty else {}
    mnemonic_raw_counts = (
        MNEMONIC_CANDIDATES_DF.groupby("Source Row Order", sort=False)["Source Candidate Order"].nunique().to_dict()
        if not MNEMONIC_CANDIDATES_DF.empty else {}
    )

    for _, validation_row in MNEMONIC_VALIDATION_DF.iterrows():
        source_row_order = validation_row["Source Row Order"]

        if validation_row["Company Raw Match Count"] != company_counts.get(source_row_order, 0):
            raise AssertionError(f"Company candidate count mismatch for Source Row Order {source_row_order}.")

        if validation_row["Property Candidate Row Count"] != property_counts.get(source_row_order, 0):
            raise AssertionError(f"Property candidate count mismatch for Source Row Order {source_row_order}.")

        if validation_row["Mnemonic Raw Candidate Count"] != mnemonic_raw_counts.get(source_row_order, 0):
            raise AssertionError(f"Mnemonic raw candidate count mismatch for Source Row Order {source_row_order}.")

    expected_review_orders = MNEMONIC_VALIDATION_DF.loc[MNEMONIC_VALIDATION_DF["Review Required"].eq(True), "Source Row Order"].tolist()

    if MNEMONIC_REVIEW_DF["Source Row Order"].tolist() != expected_review_orders:
        raise AssertionError("MNEMONIC_REVIEW_DF does not exactly represent the rows marked for review.")


try:
    cell9_validate_inputs()

    # Deep snapshots make the non-destructive behaviour auditable rather than merely assumed.
    _CELL9_UPSTREAM_SNAPSHOTS = {
        object_name: globals()[object_name].copy(deep=True)
        for object_name in [
            "METADATA_DF",
            "METADATA_ORIGINAL_DF",
            "METADATA_INPUT_PREPROCESSING_DF",
            "PWLS_COMPANY_CODES_DF",
            "PWLS_CURVES_DF",
            "PWLS_ARCHIVE_DF",
            "PWLS_PROPERTY_KINDS_DF",
            "PWLS_PROPERTY_ALIASES_DF",
            "MNEMONIC_ALIASES_DF",
            "COMPANY_NAME_ALIASES_DF",
        ]
    }
    _CELL9_INDEXES = cell9_build_indexes()

    COMPANY_CANDIDATES_DF, MNEMONIC_CANDIDATES_DF, MNEMONIC_VALIDATION_DF, MNEMONIC_REVIEW_DF, MNEMONIC_VALIDATION_STATS = cell9_apply_validation(
        METADATA_ORIGINAL_DF,
        METADATA_INPUT_PREPROCESSING_DF,
        _CELL9_INDEXES,
    )

    cell9_verify_outputs(_CELL9_UPSTREAM_SNAPSHOTS)

    print("Cell 9 company and mnemonic validation completed successfully.")

    for statistic_name, statistic_value in MNEMONIC_VALIDATION_STATS.items():
        print(f"{statistic_name}: {statistic_value:,}")

    print("\nCompany and mnemonic validation preview:")
    display(MNEMONIC_VALIDATION_DF.head(10))

    if not COMPANY_CANDIDATES_DF.empty:
        print("\nCompany candidate provenance preview:")
        display(COMPANY_CANDIDATES_DF.head(20))

    if not MNEMONIC_CANDIDATES_DF.empty:
        print("\nMnemonic candidate provenance preview:")
        display(MNEMONIC_CANDIDATES_DF.head(20))

    if MNEMONIC_REVIEW_DF.empty:
        print("\nNo company or mnemonic rows require expert review.")
    else:
        print("\nRows requiring expert review:")
        display(MNEMONIC_REVIEW_DF.head(20))

except Exception as error:
    # Reset every output so later cells cannot consume partial validation results.
    COMPANY_CANDIDATES_DF = pd.DataFrame(columns=COMPANY_CANDIDATE_COLUMNS)
    COMPANY_CANDIDATES_DF.index = pd.RangeIndex(start=1, stop=1, name="company_candidate_order")
    MNEMONIC_CANDIDATES_DF = pd.DataFrame(columns=MNEMONIC_CANDIDATE_COLUMNS)
    MNEMONIC_CANDIDATES_DF.index = pd.RangeIndex(start=1, stop=1, name="mnemonic_candidate_order")
    MNEMONIC_VALIDATION_DF = pd.DataFrame(columns=MNEMONIC_VALIDATION_COLUMNS)
    MNEMONIC_REVIEW_DF = pd.DataFrame(columns=MNEMONIC_VALIDATION_COLUMNS)
    MNEMONIC_REVIEW_DF.index = pd.RangeIndex(start=1, stop=1, name="review_order")
    MNEMONIC_VALIDATION_STATS = {}

    print("Cell 9 failed.")
    print(f"Error: {error}")


# In[ ]:


# ============================================================
# CELL 10 — UNIT VALIDATION
# ============================================================

# Requires pandas as pd and display from Cell 2.
# Uses prepared metadata from Cell 8, UOM references from Cell 6,
# and governed unit aliases from Cell 7.
# This cell recognizes units and preserves their UOM-side evidence only.
# Expected-versus-actual Quantity Class compatibility belongs exclusively to Cell 11.

UNIT_VALIDATION_INPUT_COLUMNS = ["Service company", "mnemonic", "unit", "description"]

UOM_UNIT_REQUIRED_COLUMNS = [
    "Source File", "Source Commit", "Set Version", "Unit Record Order", "Symbol", "Name",
    "Dimension", "Is SI", "Category", "Base Unit", "Conversion Reference", "Is Exact",
    "A", "B", "C", "D", "Underlying Definition", "Description", "Is Base",
]
UOM_MEMBERSHIP_REQUIRED_COLUMNS = [
    "Source File", "Source Commit", "Quantity Class Record Order", "Quantity Class",
    "Quantity Class Dimension", "Base For Conversion", "Member Unit Order", "Member Unit",
]
UOM_MAPPING_REQUIRED_COLUMNS = [
    "Source File", "Source Commit", "Mapping Source", "Mapping Record Order",
    "Maps From", "Maps To", "State", "Note",
]
UOM_UNIT_CODE_REQUIRED_COLUMNS = [
    "Source File", "Source Commit", "Unit Code Record Order", "Term", "Code", "Deprecated",
]
UNIT_ALIAS_REQUIRED_COLUMNS = ["Maps From", "Maps To", "Note", "Reviewer Name", "Date (dd-mm-yyyy)"]

# Only mappings that describe a direct identity or an authoritative correction are safe
# for automatic symbol recognition. Other states remain visible evidence for review.
UNIT_AUTOMATIC_MAPPING_STATES = {"identical", "corrected"}

UNIT_METHOD_EXACT = "Exact UOM symbol"
UNIT_METHOD_MAPPING = "Exact automatic official mapping"
UNIT_METHOD_MAPPING_REVIEW = "Official mapping requires review"
UNIT_METHOD_ALIAS = "Exact approved unit alias"
UNIT_METHOD_CASE = "Case-insensitive UOM candidate"
UNIT_METHOD_NONE = "No unit match"
UNIT_METHOD_MISSING = "No unit supplied"

UNIT_CANDIDATE_COLUMNS = [
    "Source Row Order",
    "Source Index",
    "Original Unit",
    "Prepared Unit",
    "Search Stage",
    "Candidate Source",
    "Candidate Source Position",
    "Unit Match Method",
    "Source Match Value",
    "Canonical Unit",
    "Candidate Resolution Status",
    "Candidate Issue",
    "UOM Unit Reference Position",
    "Unit Name",
    "Unit Dimension",
    "Unit Is SI",
    "Unit Category",
    "Unit Base Unit",
    "Unit Conversion Reference",
    "Unit Definition Is Exact",
    "Conversion A",
    "Conversion B",
    "Conversion C",
    "Conversion D",
    "Unit Underlying Definition",
    "Unit Description",
    "Unit Is Base",
    "Unit Source File",
    "Unit Source Commit",
    "Unit Source Record Order",
    "Quantity Class Membership Row Count",
    "Distinct Quantity Class Membership Count",
    "Actual Quantity Class Memberships",
    "Quantity Class Membership Reference Positions",
    "Quantity Class Membership Source Files",
    "Unit Code Record Count",
    "Unit Integer Codes",
    "Unit Code Deprecated Values",
    "Unit Code Reference Positions",
    "Mapping Source",
    "Mapping Record Order",
    "Mapping State",
    "Mapping Note",
    "Mapping Source File",
    "Mapping Source Commit",
    "Mapping Target Exists",
    "Mapping State Automatically Accepted",
    "Unit Alias Workbook Row",
    "Alias Maps From",
    "Alias Maps To",
    "Alias Note",
    "Alias Reviewer Name",
    "Alias Review Date",
    "Alias Target Exists",
]

UNIT_VALIDATION_COLUMNS = [
    "Source Row Order",
    "Source Index",
    "Original Service Company",
    "Original Mnemonic",
    "Original Unit",
    "Original Description",
    "Prepared Service Company",
    "Prepared Mnemonic",
    "Prepared Unit",
    "Prepared Description",
    "Unit Recognition Status",
    "Unit Match Method",
    "Unit Raw Candidate Count",
    "Distinct Candidate Canonical Unit Count",
    "Canonical Unit",
    "Unit Name",
    "Unit Dimension",
    "Unit Is SI",
    "Unit Category",
    "Unit Base Unit",
    "Unit Conversion Reference",
    "Unit Definition Is Exact",
    "Conversion A",
    "Conversion B",
    "Conversion C",
    "Conversion D",
    "Unit Underlying Definition",
    "Unit Description",
    "Unit Is Base",
    "Quantity Class Membership Row Count",
    "Distinct Quantity Class Membership Count",
    "Actual Quantity Class Memberships",
    "Quantity Class Membership Reference Positions",
    "Unit Code Record Count",
    "Unit Integer Codes",
    "Unit Code Deprecated Values",
    "Candidate Canonical Units",
    "Candidate Unit Names",
    "Candidate Unit Dimensions",
    "Official Mapping Sources",
    "Official Mapping States",
    "Official Mapping Notes",
    "Unit Alias Workbook Rows",
    "Unit Alias Notes",
    "Unit Alias Reviewers",
    "Unit Alias Review Dates",
    "Unit Validation Status",
    "Review Required",
    "Cell 10 Status",
    "Review Reason",
    "UOM Unit Source Files",
    "UOM Membership Source Files",
    "UOM Mapping Source Files",
    "UOM Source Commit",
    "Alias Workbook Path",
]

# Initialize every output before processing so a rerun cannot expose stale results.
UNIT_CANDIDATES_DF = pd.DataFrame(columns=UNIT_CANDIDATE_COLUMNS)
UNIT_VALIDATION_DF = pd.DataFrame(columns=UNIT_VALIDATION_COLUMNS)
UNIT_REVIEW_DF = pd.DataFrame(columns=UNIT_VALIDATION_COLUMNS)
UNIT_VALIDATION_STATS = {}


def cell10_stable_text(value):
    """Return a stable textual value without changing case, spacing, or punctuation."""

    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass

    return str(value)


def cell10_ordered_unique(values, include_empty=False):
    """Return distinct textual values in their first-seen order."""

    ordered_values = []
    seen_values = set()

    for value in values:
        text_value = cell10_stable_text(value)

        if not include_empty and not text_value:
            continue

        if text_value not in seen_values:
            seen_values.add(text_value)
            ordered_values.append(text_value)

    return tuple(ordered_values)


def cell10_scalar_or_tuple(values):
    """Return one value directly, multiple values as a tuple, or an empty string."""

    unique_values = cell10_ordered_unique(values)

    if len(unique_values) == 1:
        return unique_values[0]

    return unique_values if unique_values else ""


def cell10_append_reason(reasons, reason):
    """Append one review reason without introducing repeated messages."""

    if reason and reason not in reasons:
        reasons.append(reason)


def cell10_require_dataframe(object_name, required_columns, allow_empty):
    """Validate one required DataFrame and report its actual schema clearly."""

    if object_name not in globals() or globals()[object_name] is None:
        raise NameError(f"{object_name} is unavailable.")

    dataframe = globals()[object_name]

    if not isinstance(dataframe, pd.DataFrame):
        raise TypeError(f"{object_name} must be a pandas DataFrame.")

    actual_columns = list(dataframe.columns)
    missing_columns = [column for column in required_columns if column not in actual_columns]

    if missing_columns:
        raise ValueError(
            f"{object_name} is missing required columns. "
            f"Expected at least: {required_columns}. Received: {actual_columns}"
        )

    if not allow_empty and dataframe.empty:
        raise ValueError(f"{object_name} is empty.")


def cell10_validate_inputs():
    """Verify the independent unit-side inputs from Cells 3, 6, 7, and 8."""

    dependency_messages = {
        "METADATA_DF": "METADATA_DF is unavailable. Run Cell 3 before running Cell 10.",
        "METADATA_ORIGINAL_DF": "METADATA_ORIGINAL_DF is unavailable. Run Cell 8 before running Cell 10.",
        "METADATA_INPUT_PREPROCESSING_DF": "METADATA_INPUT_PREPROCESSING_DF is unavailable. Run Cell 8 before running Cell 10.",
        "UOM_UNITS_DF": "UOM_UNITS_DF is unavailable. Run Cell 6 before running Cell 10.",
        "UOM_QUANTITY_CLASS_MEMBERS_DF": "UOM_QUANTITY_CLASS_MEMBERS_DF is unavailable. Run Cell 6 before running Cell 10.",
        "UOM_UNIT_MAPPINGS_DF": "UOM_UNIT_MAPPINGS_DF is unavailable. Run Cell 6 before running Cell 10.",
        "UOM_UNIT_CODES_DF": "UOM_UNIT_CODES_DF is unavailable. Run Cell 6 before running Cell 10.",
        "UNIT_ALIASES_DF": "UNIT_ALIASES_DF is unavailable. Run Cell 7 before running Cell 10.",
    }

    for object_name, error_message in dependency_messages.items():
        if object_name not in globals() or globals()[object_name] is None:
            raise NameError(error_message)

    for object_name in ["METADATA_DF", "METADATA_ORIGINAL_DF", "METADATA_INPUT_PREPROCESSING_DF"]:
        dataframe = globals()[object_name]

        if not isinstance(dataframe, pd.DataFrame):
            raise TypeError(f"{object_name} must be a pandas DataFrame.")

        actual_columns = list(dataframe.columns)

        if actual_columns != UNIT_VALIDATION_INPUT_COLUMNS:
            raise ValueError(
                f"{object_name} requires these columns in this exact order: "
                f"{UNIT_VALIDATION_INPUT_COLUMNS}. Received: {actual_columns}"
            )

    if len(METADATA_ORIGINAL_DF) != len(METADATA_INPUT_PREPROCESSING_DF):
        raise ValueError("Cell 8 original and prepared row counts do not agree.")

    if not METADATA_ORIGINAL_DF.index.equals(METADATA_INPUT_PREPROCESSING_DF.index):
        raise ValueError("Cell 8 original and prepared indices do not agree.")

    pd.testing.assert_frame_equal(METADATA_DF, METADATA_ORIGINAL_DF, check_dtype=True, check_exact=True)

    if "METADATA_INPUT_PREPROCESSING_STATS" not in globals() or not isinstance(METADATA_INPUT_PREPROCESSING_STATS, dict):
        raise NameError("METADATA_INPUT_PREPROCESSING_STATS is unavailable. Run Cell 8 before running Cell 10.")

    if METADATA_INPUT_PREPROCESSING_STATS.get("Input rows") != len(METADATA_INPUT_PREPROCESSING_DF):
        raise ValueError("Cell 8 Input Data Preprocessing statistics do not agree with METADATA_INPUT_PREPROCESSING_DF.")

    cell10_require_dataframe("UOM_UNITS_DF", UOM_UNIT_REQUIRED_COLUMNS, allow_empty=False)
    cell10_require_dataframe("UOM_QUANTITY_CLASS_MEMBERS_DF", UOM_MEMBERSHIP_REQUIRED_COLUMNS, allow_empty=False)
    cell10_require_dataframe("UOM_UNIT_MAPPINGS_DF", UOM_MAPPING_REQUIRED_COLUMNS, allow_empty=False)
    cell10_require_dataframe("UOM_UNIT_CODES_DF", UOM_UNIT_CODE_REQUIRED_COLUMNS, allow_empty=False)
    cell10_require_dataframe("UNIT_ALIASES_DF", UNIT_ALIAS_REQUIRED_COLUMNS, allow_empty=True)

    if "UOM_SOURCE" not in globals() or not isinstance(UOM_SOURCE, dict) or not UOM_SOURCE.get("commit"):
        raise NameError("UOM_SOURCE is unavailable. Run Cell 6 before running Cell 10.")

    if "ALIASES_WORKBOOK_PATH" not in globals() or ALIASES_WORKBOOK_PATH is None:
        raise NameError("The alias workbook objects are unavailable. Run Cell 7 before running Cell 10.")

    if "ALIASES_WORKBOOK_INFO" not in globals() or not isinstance(ALIASES_WORKBOOK_INFO, dict):
        raise NameError("ALIASES_WORKBOOK_INFO is unavailable. Run Cell 7 before running Cell 10.")


def cell10_build_indexes():
    """Build positional indexes without deduplicating any authoritative source row."""

    exact_unit_index = {}
    casefold_unit_index = {}
    membership_index = {}
    mapping_index = {}
    alias_index = {}
    unit_code_index = {}

    for position in range(len(UOM_UNITS_DF)):
        symbol = cell10_stable_text(UOM_UNITS_DF.iloc[position]["Symbol"])

        if symbol:
            exact_unit_index.setdefault(symbol, []).append(position)
            casefold_unit_index.setdefault(symbol.casefold(), []).append(position)

    for position in range(len(UOM_QUANTITY_CLASS_MEMBERS_DF)):
        member_unit = cell10_stable_text(UOM_QUANTITY_CLASS_MEMBERS_DF.iloc[position]["Member Unit"])

        if member_unit:
            membership_index.setdefault(member_unit, []).append(position)

    for position in range(len(UOM_UNIT_MAPPINGS_DF)):
        maps_from = cell10_stable_text(UOM_UNIT_MAPPINGS_DF.iloc[position]["Maps From"])

        if maps_from:
            mapping_index.setdefault(maps_from, []).append(position)

    for position in range(len(UNIT_ALIASES_DF)):
        maps_from = cell10_stable_text(UNIT_ALIASES_DF.iloc[position]["Maps From"])

        if maps_from:
            alias_index.setdefault(maps_from, []).append(position)

    for position in range(len(UOM_UNIT_CODES_DF)):
        unit_term = cell10_stable_text(UOM_UNIT_CODES_DF.iloc[position]["Term"])

        if unit_term:
            unit_code_index.setdefault(unit_term, []).append(position)

    return {
        "unit_exact": exact_unit_index,
        "unit_casefold": casefold_unit_index,
        "membership": membership_index,
        "mapping": mapping_index,
        "alias": alias_index,
        "unit_code": unit_code_index,
    }


def cell10_empty_candidate():
    """Create one schema-complete unit-evidence record."""

    return {column: "" for column in UNIT_CANDIDATE_COLUMNS}


def cell10_unit_details(canonical_unit, indexes):
    """Collect UOM definition, membership, code, and provenance details for one symbol."""

    unit_positions = indexes["unit_exact"].get(canonical_unit, [])
    membership_positions = indexes["membership"].get(canonical_unit, [])
    code_positions = indexes["unit_code"].get(canonical_unit, [])

    unit_rows = [UOM_UNITS_DF.iloc[position] for position in unit_positions]
    membership_rows = [UOM_QUANTITY_CLASS_MEMBERS_DF.iloc[position] for position in membership_positions]
    code_rows = [UOM_UNIT_CODES_DF.iloc[position] for position in code_positions]

    return {
        "unit_positions": tuple(position + 1 for position in unit_positions),
        "name": cell10_scalar_or_tuple(row["Name"] for row in unit_rows),
        "dimension": cell10_scalar_or_tuple(row["Dimension"] for row in unit_rows),
        "is_si": cell10_scalar_or_tuple(row["Is SI"] for row in unit_rows),
        "category": cell10_scalar_or_tuple(row["Category"] for row in unit_rows),
        "base_unit": cell10_scalar_or_tuple(row["Base Unit"] for row in unit_rows),
        "conversion_reference": cell10_scalar_or_tuple(row["Conversion Reference"] for row in unit_rows),
        "is_exact": cell10_scalar_or_tuple(row["Is Exact"] for row in unit_rows),
        "a": cell10_scalar_or_tuple(row["A"] for row in unit_rows),
        "b": cell10_scalar_or_tuple(row["B"] for row in unit_rows),
        "c": cell10_scalar_or_tuple(row["C"] for row in unit_rows),
        "d": cell10_scalar_or_tuple(row["D"] for row in unit_rows),
        "underlying_definition": cell10_scalar_or_tuple(row["Underlying Definition"] for row in unit_rows),
        "description": cell10_scalar_or_tuple(row["Description"] for row in unit_rows),
        "is_base": cell10_scalar_or_tuple(row["Is Base"] for row in unit_rows),
        "unit_source_files": cell10_ordered_unique(row["Source File"] for row in unit_rows),
        "unit_source_commits": cell10_ordered_unique(row["Source Commit"] for row in unit_rows),
        "unit_source_record_orders": cell10_ordered_unique(row["Unit Record Order"] for row in unit_rows),
        "membership_row_count": len(membership_rows),
        "membership_count": len(cell10_ordered_unique(row["Quantity Class"] for row in membership_rows)),
        "memberships": cell10_ordered_unique(row["Quantity Class"] for row in membership_rows),
        "membership_positions": tuple(position + 1 for position in membership_positions),
        "membership_source_files": cell10_ordered_unique(row["Source File"] for row in membership_rows),
        "membership_source_commits": cell10_ordered_unique(row["Source Commit"] for row in membership_rows),
        "code_row_count": len(code_rows),
        "codes": cell10_ordered_unique(row["Code"] for row in code_rows),
        "code_deprecated_values": cell10_ordered_unique(
            (row["Deprecated"] for row in code_rows),
            include_empty=True,
        ),
        "code_positions": tuple(position + 1 for position in code_positions),
    }


def cell10_fill_unit_fields(candidate, unit_position, indexes):
    """Attach one exact UOM definition and its complete independent unit-side evidence."""

    unit_row = UOM_UNITS_DF.iloc[unit_position]
    canonical_unit = cell10_stable_text(unit_row["Symbol"])
    details = cell10_unit_details(canonical_unit, indexes)

    candidate.update({
        "Canonical Unit": unit_row["Symbol"],
        "UOM Unit Reference Position": unit_position + 1,
        "Unit Name": unit_row["Name"],
        "Unit Dimension": unit_row["Dimension"],
        "Unit Is SI": unit_row["Is SI"],
        "Unit Category": unit_row["Category"],
        "Unit Base Unit": unit_row["Base Unit"],
        "Unit Conversion Reference": unit_row["Conversion Reference"],
        "Unit Definition Is Exact": unit_row["Is Exact"],
        "Conversion A": unit_row["A"],
        "Conversion B": unit_row["B"],
        "Conversion C": unit_row["C"],
        "Conversion D": unit_row["D"],
        "Unit Underlying Definition": unit_row["Underlying Definition"],
        "Unit Description": unit_row["Description"],
        "Unit Is Base": unit_row["Is Base"],
        "Unit Source File": unit_row["Source File"],
        "Unit Source Commit": unit_row["Source Commit"],
        "Unit Source Record Order": unit_row["Unit Record Order"],
        "Quantity Class Membership Row Count": details["membership_row_count"],
        "Distinct Quantity Class Membership Count": details["membership_count"],
        "Actual Quantity Class Memberships": details["memberships"],
        "Quantity Class Membership Reference Positions": details["membership_positions"],
        "Quantity Class Membership Source Files": details["membership_source_files"],
        "Unit Code Record Count": details["code_row_count"],
        "Unit Integer Codes": details["codes"],
        "Unit Code Deprecated Values": details["code_deprecated_values"],
        "Unit Code Reference Positions": details["code_positions"],
    })


def cell10_dictionary_candidates(source_row_order, source_index, original_unit, prepared_unit, unit_positions, method, stage, status, indexes):
    """Create one evidence row for every preserved UOM definition candidate."""

    candidates = []

    for unit_position in unit_positions:
        candidate = cell10_empty_candidate()
        candidate.update({
            "Source Row Order": source_row_order,
            "Source Index": source_index,
            "Original Unit": original_unit,
            "Prepared Unit": prepared_unit,
            "Search Stage": stage,
            "Candidate Source": "Energistics UOM Dictionary",
            "Candidate Source Position": unit_position + 1,
            "Unit Match Method": method,
            "Source Match Value": prepared_unit,
            "Candidate Resolution Status": status,
        })
        cell10_fill_unit_fields(candidate, unit_position, indexes)
        candidates.append(candidate)

    return candidates


def cell10_mapping_candidates(source_row_order, source_index, original_unit, prepared_unit, mapping_positions, indexes):
    """Preserve every exact official mapping row and evaluate its target safely."""

    candidates = []

    for mapping_position in mapping_positions:
        mapping_row = UOM_UNIT_MAPPINGS_DF.iloc[mapping_position]
        mapping_target = cell10_stable_text(mapping_row["Maps To"])
        mapping_state = cell10_stable_text(mapping_row["State"])
        target_positions = indexes["unit_exact"].get(mapping_target, []) if mapping_target else []
        state_accepted = mapping_state.casefold() in UNIT_AUTOMATIC_MAPPING_STATES

        if not mapping_target:
            candidate_status = "Mapping target missing"
            candidate_issue = "Official mapping has an empty Maps To value"
        elif not target_positions:
            candidate_status = "Mapping target not found"
            candidate_issue = "Official mapping target is absent from UOM_UNITS_DF"
        elif not state_accepted:
            candidate_status = "Mapping state requires review"
            candidate_issue = f"Mapping state '{mapping_state}' is not automatically accepted"
        else:
            candidate_status = "Resolved"
            candidate_issue = ""

        output_positions = target_positions if target_positions else [None]

        for unit_position in output_positions:
            candidate = cell10_empty_candidate()
            candidate.update({
                "Source Row Order": source_row_order,
                "Source Index": source_index,
                "Original Unit": original_unit,
                "Prepared Unit": prepared_unit,
                "Search Stage": "Official mapping",
                "Candidate Source": "Official UOM Mapping",
                "Candidate Source Position": mapping_position + 1,
                "Unit Match Method": UNIT_METHOD_MAPPING if candidate_status == "Resolved" else UNIT_METHOD_MAPPING_REVIEW,
                "Source Match Value": mapping_row["Maps From"],
                "Canonical Unit": mapping_row["Maps To"],
                "Candidate Resolution Status": candidate_status,
                "Candidate Issue": candidate_issue,
                "Mapping Source": mapping_row["Mapping Source"],
                "Mapping Record Order": mapping_row["Mapping Record Order"],
                "Mapping State": mapping_row["State"],
                "Mapping Note": mapping_row["Note"],
                "Mapping Source File": mapping_row["Source File"],
                "Mapping Source Commit": mapping_row["Source Commit"],
                "Mapping Target Exists": bool(target_positions),
                "Mapping State Automatically Accepted": state_accepted,
            })

            if unit_position is not None:
                cell10_fill_unit_fields(candidate, unit_position, indexes)

            candidates.append(candidate)

    return candidates


def cell10_alias_candidates(source_row_order, source_index, original_unit, prepared_unit, alias_positions, indexes):
    """Preserve every governed workbook alias and verify its canonical target."""

    candidates = []

    for alias_position in alias_positions:
        alias_row = UNIT_ALIASES_DF.iloc[alias_position]
        alias_target = cell10_stable_text(alias_row["Maps To"])
        target_positions = indexes["unit_exact"].get(alias_target, []) if alias_target else []

        if not alias_target:
            candidate_status = "Alias target missing"
            candidate_issue = "Approved unit alias has an empty Maps To value"
        elif not target_positions:
            candidate_status = "Alias target not found"
            candidate_issue = "Approved unit alias target is absent from UOM_UNITS_DF"
        else:
            candidate_status = "Resolved"
            candidate_issue = ""

        output_positions = target_positions if target_positions else [None]

        for unit_position in output_positions:
            candidate = cell10_empty_candidate()
            candidate.update({
                "Source Row Order": source_row_order,
                "Source Index": source_index,
                "Original Unit": original_unit,
                "Prepared Unit": prepared_unit,
                "Search Stage": "Expert Review Workbook",
                "Candidate Source": "Expert Review Workbook",
                "Candidate Source Position": alias_position + 2,
                "Unit Match Method": UNIT_METHOD_ALIAS,
                "Source Match Value": alias_row["Maps From"],
                "Canonical Unit": alias_row["Maps To"],
                "Candidate Resolution Status": candidate_status,
                "Candidate Issue": candidate_issue,
                "Unit Alias Workbook Row": alias_position + 2,
                "Alias Maps From": alias_row["Maps From"],
                "Alias Maps To": alias_row["Maps To"],
                "Alias Note": alias_row["Note"],
                "Alias Reviewer Name": alias_row["Reviewer Name"],
                "Alias Review Date": alias_row["Date (dd-mm-yyyy)"],
                "Alias Target Exists": bool(target_positions),
            })

            if unit_position is not None:
                cell10_fill_unit_fields(candidate, unit_position, indexes)

            candidates.append(candidate)

    return candidates


def cell10_candidate_units(candidate_rows, statuses=None):
    """Return distinct canonical symbols from selected candidate-resolution states."""

    return cell10_ordered_unique(
        candidate["Canonical Unit"]
        for candidate in candidate_rows
        if statuses is None or candidate["Candidate Resolution Status"] in statuses
    )


def cell10_resolve_unit(source_row_order, source_index, original_unit, prepared_unit, indexes):
    """Apply the exact-symbol, official-mapping, alias, and case-candidate hierarchy."""

    candidate_rows = []
    review_reasons = []

    if not prepared_unit:
        return {
            "status": "Not provided",
            "method": UNIT_METHOD_MISSING,
            "canonical_unit": "",
            "review_reasons": [],
        }, candidate_rows

    exact_positions = indexes["unit_exact"].get(prepared_unit, [])

    if exact_positions:
        candidate_rows.extend(cell10_dictionary_candidates(
            source_row_order,
            source_index,
            original_unit,
            prepared_unit,
            exact_positions,
            UNIT_METHOD_EXACT,
            "Exact dictionary symbol",
            "Resolved",
            indexes,
        ))
        return {
            "status": "Recognized",
            "method": UNIT_METHOD_EXACT,
            "canonical_unit": prepared_unit,
            "review_reasons": [],
        }, candidate_rows

    mapping_positions = indexes["mapping"].get(prepared_unit, [])
    mapping_rows = cell10_mapping_candidates(
        source_row_order,
        source_index,
        original_unit,
        prepared_unit,
        mapping_positions,
        indexes,
    ) if mapping_positions else []
    candidate_rows.extend(mapping_rows)

    approved_mapping_targets = cell10_candidate_units(mapping_rows, statuses={"Resolved"})
    all_existing_mapping_targets = cell10_ordered_unique(
        candidate["Canonical Unit"]
        for candidate in mapping_rows
        if candidate["Mapping Target Exists"] is True
    )
    mapping_blockers = [candidate for candidate in mapping_rows if candidate["Candidate Resolution Status"] != "Resolved"]

    if len(approved_mapping_targets) > 1:
        cell10_append_reason(review_reasons, "Exact official mappings resolve to multiple automatically accepted canonical units")
        return {
            "status": "Ambiguous",
            "method": UNIT_METHOD_MAPPING,
            "canonical_unit": "",
            "review_reasons": review_reasons,
        }, candidate_rows

    if len(approved_mapping_targets) == 1 and not mapping_blockers and len(all_existing_mapping_targets) == 1:
        return {
            "status": "Recognized",
            "method": UNIT_METHOD_MAPPING,
            "canonical_unit": approved_mapping_targets[0],
            "review_reasons": [],
        }, candidate_rows

    if mapping_rows:
        unaccepted_states = cell10_ordered_unique(
            candidate["Mapping State"]
            for candidate in mapping_rows
            if candidate["Candidate Resolution Status"] == "Mapping state requires review"
        )

        if unaccepted_states:
            cell10_append_reason(
                review_reasons,
                "Official mapping states require review: " + ", ".join(unaccepted_states),
            )

        if any(candidate["Candidate Resolution Status"] == "Mapping target missing" for candidate in mapping_rows):
            cell10_append_reason(review_reasons, "An exact official mapping has no canonical target")

        if any(candidate["Candidate Resolution Status"] == "Mapping target not found" for candidate in mapping_rows):
            cell10_append_reason(review_reasons, "An exact official mapping points to a unit absent from the UOM Dictionary")

        if len(all_existing_mapping_targets) > 1:
            cell10_append_reason(review_reasons, "Exact official mappings preserve multiple canonical-unit candidates")

    alias_positions = indexes["alias"].get(prepared_unit, [])
    alias_rows = cell10_alias_candidates(
        source_row_order,
        source_index,
        original_unit,
        prepared_unit,
        alias_positions,
        indexes,
    ) if alias_positions else []
    candidate_rows.extend(alias_rows)

    valid_alias_targets = cell10_candidate_units(alias_rows, statuses={"Resolved"})
    alias_blockers = [candidate for candidate in alias_rows if candidate["Candidate Resolution Status"] != "Resolved"]

    if len(valid_alias_targets) > 1 or (valid_alias_targets and alias_blockers):
        cell10_append_reason(review_reasons, "Approved unit aliases do not resolve to one complete canonical target")
        return {
            "status": "Ambiguous",
            "method": UNIT_METHOD_ALIAS,
            "canonical_unit": "",
            "review_reasons": review_reasons,
        }, candidate_rows

    if len(valid_alias_targets) == 1:
        alias_target = valid_alias_targets[0]
        approved_mapping_conflict = approved_mapping_targets and approved_mapping_targets[0] != alias_target

        if approved_mapping_conflict:
            cell10_append_reason(review_reasons, "The approved alias conflicts with an automatically accepted official mapping")
            return {
                "status": "Ambiguous",
                "method": UNIT_METHOD_ALIAS,
                "canonical_unit": "",
                "review_reasons": review_reasons,
            }, candidate_rows

        # A valid workbook alias is the governed resolution when no stronger mapping was safe.
        return {
            "status": "Recognized",
            "method": UNIT_METHOD_ALIAS,
            "canonical_unit": alias_target,
            "review_reasons": [],
        }, candidate_rows

    if alias_rows:
        if any(candidate["Candidate Resolution Status"] == "Alias target missing" for candidate in alias_rows):
            cell10_append_reason(review_reasons, "An approved unit alias has no canonical target")

        if any(candidate["Candidate Resolution Status"] == "Alias target not found" for candidate in alias_rows):
            cell10_append_reason(review_reasons, "An approved unit alias points to a unit absent from the UOM Dictionary")

    case_positions = indexes["unit_casefold"].get(prepared_unit.casefold(), [])

    if case_positions:
        candidate_rows.extend(cell10_dictionary_candidates(
            source_row_order,
            source_index,
            original_unit,
            prepared_unit,
            case_positions,
            UNIT_METHOD_CASE,
            "Case-insensitive dictionary candidate",
            "Candidate only",
            indexes,
        ))
        cell10_append_reason(review_reasons, "Only case-insensitive UOM symbol candidates were found")
        return {
            "status": "Candidate only",
            "method": UNIT_METHOD_CASE,
            "canonical_unit": "",
            "review_reasons": review_reasons,
        }, candidate_rows

    if mapping_rows:
        return {
            "status": "Not recognized",
            "method": UNIT_METHOD_MAPPING_REVIEW,
            "canonical_unit": "",
            "review_reasons": review_reasons,
        }, candidate_rows

    if alias_rows:
        return {
            "status": "Not recognized",
            "method": UNIT_METHOD_ALIAS,
            "canonical_unit": "",
            "review_reasons": review_reasons,
        }, candidate_rows

    cell10_append_reason(review_reasons, "Unit was not found in the UOM Dictionary, official mappings, or approved aliases")
    return {
        "status": "Not recognized",
        "method": UNIT_METHOD_NONE,
        "canonical_unit": "",
        "review_reasons": review_reasons,
    }, candidate_rows


def cell10_apply_validation(metadata_original_df, metadata_input_preprocessing_df, indexes):
    """Recognize every prepared unit while preserving source rows and evidence cardinality."""

    candidate_rows = []
    validation_rows = []

    for source_position in range(len(metadata_input_preprocessing_df)):
        source_row_order = source_position + 1
        source_index = metadata_input_preprocessing_df.index[source_position]
        original_values = [metadata_original_df.iat[source_position, column_position] for column_position in range(len(UNIT_VALIDATION_INPUT_COLUMNS))]
        prepared_values = [metadata_input_preprocessing_df.iat[source_position, column_position] for column_position in range(len(UNIT_VALIDATION_INPUT_COLUMNS))]
        original_unit = original_values[2]
        prepared_unit = cell10_stable_text(prepared_values[2])

        unit_info, row_candidates = cell10_resolve_unit(
            source_row_order,
            source_index,
            original_unit,
            prepared_unit,
            indexes,
        )
        candidate_rows.extend(row_candidates)

        canonical_unit = cell10_stable_text(unit_info["canonical_unit"])
        details = cell10_unit_details(canonical_unit, indexes) if canonical_unit else {
            "name": "",
            "dimension": "",
            "is_si": "",
            "category": "",
            "base_unit": "",
            "conversion_reference": "",
            "is_exact": "",
            "a": "",
            "b": "",
            "c": "",
            "d": "",
            "underlying_definition": "",
            "description": "",
            "is_base": "",
            "unit_source_files": (),
            "membership_row_count": 0,
            "membership_count": 0,
            "memberships": (),
            "membership_positions": (),
            "membership_source_files": (),
            "code_row_count": 0,
            "codes": (),
            "code_deprecated_values": (),
        }

        candidate_units = cell10_candidate_units(row_candidates)
        candidate_names = cell10_ordered_unique(candidate["Unit Name"] for candidate in row_candidates)
        candidate_dimensions = cell10_ordered_unique(candidate["Unit Dimension"] for candidate in row_candidates)
        mapping_sources = cell10_ordered_unique(candidate["Mapping Source"] for candidate in row_candidates)
        mapping_states = cell10_ordered_unique(candidate["Mapping State"] for candidate in row_candidates)
        mapping_notes = cell10_ordered_unique(candidate["Mapping Note"] for candidate in row_candidates)
        mapping_source_files = cell10_ordered_unique(candidate["Mapping Source File"] for candidate in row_candidates)
        alias_rows = cell10_ordered_unique(candidate["Unit Alias Workbook Row"] for candidate in row_candidates)
        alias_notes = cell10_ordered_unique(candidate["Alias Note"] for candidate in row_candidates)
        alias_reviewers = cell10_ordered_unique(candidate["Alias Reviewer Name"] for candidate in row_candidates)
        alias_dates = cell10_ordered_unique(candidate["Alias Review Date"] for candidate in row_candidates)

        if unit_info["status"] == "Recognized":
            unit_validation_status = "Pass"
            review_required = False
            cell10_status = "Pass"
        elif unit_info["status"] == "Not provided":
            # Missing units remain undecided until Cell 11 knows the expected Quantity Class.
            unit_validation_status = "Not assessed"
            review_required = False
            cell10_status = "Not assessed"
        else:
            unit_validation_status = "Review required"
            review_required = True
            cell10_status = "Review required"

        validation_rows.append({
            "Source Row Order": source_row_order,
            "Source Index": source_index,
            "Original Service Company": original_values[0],
            "Original Mnemonic": original_values[1],
            "Original Unit": original_values[2],
            "Original Description": original_values[3],
            "Prepared Service Company": prepared_values[0],
            "Prepared Mnemonic": prepared_values[1],
            "Prepared Unit": prepared_values[2],
            "Prepared Description": prepared_values[3],
            "Unit Recognition Status": unit_info["status"],
            "Unit Match Method": unit_info["method"],
            "Unit Raw Candidate Count": len(row_candidates),
            "Distinct Candidate Canonical Unit Count": len(candidate_units),
            "Canonical Unit": canonical_unit,
            "Unit Name": details["name"],
            "Unit Dimension": details["dimension"],
            "Unit Is SI": details["is_si"],
            "Unit Category": details["category"],
            "Unit Base Unit": details["base_unit"],
            "Unit Conversion Reference": details["conversion_reference"],
            "Unit Definition Is Exact": details["is_exact"],
            "Conversion A": details["a"],
            "Conversion B": details["b"],
            "Conversion C": details["c"],
            "Conversion D": details["d"],
            "Unit Underlying Definition": details["underlying_definition"],
            "Unit Description": details["description"],
            "Unit Is Base": details["is_base"],
            "Quantity Class Membership Row Count": details["membership_row_count"],
            "Distinct Quantity Class Membership Count": details["membership_count"],
            "Actual Quantity Class Memberships": details["memberships"],
            "Quantity Class Membership Reference Positions": details["membership_positions"],
            "Unit Code Record Count": details["code_row_count"],
            "Unit Integer Codes": details["codes"],
            "Unit Code Deprecated Values": details["code_deprecated_values"],
            "Candidate Canonical Units": candidate_units,
            "Candidate Unit Names": candidate_names,
            "Candidate Unit Dimensions": candidate_dimensions,
            "Official Mapping Sources": mapping_sources,
            "Official Mapping States": mapping_states,
            "Official Mapping Notes": mapping_notes,
            "Unit Alias Workbook Rows": alias_rows,
            "Unit Alias Notes": alias_notes,
            "Unit Alias Reviewers": alias_reviewers,
            "Unit Alias Review Dates": alias_dates,
            "Unit Validation Status": unit_validation_status,
            "Review Required": review_required,
            "Cell 10 Status": cell10_status,
            "Review Reason": "; ".join(unit_info["review_reasons"]),
            "UOM Unit Source Files": details["unit_source_files"],
            "UOM Membership Source Files": details["membership_source_files"],
            "UOM Mapping Source Files": mapping_source_files,
            "UOM Source Commit": UOM_SOURCE["commit"],
            "Alias Workbook Path": str(ALIASES_WORKBOOK_PATH),
        })

    candidates_df = pd.DataFrame(candidate_rows, columns=UNIT_CANDIDATE_COLUMNS)
    candidates_df.index = pd.RangeIndex(start=1, stop=len(candidates_df) + 1, name="unit_candidate_order")
    validation_df = pd.DataFrame(validation_rows, columns=UNIT_VALIDATION_COLUMNS)
    validation_df.index = metadata_input_preprocessing_df.index.copy()
    review_df = validation_df.loc[validation_df["Review Required"].eq(True)].copy()
    review_df.index = pd.RangeIndex(start=1, stop=len(review_df) + 1, name="review_order")

    recognized_rows = validation_df["Unit Recognition Status"].eq("Recognized")
    stats = {
        "Input rows": len(validation_df),
        "Units not provided": int(validation_df["Unit Recognition Status"].eq("Not provided").sum()),
        "Units recognized by exact UOM symbol": int(((validation_df["Unit Recognition Status"] == "Recognized") & (validation_df["Unit Match Method"] == UNIT_METHOD_EXACT)).sum()),
        "Units recognized by automatic official mapping": int(((validation_df["Unit Recognition Status"] == "Recognized") & (validation_df["Unit Match Method"] == UNIT_METHOD_MAPPING)).sum()),
        "Units recognized by approved alias": int(((validation_df["Unit Recognition Status"] == "Recognized") & (validation_df["Unit Match Method"] == UNIT_METHOD_ALIAS)).sum()),
        "Units with case-insensitive candidates": int(validation_df["Unit Match Method"].eq(UNIT_METHOD_CASE).sum()),
        "Ambiguous unit resolutions": int(validation_df["Unit Recognition Status"].eq("Ambiguous").sum()),
        "Units not recognized": int(validation_df["Unit Recognition Status"].eq("Not recognized").sum()),
        "Rows with multiple Quantity Class memberships": int((recognized_rows & validation_df["Distinct Quantity Class Membership Count"].gt(1)).sum()),
        "Recognized rows without Quantity Class memberships": int((recognized_rows & validation_df["Distinct Quantity Class Membership Count"].eq(0)).sum()),
        "Official mapping candidate rows requiring review": int(sum(candidate["Search Stage"] == "Official mapping" and candidate["Candidate Resolution Status"] != "Resolved" for candidate in candidate_rows)),
        "Unit alias target issue rows": int(sum(candidate["Search Stage"] == "Expert Review Workbook" and candidate["Candidate Resolution Status"] != "Resolved" for candidate in candidate_rows)),
        "Cell 10 passes": int(validation_df["Cell 10 Status"].eq("Pass").sum()),
        "Cell 10 assessments deferred": int(validation_df["Cell 10 Status"].eq("Not assessed").sum()),
        "Cell 10 reviews required": int(validation_df["Cell 10 Status"].eq("Review required").sum()),
    }

    return candidates_df, validation_df, review_df, stats


def cell10_verify_outputs(upstream_snapshots):
    """Confirm that Cell 10 preserved every source object and every metadata row."""

    for object_name, snapshot in upstream_snapshots.items():
        pd.testing.assert_frame_equal(globals()[object_name], snapshot, check_dtype=True, check_exact=True)

    if len(UNIT_VALIDATION_DF) != len(METADATA_INPUT_PREPROCESSING_DF):
        raise AssertionError("UNIT_VALIDATION_DF does not contain exactly one row per input row.")

    if not UNIT_VALIDATION_DF.index.equals(METADATA_INPUT_PREPROCESSING_DF.index):
        raise AssertionError("UNIT_VALIDATION_DF did not preserve the Input Data Preprocessing input index.")

    expected_source_orders = list(range(1, len(METADATA_INPUT_PREPROCESSING_DF) + 1))

    if UNIT_VALIDATION_DF["Source Row Order"].tolist() != expected_source_orders:
        raise AssertionError("Source Row Order is not the complete one-based positional input sequence.")

    if not UNIT_CANDIDATES_DF.empty:
        valid_source_orders = set(expected_source_orders)

        if not set(UNIT_CANDIDATES_DF["Source Row Order"]).issubset(valid_source_orders):
            raise AssertionError("UNIT_CANDIDATES_DF contains an invalid Source Row Order.")

    candidate_counts = UNIT_CANDIDATES_DF.groupby("Source Row Order", sort=False).size().to_dict() if not UNIT_CANDIDATES_DF.empty else {}

    for _, validation_row in UNIT_VALIDATION_DF.iterrows():
        source_row_order = validation_row["Source Row Order"]

        if validation_row["Unit Raw Candidate Count"] != candidate_counts.get(source_row_order, 0):
            raise AssertionError(f"Unit candidate count mismatch for Source Row Order {source_row_order}.")

        if validation_row["Unit Validation Status"] == "Pass" and validation_row["Unit Recognition Status"] != "Recognized":
            raise AssertionError(f"A non-recognized unit passed at Source Row Order {source_row_order}.")

        if validation_row["Unit Recognition Status"] == "Recognized" and not cell10_stable_text(validation_row["Canonical Unit"]):
            raise AssertionError(f"A recognized unit lacks a canonical symbol at Source Row Order {source_row_order}.")

    expected_review_orders = UNIT_VALIDATION_DF.loc[UNIT_VALIDATION_DF["Review Required"].eq(True), "Source Row Order"].tolist()

    if UNIT_REVIEW_DF["Source Row Order"].tolist() != expected_review_orders:
        raise AssertionError("UNIT_REVIEW_DF does not exactly represent the rows marked for review.")

    forbidden_cell11_columns = {"Expected Quantity Class", "Unit Compatibility Status", "Compatibility Status"}

    if forbidden_cell11_columns.intersection(UNIT_VALIDATION_DF.columns):
        raise AssertionError("Cell 10 contains fields reserved for Cell 11 compatibility validation.")


try:
    cell10_validate_inputs()

    # Deep snapshots make the non-destructive behaviour auditable.
    _CELL10_UPSTREAM_SNAPSHOTS = {
        object_name: globals()[object_name].copy(deep=True)
        for object_name in [
            "METADATA_DF",
            "METADATA_ORIGINAL_DF",
            "METADATA_INPUT_PREPROCESSING_DF",
            "UOM_UNITS_DF",
            "UOM_QUANTITY_CLASS_MEMBERS_DF",
            "UOM_UNIT_MAPPINGS_DF",
            "UOM_UNIT_CODES_DF",
            "UNIT_ALIASES_DF",
        ]
    }
    _CELL10_INDEXES = cell10_build_indexes()

    UNIT_CANDIDATES_DF, UNIT_VALIDATION_DF, UNIT_REVIEW_DF, UNIT_VALIDATION_STATS = cell10_apply_validation(
        METADATA_ORIGINAL_DF,
        METADATA_INPUT_PREPROCESSING_DF,
        _CELL10_INDEXES,
    )

    cell10_verify_outputs(_CELL10_UPSTREAM_SNAPSHOTS)

    print("Cell 10 unit validation completed successfully.")

    for statistic_name, statistic_value in UNIT_VALIDATION_STATS.items():
        print(f"{statistic_name}: {statistic_value:,}")

    print("\nUnit validation preview:")
    display(UNIT_VALIDATION_DF.head(10))

    if not UNIT_CANDIDATES_DF.empty:
        print("\nUnit candidate provenance preview:")
        display(UNIT_CANDIDATES_DF.head(20))

    if UNIT_REVIEW_DF.empty:
        print("\nNo supplied units require expert review in Cell 10.")
    else:
        print("\nUnits requiring expert review:")
        display(UNIT_REVIEW_DF.head(20))

except Exception as error:
    # Reset every output so Cell 11 cannot consume partial unit-validation results.
    UNIT_CANDIDATES_DF = pd.DataFrame(columns=UNIT_CANDIDATE_COLUMNS)
    UNIT_CANDIDATES_DF.index = pd.RangeIndex(start=1, stop=1, name="unit_candidate_order")
    UNIT_VALIDATION_DF = pd.DataFrame(columns=UNIT_VALIDATION_COLUMNS)
    UNIT_REVIEW_DF = pd.DataFrame(columns=UNIT_VALIDATION_COLUMNS)
    UNIT_REVIEW_DF.index = pd.RangeIndex(start=1, stop=1, name="review_order")
    UNIT_VALIDATION_STATS = {}

    print("Cell 10 failed.")
    print(f"Error: {error}")


# In[ ]:


# ==============================================================================
# CELL 10+ — MULTIPLE QUANTITY CLASS MEMBERSHIPS UNIT COMPATIBILITY CHECK
# ==============================================================================

# Requires pandas as pd and display from Cell 2.
# Uses the final mnemonic-side result from Cell 9 and unit-side result from Cell 10.
# This supplementary cell evaluates only recognized units that belong to more than one
# Quantity Class. It never changes Cell 9 or Cell 10 outputs.

CELL10PLUS_MNEMONIC_REQUIRED_COLUMNS = [
    "Source Row Order",
    "Source Index",
    "Prepared Service Company",
    "Prepared Mnemonic",
    "Prepared Unit",
    "Property Kind UUID",
    "Property Kind Title",
    "Expected Quantity Class",
    "Mnemonic Resolution Status",
    "Mnemonic Validation Status",
    "Cell 9 Status",
    "Review Reason",
    "PWLS Curve Source Files",
    "PWLS Curve Source Commit",
    "Property Kind Source Files",
    "Property Kind Source Commit",
]

CELL10PLUS_UNIT_REQUIRED_COLUMNS = [
    "Source Row Order",
    "Source Index",
    "Prepared Unit",
    "Unit Recognition Status",
    "Unit Match Method",
    "Canonical Unit",
    "Unit Name",
    "Unit Dimension",
    "Distinct Quantity Class Membership Count",
    "Actual Quantity Class Memberships",
    "Unit Validation Status",
    "Cell 10 Status",
    "Review Reason",
    "UOM Unit Source Files",
    "UOM Membership Source Files",
    "UOM Source Commit",
]

UNIT_MULTI_CLASS_VALIDATION_COLUMNS = [
    "Source Row Order",
    "Source Index",
    "Prepared Service Company",
    "Prepared Mnemonic",
    "Prepared Unit",
    "Canonical Unit",
    "Unit Name",
    "Unit Dimension",
    "Property Kind UUID",
    "Property Kind Title",
    "Expected Quantity Class",
    "Actual Quantity Class Memberships",
    "Distinct Quantity Class Membership Count",
    "Expected Class Membership Found",
    "Matching Quantity Class",
    "Unit Compatibility Status",
    "Cell 10+ Status",
    "Review Required",
    "Review Reason",
    "Mnemonic Resolution Status",
    "Mnemonic Validation Status",
    "Cell 9 Status",
    "Cell 9 Review Reason",
    "Unit Recognition Status",
    "Unit Match Method",
    "Unit Validation Status",
    "Cell 10 Status",
    "Cell 10 Review Reason",
    "PWLS Curve Source Files",
    "PWLS Curve Source Commit",
    "Property Kind Source Files",
    "Property Kind Source Commit",
    "UOM Unit Source Files",
    "UOM Membership Source Files",
    "UOM Source Commit",
]

# Initialize every output before execution so rerunning the cell cannot expose stale results.
UNIT_MULTI_CLASS_VALIDATION_DF = pd.DataFrame(columns=UNIT_MULTI_CLASS_VALIDATION_COLUMNS)
UNIT_MULTI_CLASS_REVIEW_DF = pd.DataFrame(columns=UNIT_MULTI_CLASS_VALIDATION_COLUMNS)
UNIT_MULTI_CLASS_STATS = {}


def cell10plus_stable_text(value):
    """Return a stable text value without changing capitalization or spacing."""

    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass

    return str(value)


def cell10plus_ordered_unique(values):
    """Return distinct non-empty text values in their original order."""

    ordered_values = []
    seen_values = set()

    for value in values:
        text_value = cell10plus_stable_text(value)

        if text_value and text_value not in seen_values:
            seen_values.add(text_value)
            ordered_values.append(text_value)

    return tuple(ordered_values)


def cell10plus_membership_tuple(value):
    """Read the membership collection created by Cell 10 without parsing or normalizing it."""

    if isinstance(value, (tuple, list)):
        return cell10plus_ordered_unique(value)

    if value is None:
        return ()

    try:
        if pd.isna(value):
            return ()
    except (TypeError, ValueError):
        pass

    text_value = cell10plus_stable_text(value)
    return (text_value,) if text_value else ()


def cell10plus_numeric_count(value, fallback):
    """Read a stored membership count and fall back to the preserved membership collection."""

    try:
        if pd.isna(value):
            return fallback
    except (TypeError, ValueError):
        pass

    try:
        return int(value)
    except (TypeError, ValueError):
        return fallback


def cell10plus_values_equal(left_value, right_value):
    """Compare two source identifiers safely, including missing index labels."""

    try:
        if pd.isna(left_value) and pd.isna(right_value):
            return True
    except (TypeError, ValueError):
        pass

    try:
        comparison = left_value == right_value
        return bool(comparison)
    except (TypeError, ValueError):
        return False


def cell10plus_require_dataframe(object_name, required_columns):
    """Validate one upstream DataFrame and report its actual schema clearly."""

    if object_name not in globals() or globals()[object_name] is None:
        raise NameError(f"{object_name} is unavailable.")

    dataframe = globals()[object_name]

    if not isinstance(dataframe, pd.DataFrame):
        raise TypeError(f"{object_name} must be a pandas DataFrame.")

    actual_columns = list(dataframe.columns)
    missing_columns = [column for column in required_columns if column not in actual_columns]

    if missing_columns:
        raise ValueError(
            f"{object_name} is missing required columns. "
            f"Expected at least: {required_columns}. Received: {actual_columns}"
        )


def cell10plus_validate_inputs():
    """Verify that Cell 9 and Cell 10 provide aligned one-row-per-source results."""

    if "MNEMONIC_VALIDATION_DF" not in globals() or MNEMONIC_VALIDATION_DF is None:
        raise NameError("MNEMONIC_VALIDATION_DF is unavailable. Run Cell 9 before running Cell 10+.")

    if "UNIT_VALIDATION_DF" not in globals() or UNIT_VALIDATION_DF is None:
        raise NameError("UNIT_VALIDATION_DF is unavailable. Run Cell 10 before running Cell 10+.")

    cell10plus_require_dataframe("MNEMONIC_VALIDATION_DF", CELL10PLUS_MNEMONIC_REQUIRED_COLUMNS)
    cell10plus_require_dataframe("UNIT_VALIDATION_DF", CELL10PLUS_UNIT_REQUIRED_COLUMNS)

    if MNEMONIC_VALIDATION_DF.empty and not UNIT_VALIDATION_DF.empty:
        raise ValueError("Cell 9 and Cell 10 row counts do not agree.")

    if len(MNEMONIC_VALIDATION_DF) != len(UNIT_VALIDATION_DF):
        raise ValueError(
            "Cell 9 and Cell 10 row counts do not agree. "
            f"Cell 9: {len(MNEMONIC_VALIDATION_DF)}; Cell 10: {len(UNIT_VALIDATION_DF)}."
        )

    if not MNEMONIC_VALIDATION_DF.index.equals(UNIT_VALIDATION_DF.index):
        raise ValueError("Cell 9 and Cell 10 result indices do not agree.")

    expected_source_orders = list(range(1, len(UNIT_VALIDATION_DF) + 1))

    if MNEMONIC_VALIDATION_DF["Source Row Order"].tolist() != expected_source_orders:
        raise ValueError("Cell 9 Source Row Order is not the complete one-based input sequence.")

    if UNIT_VALIDATION_DF["Source Row Order"].tolist() != expected_source_orders:
        raise ValueError("Cell 10 Source Row Order is not the complete one-based input sequence.")

    for source_position in range(len(UNIT_VALIDATION_DF)):
        mnemonic_source_index = MNEMONIC_VALIDATION_DF.iloc[source_position]["Source Index"]
        unit_source_index = UNIT_VALIDATION_DF.iloc[source_position]["Source Index"]

        if not cell10plus_values_equal(mnemonic_source_index, unit_source_index):
            raise ValueError(f"Cell 9 and Cell 10 Source Index values disagree at Source Row Order {source_position + 1}.")

        mnemonic_prepared_unit = cell10plus_stable_text(MNEMONIC_VALIDATION_DF.iloc[source_position]["Prepared Unit"])
        unit_prepared_unit = cell10plus_stable_text(UNIT_VALIDATION_DF.iloc[source_position]["Prepared Unit"])

        if mnemonic_prepared_unit != unit_prepared_unit:
            raise ValueError(f"Cell 9 and Cell 10 Prepared Unit values disagree at Source Row Order {source_position + 1}.")

    if "UNIT_VALIDATION_STATS" not in globals() or not isinstance(UNIT_VALIDATION_STATS, dict):
        raise NameError("UNIT_VALIDATION_STATS is unavailable. Run Cell 10 before running Cell 10+.")


def cell10plus_apply_check(mnemonic_validation_df, unit_validation_df):
    """Evaluate the expected class against every recognized multi-membership unit."""

    validation_rows = []

    for source_position in range(len(unit_validation_df)):
        mnemonic_row = mnemonic_validation_df.iloc[source_position]
        unit_row = unit_validation_df.iloc[source_position]
        memberships = cell10plus_membership_tuple(unit_row["Actual Quantity Class Memberships"])
        membership_count = cell10plus_numeric_count(
            unit_row["Distinct Quantity Class Membership Count"],
            len(memberships),
        )
        unit_recognized = cell10plus_stable_text(unit_row["Unit Recognition Status"]) == "Recognized"

        # Reproduce the exact population counted by Cell 10's multi-membership statistic.
        if not unit_recognized or membership_count <= 1:
            continue

        expected_class = cell10plus_stable_text(mnemonic_row["Expected Quantity Class"])

        if not expected_class:
            membership_found = ""
            matching_class = ""
            compatibility_status = "Not assessed"
            cell10plus_status = "Review required"
            review_required = True
            review_reason = "Cell 9 did not provide one resolved expected Quantity Class"

        elif not memberships:
            membership_found = ""
            matching_class = ""
            compatibility_status = "Not assessed"
            cell10plus_status = "Review required"
            review_required = True
            review_reason = "Cell 10 did not preserve the unit's Quantity Class memberships"

        elif expected_class in memberships:
            # Exact, case-sensitive membership is sufficient even when the unit has other valid classes.
            membership_found = True
            matching_class = expected_class
            compatibility_status = "Compatible"
            cell10plus_status = "Pass"
            review_required = False
            review_reason = ""

        else:
            membership_found = False
            matching_class = ""
            compatibility_status = "Incompatible"
            cell10plus_status = "Review required"
            review_required = True
            review_reason = "Expected Quantity Class is not among the recognized unit's memberships"

        validation_rows.append({
            "Source Row Order": unit_row["Source Row Order"],
            "Source Index": unit_row["Source Index"],
            "Prepared Service Company": mnemonic_row["Prepared Service Company"],
            "Prepared Mnemonic": mnemonic_row["Prepared Mnemonic"],
            "Prepared Unit": unit_row["Prepared Unit"],
            "Canonical Unit": unit_row["Canonical Unit"],
            "Unit Name": unit_row["Unit Name"],
            "Unit Dimension": unit_row["Unit Dimension"],
            "Property Kind UUID": mnemonic_row["Property Kind UUID"],
            "Property Kind Title": mnemonic_row["Property Kind Title"],
            "Expected Quantity Class": mnemonic_row["Expected Quantity Class"],
            "Actual Quantity Class Memberships": memberships,
            "Distinct Quantity Class Membership Count": membership_count,
            "Expected Class Membership Found": membership_found,
            "Matching Quantity Class": matching_class,
            "Unit Compatibility Status": compatibility_status,
            "Cell 10+ Status": cell10plus_status,
            "Review Required": review_required,
            "Review Reason": review_reason,
            "Mnemonic Resolution Status": mnemonic_row["Mnemonic Resolution Status"],
            "Mnemonic Validation Status": mnemonic_row["Mnemonic Validation Status"],
            "Cell 9 Status": mnemonic_row["Cell 9 Status"],
            "Cell 9 Review Reason": mnemonic_row["Review Reason"],
            "Unit Recognition Status": unit_row["Unit Recognition Status"],
            "Unit Match Method": unit_row["Unit Match Method"],
            "Unit Validation Status": unit_row["Unit Validation Status"],
            "Cell 10 Status": unit_row["Cell 10 Status"],
            "Cell 10 Review Reason": unit_row["Review Reason"],
            "PWLS Curve Source Files": mnemonic_row["PWLS Curve Source Files"],
            "PWLS Curve Source Commit": mnemonic_row["PWLS Curve Source Commit"],
            "Property Kind Source Files": mnemonic_row["Property Kind Source Files"],
            "Property Kind Source Commit": mnemonic_row["Property Kind Source Commit"],
            "UOM Unit Source Files": unit_row["UOM Unit Source Files"],
            "UOM Membership Source Files": unit_row["UOM Membership Source Files"],
            "UOM Source Commit": unit_row["UOM Source Commit"],
        })

    validation_df = pd.DataFrame(validation_rows, columns=UNIT_MULTI_CLASS_VALIDATION_COLUMNS)
    validation_df.index = pd.RangeIndex(start=1, stop=len(validation_df) + 1, name="check_order")
    review_df = validation_df.loc[validation_df["Review Required"].eq(True)].copy()
    review_df.index = pd.RangeIndex(start=1, stop=len(review_df) + 1, name="review_order")

    checked_rows = len(validation_df)
    compatibility_passes = int(validation_df["Unit Compatibility Status"].eq("Compatible").sum())
    stats = {
        "Multiple-membership rows checked": checked_rows,
        "Compatibility passes": compatibility_passes,
        "Compatibility incompatibilities": int(validation_df["Unit Compatibility Status"].eq("Incompatible").sum()),
        "Compatibility assessments unavailable": int(validation_df["Unit Compatibility Status"].eq("Not assessed").sum()),
        "Cell 10+ reviews required": int(validation_df["Review Required"].eq(True).sum()),
        "Cell 10+ compatibility pass rate (%)": round((compatibility_passes / checked_rows) * 100, 2) if checked_rows else 0.0,
    }

    return validation_df, review_df, stats


def cell10plus_verify_outputs(upstream_snapshots):
    """Verify exact comparison behaviour and confirm both upstream results remained unchanged."""

    for object_name, snapshot in upstream_snapshots.items():
        pd.testing.assert_frame_equal(globals()[object_name], snapshot, check_dtype=True, check_exact=True)

    expected_multi_membership_rows = 0

    for source_position in range(len(UNIT_VALIDATION_DF)):
        unit_row = UNIT_VALIDATION_DF.iloc[source_position]
        memberships = cell10plus_membership_tuple(unit_row["Actual Quantity Class Memberships"])
        membership_count = cell10plus_numeric_count(unit_row["Distinct Quantity Class Membership Count"], len(memberships))

        if cell10plus_stable_text(unit_row["Unit Recognition Status"]) == "Recognized" and membership_count > 1:
            expected_multi_membership_rows += 1

    reported_multi_membership_rows = UNIT_VALIDATION_STATS.get("Rows with multiple Quantity Class memberships")

    if reported_multi_membership_rows != expected_multi_membership_rows:
        raise AssertionError(
            "Cell 10's reported multiple-membership count does not agree with UNIT_VALIDATION_DF. "
            f"Reported: {reported_multi_membership_rows}; calculated: {expected_multi_membership_rows}."
        )

    if len(UNIT_MULTI_CLASS_VALIDATION_DF) != expected_multi_membership_rows:
        raise AssertionError("Cell 10+ did not check every recognized multiple-membership unit exactly once.")

    expected_source_orders = []

    for source_position in range(len(UNIT_VALIDATION_DF)):
        unit_row = UNIT_VALIDATION_DF.iloc[source_position]
        memberships = cell10plus_membership_tuple(unit_row["Actual Quantity Class Memberships"])
        membership_count = cell10plus_numeric_count(unit_row["Distinct Quantity Class Membership Count"], len(memberships))

        if cell10plus_stable_text(unit_row["Unit Recognition Status"]) == "Recognized" and membership_count > 1:
            expected_source_orders.append(source_position + 1)

    if UNIT_MULTI_CLASS_VALIDATION_DF["Source Row Order"].tolist() != expected_source_orders:
        raise AssertionError("Cell 10+ did not preserve the source-row order of the checked units.")

    for _, result_row in UNIT_MULTI_CLASS_VALIDATION_DF.iterrows():
        expected_class = cell10plus_stable_text(result_row["Expected Quantity Class"])
        memberships = cell10plus_membership_tuple(result_row["Actual Quantity Class Memberships"])

        if result_row["Cell 10+ Status"] == "Pass":
            if not expected_class or expected_class not in memberships:
                raise AssertionError(f"Cell 10+ produced an unsupported pass at Source Row Order {result_row['Source Row Order']}.")

            if result_row["Unit Compatibility Status"] != "Compatible":
                raise AssertionError(f"Cell 10+ pass and compatibility status disagree at Source Row Order {result_row['Source Row Order']}.")

        if result_row["Unit Compatibility Status"] == "Incompatible" and expected_class in memberships:
            raise AssertionError(f"Cell 10+ rejected an exact membership at Source Row Order {result_row['Source Row Order']}.")

    expected_review_orders = UNIT_MULTI_CLASS_VALIDATION_DF.loc[
        UNIT_MULTI_CLASS_VALIDATION_DF["Review Required"].eq(True),
        "Source Row Order",
    ].tolist()

    if UNIT_MULTI_CLASS_REVIEW_DF["Source Row Order"].tolist() != expected_review_orders:
        raise AssertionError("UNIT_MULTI_CLASS_REVIEW_DF does not exactly represent the rows marked for review.")


try:
    cell10plus_validate_inputs()

    _CELL10PLUS_UPSTREAM_SNAPSHOTS = {
        "MNEMONIC_VALIDATION_DF": MNEMONIC_VALIDATION_DF.copy(deep=True),
        "UNIT_VALIDATION_DF": UNIT_VALIDATION_DF.copy(deep=True),
    }

    UNIT_MULTI_CLASS_VALIDATION_DF, UNIT_MULTI_CLASS_REVIEW_DF, UNIT_MULTI_CLASS_STATS = cell10plus_apply_check(
        MNEMONIC_VALIDATION_DF,
        UNIT_VALIDATION_DF,
    )

    cell10plus_verify_outputs(_CELL10PLUS_UPSTREAM_SNAPSHOTS)

    print("Cell 10+ multiple-membership unit compatibility check completed successfully.")

    for statistic_name, statistic_value in UNIT_MULTI_CLASS_STATS.items():
        print(f"{statistic_name}: {statistic_value:,}")

    if UNIT_MULTI_CLASS_VALIDATION_DF.empty:
        print("\nNo recognized units have multiple Quantity Class memberships.")
    else:
        print("\nMultiple-membership unit compatibility results:")
        display(UNIT_MULTI_CLASS_VALIDATION_DF.head(20))

    if UNIT_MULTI_CLASS_REVIEW_DF.empty:
        print("\nAll checked multiple-membership units passed compatibility.")
    else:
        print("\nMultiple-membership units requiring expert review:")
        display(UNIT_MULTI_CLASS_REVIEW_DF.head(20))

except Exception as error:
    # Reset all supplementary outputs so later cells cannot consume partial results.
    UNIT_MULTI_CLASS_VALIDATION_DF = pd.DataFrame(columns=UNIT_MULTI_CLASS_VALIDATION_COLUMNS)
    UNIT_MULTI_CLASS_VALIDATION_DF.index = pd.RangeIndex(start=1, stop=1, name="check_order")
    UNIT_MULTI_CLASS_REVIEW_DF = pd.DataFrame(columns=UNIT_MULTI_CLASS_VALIDATION_COLUMNS)
    UNIT_MULTI_CLASS_REVIEW_DF.index = pd.RangeIndex(start=1, stop=1, name="review_order")
    UNIT_MULTI_CLASS_STATS = {}

    print("Cell 10+ failed.")
    print(f"Error: {error}")


# In[ ]:


# ============================================================
# CELL 11 — SEMANTIC QUANTITY CLASS VALIDATION
# ============================================================

# Requires pandas as pd and display from Cell 2.
# Converges the independent mnemonic evidence from Cell 9 with the unit evidence
# from Cell 10 and the multiple-membership decisions from Cell 10+.
# Only rows that passed every applicable upstream gate receive a Cell 11 comparison.
# Every other row remains preserved and is routed to the final expert-review table.

CELL11_MNEMONIC_REQUIRED_COLUMNS = [
    "Source Row Order",
    "Source Index",
    "Original Service Company",
    "Original Mnemonic",
    "Original Unit",
    "Original Description",
    "Prepared Service Company",
    "Prepared Mnemonic",
    "Prepared Unit",
    "Prepared Description",
    "Company Resolution Status",
    "Company Match Method",
    "Mnemonic Resolution Status",
    "Mnemonic Match Method",
    "Mnemonic Validation Status",
    "Property Kind UUID",
    "Property Kind Title",
    "Expected Quantity Class",
    "Cell 9 Status",
    "Review Reason",
    "PWLS Curve Source Files",
    "PWLS Curve Source Commit",
    "Property Kind Source Files",
    "Property Kind Source Commit",
]

CELL11_UNIT_REQUIRED_COLUMNS = [
    "Source Row Order",
    "Source Index",
    "Prepared Service Company",
    "Prepared Mnemonic",
    "Prepared Unit",
    "Prepared Description",
    "Unit Recognition Status",
    "Unit Match Method",
    "Canonical Unit",
    "Unit Name",
    "Unit Dimension",
    "Quantity Class Membership Row Count",
    "Distinct Quantity Class Membership Count",
    "Actual Quantity Class Memberships",
    "Unit Validation Status",
    "Cell 10 Status",
    "Review Reason",
    "UOM Unit Source Files",
    "UOM Membership Source Files",
    "UOM Source Commit",
]

CELL11_MULTI_REQUIRED_COLUMNS = [
    "Source Row Order",
    "Source Index",
    "Prepared Mnemonic",
    "Prepared Unit",
    "Canonical Unit",
    "Property Kind UUID",
    "Expected Quantity Class",
    "Actual Quantity Class Memberships",
    "Distinct Quantity Class Membership Count",
    "Expected Class Membership Found",
    "Unit Compatibility Status",
    "Cell 10+ Status",
    "Review Required",
    "Review Reason",
]

SEMANTIC_VALIDATION_COLUMNS = [
    "Source Row Order",
    "Source Index",
    "Original Service Company",
    "Original Mnemonic",
    "Original Unit",
    "Original Description",
    "Prepared Service Company",
    "Prepared Mnemonic",
    "Prepared Unit",
    "Prepared Description",
    "Resolved Company Code",
    "Resolved Company Name",
    "Company Resolution Status",
    "Company Match Method",
    "Mnemonic Resolution Status",
    "Mnemonic Match Method",
    "Mnemonic Validation Status",
    "Property Kind UUID",
    "Property Kind Title",
    "Expected Quantity Class",
    "Unit Recognition Status",
    "Unit Match Method",
    "Canonical Unit",
    "Unit Name",
    "Unit Dimension",
    "Quantity Class Membership Row Count",
    "Distinct Quantity Class Membership Count",
    "Actual Quantity Class Memberships",
    "Unit Validation Status",
    "Membership Structure",
    "Cell 9 Gate Passed",
    "Cell 10 Gate Passed",
    "Cell 10+ Check Required",
    "Cell 10+ Gate Passed",
    "Semantic Comparison Eligible",
    "Semantic Comparison Performed",
    "Semantic Result Source",
    "Expected Class Membership Found",
    "Matching Quantity Class",
    "Semantic Compatibility Status",
    "Semantic Validation Status",
    "Cell 11 Status",
    "Final Workflow Status",
    "Expert Review Required",
    "Review Origin",
    "Review Reason",
    "Cell 9 Status",
    "Cell 9 Review Reason",
    "Cell 10 Status",
    "Cell 10 Review Reason",
    "Cell 10+ Status",
    "Cell 10+ Compatibility Status",
    "Cell 10+ Review Reason",
    "PWLS Curve Source Files",
    "PWLS Curve Source Commit",
    "Property Kind Source Files",
    "Property Kind Source Commit",
    "UOM Unit Source Files",
    "UOM Membership Source Files",
    "UOM Source Commit",
]

# Initialize all outputs before execution so rerunning Cell 11 cannot expose stale results.
SEMANTIC_VALIDATION_DF = pd.DataFrame(columns=SEMANTIC_VALIDATION_COLUMNS)
SEMANTIC_PASS_DF = pd.DataFrame(columns=SEMANTIC_VALIDATION_COLUMNS)
SEMANTIC_REVIEW_DF = pd.DataFrame(columns=SEMANTIC_VALIDATION_COLUMNS)
SEMANTIC_VALIDATION_STATS = {}


def cell11_stable_text(value):
    """Return a stable text value without changing capitalization, spaces, or punctuation."""

    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass

    return str(value)


def cell11_ordered_unique(values):
    """Return distinct non-empty textual values in first-seen order."""

    ordered_values = []
    seen_values = set()

    for value in values:
        text_value = cell11_stable_text(value)

        if text_value and text_value not in seen_values:
            seen_values.add(text_value)
            ordered_values.append(text_value)

    return tuple(ordered_values)


def cell11_membership_tuple(value):
    """Read Cell 10's preserved membership collection without normalizing class names."""

    if isinstance(value, (tuple, list)):
        return cell11_ordered_unique(value)

    if value is None:
        return ()

    try:
        if pd.isna(value):
            return ()
    except (TypeError, ValueError):
        pass

    text_value = cell11_stable_text(value)
    return (text_value,) if text_value else ()


def cell11_numeric_count(value, fallback):
    """Read a stored count and fall back to the preserved membership collection."""

    try:
        if pd.isna(value):
            return fallback
    except (TypeError, ValueError):
        pass

    try:
        return int(value)
    except (TypeError, ValueError):
        return fallback


def cell11_values_equal(left_value, right_value):
    """Compare source identifiers safely, including missing index labels."""

    try:
        if pd.isna(left_value) and pd.isna(right_value):
            return True
    except (TypeError, ValueError):
        pass

    try:
        return bool(left_value == right_value)
    except (TypeError, ValueError):
        return False


def cell11_is_true(value):
    """Interpret stored Boolean evidence without treating non-empty text as true."""

    try:
        if pd.isna(value):
            return False
    except (TypeError, ValueError):
        pass

    try:
        return bool(value == True)
    except (TypeError, ValueError):
        return False


def cell11_append_reason(reasons, reason):
    """Append one review reason without creating repeated text."""

    reason_text = cell11_stable_text(reason)

    if reason_text and reason_text not in reasons:
        reasons.append(reason_text)


def cell11_append_origin(origins, origin):
    """Record each stage that caused expert review exactly once."""

    if origin and origin not in origins:
        origins.append(origin)


def cell11_require_dataframe(object_name, required_columns):
    """Validate one upstream DataFrame and report its actual schema clearly."""

    if object_name not in globals() or globals()[object_name] is None:
        raise NameError(f"{object_name} is unavailable.")

    dataframe = globals()[object_name]

    if not isinstance(dataframe, pd.DataFrame):
        raise TypeError(f"{object_name} must be a pandas DataFrame.")

    actual_columns = list(dataframe.columns)
    missing_columns = [column for column in required_columns if column not in actual_columns]

    if missing_columns:
        raise ValueError(
            f"{object_name} is missing required columns. "
            f"Expected at least: {required_columns}. Received: {actual_columns}"
        )


def cell11_build_multi_lookup(multi_validation_df):
    """Index Cell 10+ by immutable Source Row Order rather than duplicate-prone labels."""

    multi_lookup = {}

    for position in range(len(multi_validation_df)):
        source_row_order = multi_validation_df.iloc[position]["Source Row Order"]

        if source_row_order in multi_lookup:
            raise ValueError(f"Cell 10+ contains repeated Source Row Order {source_row_order}.")

        multi_lookup[source_row_order] = position

    return multi_lookup


def cell11_validate_inputs():
    """Verify alignment and cross-cell integrity before semantic convergence."""

    dependency_messages = {
        "MNEMONIC_VALIDATION_DF": "MNEMONIC_VALIDATION_DF is unavailable. Run Cell 9 before running Cell 11.",
        "UNIT_VALIDATION_DF": "UNIT_VALIDATION_DF is unavailable. Run Cell 10 before running Cell 11.",
        "UNIT_MULTI_CLASS_VALIDATION_DF": "UNIT_MULTI_CLASS_VALIDATION_DF is unavailable. Run Cell 10+ before running Cell 11.",
    }

    for object_name, error_message in dependency_messages.items():
        if object_name not in globals() or globals()[object_name] is None:
            raise NameError(error_message)

    cell11_require_dataframe("MNEMONIC_VALIDATION_DF", CELL11_MNEMONIC_REQUIRED_COLUMNS)
    cell11_require_dataframe("UNIT_VALIDATION_DF", CELL11_UNIT_REQUIRED_COLUMNS)
    cell11_require_dataframe("UNIT_MULTI_CLASS_VALIDATION_DF", CELL11_MULTI_REQUIRED_COLUMNS)

    if len(MNEMONIC_VALIDATION_DF) != len(UNIT_VALIDATION_DF):
        raise ValueError(
            "Cell 9 and Cell 10 row counts do not agree. "
            f"Cell 9: {len(MNEMONIC_VALIDATION_DF)}; Cell 10: {len(UNIT_VALIDATION_DF)}."
        )

    if not MNEMONIC_VALIDATION_DF.index.equals(UNIT_VALIDATION_DF.index):
        raise ValueError("Cell 9 and Cell 10 result indices do not agree.")

    expected_source_orders = list(range(1, len(UNIT_VALIDATION_DF) + 1))

    if MNEMONIC_VALIDATION_DF["Source Row Order"].tolist() != expected_source_orders:
        raise ValueError("Cell 9 Source Row Order is not the complete one-based input sequence.")

    if UNIT_VALIDATION_DF["Source Row Order"].tolist() != expected_source_orders:
        raise ValueError("Cell 10 Source Row Order is not the complete one-based input sequence.")

    multi_lookup = cell11_build_multi_lookup(UNIT_MULTI_CLASS_VALIDATION_DF)
    expected_multi_source_orders = []

    for source_position in range(len(UNIT_VALIDATION_DF)):
        mnemonic_row = MNEMONIC_VALIDATION_DF.iloc[source_position]
        unit_row = UNIT_VALIDATION_DF.iloc[source_position]
        source_row_order = source_position + 1

        if not cell11_values_equal(mnemonic_row["Source Index"], unit_row["Source Index"]):
            raise ValueError(f"Cell 9 and Cell 10 Source Index values disagree at Source Row Order {source_row_order}.")

        for column_name in ["Prepared Service Company", "Prepared Mnemonic", "Prepared Unit", "Prepared Description"]:
            mnemonic_value = cell11_stable_text(mnemonic_row[column_name])
            unit_value = cell11_stable_text(unit_row[column_name])

            if mnemonic_value != unit_value:
                raise ValueError(f"Cell 9 and Cell 10 {column_name} values disagree at Source Row Order {source_row_order}.")

        memberships = cell11_membership_tuple(unit_row["Actual Quantity Class Memberships"])
        stored_count = cell11_numeric_count(unit_row["Distinct Quantity Class Membership Count"], len(memberships))

        if stored_count != len(memberships):
            raise ValueError(f"Cell 10 membership count and membership values disagree at Source Row Order {source_row_order}.")

        if cell11_stable_text(unit_row["Unit Recognition Status"]) == "Recognized" and stored_count > 1:
            expected_multi_source_orders.append(source_row_order)

    if sorted(multi_lookup) != expected_multi_source_orders:
        raise ValueError(
            "Cell 10+ rows do not exactly represent Cell 10's recognized multiple-membership units. "
            f"Expected Source Row Orders: {expected_multi_source_orders}; received: {sorted(multi_lookup)}."
        )

    for source_row_order, multi_position in multi_lookup.items():
        source_position = source_row_order - 1
        mnemonic_row = MNEMONIC_VALIDATION_DF.iloc[source_position]
        unit_row = UNIT_VALIDATION_DF.iloc[source_position]
        multi_row = UNIT_MULTI_CLASS_VALIDATION_DF.iloc[multi_position]

        if not cell11_values_equal(unit_row["Source Index"], multi_row["Source Index"]):
            raise ValueError(f"Cell 10 and Cell 10+ Source Index values disagree at Source Row Order {source_row_order}.")

        if cell11_stable_text(mnemonic_row["Expected Quantity Class"]) != cell11_stable_text(multi_row["Expected Quantity Class"]):
            raise ValueError(f"Cell 9 and Cell 10+ Expected Quantity Class values disagree at Source Row Order {source_row_order}.")

        if cell11_membership_tuple(unit_row["Actual Quantity Class Memberships"]) != cell11_membership_tuple(multi_row["Actual Quantity Class Memberships"]):
            raise ValueError(f"Cell 10 and Cell 10+ Quantity Class memberships disagree at Source Row Order {source_row_order}.")

        if cell11_stable_text(unit_row["Canonical Unit"]) != cell11_stable_text(multi_row["Canonical Unit"]):
            raise ValueError(f"Cell 10 and Cell 10+ Canonical Unit values disagree at Source Row Order {source_row_order}.")

    if "UNIT_VALIDATION_STATS" not in globals() or not isinstance(UNIT_VALIDATION_STATS, dict):
        raise NameError("UNIT_VALIDATION_STATS is unavailable. Run Cell 10 before running Cell 11.")

    if "UNIT_MULTI_CLASS_STATS" not in globals() or not isinstance(UNIT_MULTI_CLASS_STATS, dict):
        raise NameError("UNIT_MULTI_CLASS_STATS is unavailable. Run Cell 10+ before running Cell 11.")

    reported_multi_rows = UNIT_VALIDATION_STATS.get("Rows with multiple Quantity Class memberships")
    checked_multi_rows = UNIT_MULTI_CLASS_STATS.get("Multiple-membership rows checked")

    if reported_multi_rows != len(expected_multi_source_orders) or checked_multi_rows != len(expected_multi_source_orders):
        raise ValueError(
            "The Cell 10 and Cell 10+ multiple-membership statistics do not agree with their result tables."
        )

    return multi_lookup


def cell11_apply_validation(mnemonic_validation_df, unit_validation_df, multi_validation_df, multi_lookup):
    """Apply gated semantic comparison and preserve all upstream review decisions."""

    validation_rows = []

    for source_position in range(len(unit_validation_df)):
        source_row_order = source_position + 1
        mnemonic_row = mnemonic_validation_df.iloc[source_position]
        unit_row = unit_validation_df.iloc[source_position]
        expected_class = cell11_stable_text(mnemonic_row["Expected Quantity Class"])
        memberships = cell11_membership_tuple(unit_row["Actual Quantity Class Memberships"])
        membership_count = cell11_numeric_count(unit_row["Distinct Quantity Class Membership Count"], len(memberships))

        if membership_count == 0:
            membership_structure = "No memberships"
        elif membership_count == 1:
            membership_structure = "Single membership"
        else:
            membership_structure = "Multiple memberships"

        property_uuid = cell11_stable_text(mnemonic_row["Property Kind UUID"])
        canonical_unit = cell11_stable_text(unit_row["Canonical Unit"])

        # The gates require both the cell-level pass and the specific resolved evidence.
        cell9_gate_passed = (
            cell11_stable_text(mnemonic_row["Cell 9 Status"]) == "Pass"
            and cell11_stable_text(mnemonic_row["Mnemonic Validation Status"]) == "Pass"
            and cell11_stable_text(mnemonic_row["Mnemonic Resolution Status"]) == "Resolved"
            and bool(property_uuid)
            and bool(expected_class)
        )
        cell10_gate_passed = (
            cell11_stable_text(unit_row["Cell 10 Status"]) == "Pass"
            and cell11_stable_text(unit_row["Unit Validation Status"]) == "Pass"
            and cell11_stable_text(unit_row["Unit Recognition Status"]) == "Recognized"
            and bool(canonical_unit)
        )

        multi_check_required = cell10_gate_passed and membership_count > 1
        multi_position = multi_lookup.get(source_row_order)
        multi_row = multi_validation_df.iloc[multi_position] if multi_position is not None else None

        if multi_check_required and multi_row is not None:
            multi_status = cell11_stable_text(multi_row["Cell 10+ Status"])
            multi_compatibility = cell11_stable_text(multi_row["Unit Compatibility Status"])
            multi_review_reason = cell11_stable_text(multi_row["Review Reason"])
            multi_gate_passed = (
                multi_status == "Pass"
                and multi_compatibility == "Compatible"
                and cell11_is_true(multi_row["Expected Class Membership Found"])
            )
        elif multi_check_required:
            multi_status = "Unavailable"
            multi_compatibility = "Not assessed"
            multi_review_reason = "Cell 10+ result is missing"
            multi_gate_passed = False
        else:
            multi_status = "Not required"
            multi_compatibility = "Not required"
            multi_review_reason = ""
            multi_gate_passed = ""

        semantic_eligible = (
            cell9_gate_passed
            and cell10_gate_passed
            and membership_count >= 1
            and (not multi_check_required or cell11_is_true(multi_gate_passed))
        )

        review_reasons = []
        review_origins = []

        if not cell9_gate_passed:
            cell11_append_origin(review_origins, "Cell 9")
            cell11_append_reason(
                review_reasons,
                mnemonic_row["Review Reason"] or "Cell 9 did not provide a passing resolved mnemonic and expected Quantity Class",
            )

        if not cell10_gate_passed:
            cell11_append_origin(review_origins, "Cell 10")
            cell11_append_reason(
                review_reasons,
                unit_row["Review Reason"] or "Cell 10 did not provide a passing recognized canonical unit",
            )

        if cell10_gate_passed and membership_count == 0:
            cell11_append_origin(review_origins, "Cell 11")
            cell11_append_reason(review_reasons, "The recognized canonical unit has no Quantity Class membership")

        if multi_check_required and not cell11_is_true(multi_gate_passed):
            cell11_append_origin(review_origins, "Cell 10+")
            cell11_append_reason(
                review_reasons,
                multi_review_reason or "Cell 10+ did not pass the multiple-membership compatibility check",
            )

        if semantic_eligible:
            semantic_comparison_performed = True
            expected_membership_found = expected_class in memberships

            if expected_membership_found:
                matching_class = expected_class
                compatibility_status = "Compatible"
                semantic_validation_status = "Pass"
                cell11_status = "Pass"
                final_workflow_status = "Pass"
                expert_review_required = False
                semantic_result_source = "Cell 11 exact membership comparison"
            else:
                matching_class = ""
                compatibility_status = "Incompatible"
                semantic_validation_status = "Review required"
                cell11_status = "Review required"
                final_workflow_status = "Expert review required"
                expert_review_required = True
                semantic_result_source = "Cell 11 exact membership comparison"
                cell11_append_origin(review_origins, "Cell 11")
                cell11_append_reason(review_reasons, "Expected Quantity Class is not among the canonical unit's memberships")

        else:
            semantic_comparison_performed = False
            expected_membership_found = ""
            matching_class = ""
            cell11_status = "Review required"
            final_workflow_status = "Expert review required"
            expert_review_required = True

            if multi_check_required and multi_compatibility == "Incompatible":
                compatibility_status = "Incompatible"
                semantic_validation_status = "Review required"
                semantic_result_source = "Cell 10+ multiple-membership check"
            else:
                compatibility_status = "Not assessed"
                semantic_validation_status = "Not assessed"
                semantic_result_source = "Upstream review routing"

        validation_rows.append({
            "Source Row Order": source_row_order,
            "Source Index": unit_row["Source Index"],
            "Original Service Company": mnemonic_row["Original Service Company"],
            "Original Mnemonic": mnemonic_row["Original Mnemonic"],
            "Original Unit": mnemonic_row["Original Unit"],
            "Original Description": mnemonic_row["Original Description"],
            "Prepared Service Company": mnemonic_row["Prepared Service Company"],
            "Prepared Mnemonic": mnemonic_row["Prepared Mnemonic"],
            "Prepared Unit": mnemonic_row["Prepared Unit"],
            "Prepared Description": mnemonic_row["Prepared Description"],
            "Resolved Company Code": mnemonic_row["Resolved Company Code"] if "Resolved Company Code" in mnemonic_validation_df.columns else "",
            "Resolved Company Name": mnemonic_row["Resolved Company Name"] if "Resolved Company Name" in mnemonic_validation_df.columns else "",
            "Company Resolution Status": mnemonic_row["Company Resolution Status"],
            "Company Match Method": mnemonic_row["Company Match Method"],
            "Mnemonic Resolution Status": mnemonic_row["Mnemonic Resolution Status"],
            "Mnemonic Match Method": mnemonic_row["Mnemonic Match Method"],
            "Mnemonic Validation Status": mnemonic_row["Mnemonic Validation Status"],
            "Property Kind UUID": mnemonic_row["Property Kind UUID"],
            "Property Kind Title": mnemonic_row["Property Kind Title"],
            "Expected Quantity Class": mnemonic_row["Expected Quantity Class"],
            "Unit Recognition Status": unit_row["Unit Recognition Status"],
            "Unit Match Method": unit_row["Unit Match Method"],
            "Canonical Unit": unit_row["Canonical Unit"],
            "Unit Name": unit_row["Unit Name"],
            "Unit Dimension": unit_row["Unit Dimension"],
            "Quantity Class Membership Row Count": unit_row["Quantity Class Membership Row Count"],
            "Distinct Quantity Class Membership Count": membership_count,
            "Actual Quantity Class Memberships": memberships,
            "Unit Validation Status": unit_row["Unit Validation Status"],
            "Membership Structure": membership_structure,
            "Cell 9 Gate Passed": cell9_gate_passed,
            "Cell 10 Gate Passed": cell10_gate_passed,
            "Cell 10+ Check Required": multi_check_required,
            "Cell 10+ Gate Passed": multi_gate_passed,
            "Semantic Comparison Eligible": semantic_eligible,
            "Semantic Comparison Performed": semantic_comparison_performed,
            "Semantic Result Source": semantic_result_source,
            "Expected Class Membership Found": expected_membership_found,
            "Matching Quantity Class": matching_class,
            "Semantic Compatibility Status": compatibility_status,
            "Semantic Validation Status": semantic_validation_status,
            "Cell 11 Status": cell11_status,
            "Final Workflow Status": final_workflow_status,
            "Expert Review Required": expert_review_required,
            "Review Origin": "; ".join(review_origins),
            "Review Reason": "; ".join(review_reasons),
            "Cell 9 Status": mnemonic_row["Cell 9 Status"],
            "Cell 9 Review Reason": mnemonic_row["Review Reason"],
            "Cell 10 Status": unit_row["Cell 10 Status"],
            "Cell 10 Review Reason": unit_row["Review Reason"],
            "Cell 10+ Status": multi_status,
            "Cell 10+ Compatibility Status": multi_compatibility,
            "Cell 10+ Review Reason": multi_review_reason,
            "PWLS Curve Source Files": mnemonic_row["PWLS Curve Source Files"],
            "PWLS Curve Source Commit": mnemonic_row["PWLS Curve Source Commit"],
            "Property Kind Source Files": mnemonic_row["Property Kind Source Files"],
            "Property Kind Source Commit": mnemonic_row["Property Kind Source Commit"],
            "UOM Unit Source Files": unit_row["UOM Unit Source Files"],
            "UOM Membership Source Files": unit_row["UOM Membership Source Files"],
            "UOM Source Commit": unit_row["UOM Source Commit"],
        })

    validation_df = pd.DataFrame(validation_rows, columns=SEMANTIC_VALIDATION_COLUMNS)
    validation_df.index = unit_validation_df.index.copy()
    pass_df = validation_df.loc[validation_df["Cell 11 Status"].eq("Pass")].copy()
    pass_df.index = pd.RangeIndex(start=1, stop=len(pass_df) + 1, name="pass_order")
    review_df = validation_df.loc[validation_df["Expert Review Required"].eq(True)].copy()
    review_df.index = pd.RangeIndex(start=1, stop=len(review_df) + 1, name="review_order")

    input_rows = len(validation_df)
    final_passes = int(validation_df["Cell 11 Status"].eq("Pass").sum())
    semantic_comparisons = int(validation_df["Semantic Comparison Performed"].eq(True).sum())
    stats = {
        "Input rows": input_rows,
        "Rows passing the Cell 9 gate": int(validation_df["Cell 9 Gate Passed"].eq(True).sum()),
        "Rows failing the Cell 9 gate": int(validation_df["Cell 9 Gate Passed"].eq(False).sum()),
        "Rows passing the Cell 10 gate": int(validation_df["Cell 10 Gate Passed"].eq(True).sum()),
        "Rows failing the Cell 10 gate": int(validation_df["Cell 10 Gate Passed"].eq(False).sum()),
        "Rows requiring the Cell 10+ gate": int(validation_df["Cell 10+ Check Required"].eq(True).sum()),
        "Rows passing the Cell 10+ gate": int(validation_df["Cell 10+ Gate Passed"].eq(True).sum()),
        "Recognized units without Quantity Class memberships": int(((validation_df["Cell 10 Gate Passed"] == True) & (validation_df["Distinct Quantity Class Membership Count"] == 0)).sum()),
        "Rows eligible for semantic comparison": int(validation_df["Semantic Comparison Eligible"].eq(True).sum()),
        "Rows routed to review before semantic comparison": int(validation_df["Semantic Comparison Eligible"].eq(False).sum()),
        "Semantic comparisons performed": semantic_comparisons,
        "Semantic compatibility passes": int(validation_df["Semantic Compatibility Status"].eq("Compatible").sum()),
        "Semantic incompatibilities": int(validation_df["Semantic Compatibility Status"].eq("Incompatible").sum()),
        "Cell 10+ incompatibilities carried to review": int(((validation_df["Semantic Compatibility Status"] == "Incompatible") & (validation_df["Semantic Result Source"] == "Cell 10+ multiple-membership check")).sum()),
        "New incompatibilities found by Cell 11": int(((validation_df["Semantic Compatibility Status"] == "Incompatible") & (validation_df["Semantic Result Source"] == "Cell 11 exact membership comparison")).sum()),
        "Semantic assessments not performed": int(validation_df["Semantic Validation Status"].eq("Not assessed").sum()),
        "Cell 11 passes": final_passes,
        "Cell 11 expert-review rows": int(validation_df["Expert Review Required"].eq(True).sum()),
        "Cell 11 pass rate among compared rows (%)": round((final_passes / semantic_comparisons) * 100, 2) if semantic_comparisons else 0.0,
        "Final workflow pass rate (%)": round((final_passes / input_rows) * 100, 2) if input_rows else 0.0,
    }

    return validation_df, pass_df, review_df, stats


def cell11_verify_outputs(upstream_snapshots):
    """Verify complete row preservation, gate enforcement, and exact semantic passes."""

    for object_name, snapshot in upstream_snapshots.items():
        pd.testing.assert_frame_equal(globals()[object_name], snapshot, check_dtype=True, check_exact=True)

    if len(SEMANTIC_VALIDATION_DF) != len(MNEMONIC_VALIDATION_DF):
        raise AssertionError("SEMANTIC_VALIDATION_DF does not contain exactly one row per source row.")

    if not SEMANTIC_VALIDATION_DF.index.equals(MNEMONIC_VALIDATION_DF.index):
        raise AssertionError("SEMANTIC_VALIDATION_DF did not preserve the source result index.")

    expected_source_orders = list(range(1, len(SEMANTIC_VALIDATION_DF) + 1))

    if SEMANTIC_VALIDATION_DF["Source Row Order"].tolist() != expected_source_orders:
        raise AssertionError("Cell 11 did not preserve the complete one-based Source Row Order sequence.")

    for _, result_row in SEMANTIC_VALIDATION_DF.iterrows():
        memberships = cell11_membership_tuple(result_row["Actual Quantity Class Memberships"])
        expected_class = cell11_stable_text(result_row["Expected Quantity Class"])

        if result_row["Cell 11 Status"] == "Pass":
            if not cell11_is_true(result_row["Cell 9 Gate Passed"]) or not cell11_is_true(result_row["Cell 10 Gate Passed"]):
                raise AssertionError(f"Cell 11 bypassed an upstream gate at Source Row Order {result_row['Source Row Order']}.")

            if cell11_is_true(result_row["Cell 10+ Check Required"]) and not cell11_is_true(result_row["Cell 10+ Gate Passed"]):
                raise AssertionError(f"Cell 11 bypassed Cell 10+ at Source Row Order {result_row['Source Row Order']}.")

            if not expected_class or expected_class not in memberships:
                raise AssertionError(f"Cell 11 produced an unsupported semantic pass at Source Row Order {result_row['Source Row Order']}.")

            if result_row["Semantic Compatibility Status"] != "Compatible":
                raise AssertionError(f"Cell 11 pass and compatibility status disagree at Source Row Order {result_row['Source Row Order']}.")

        if cell11_is_true(result_row["Semantic Comparison Performed"]) and not cell11_is_true(result_row["Semantic Comparison Eligible"]):
            raise AssertionError(f"Cell 11 compared an ineligible row at Source Row Order {result_row['Source Row Order']}.")

        if result_row["Semantic Compatibility Status"] == "Incompatible" and result_row["Semantic Result Source"].startswith("Cell 11"):
            if expected_class in memberships:
                raise AssertionError(f"Cell 11 rejected an exact membership at Source Row Order {result_row['Source Row Order']}.")

    pass_source_orders = SEMANTIC_VALIDATION_DF.loc[SEMANTIC_VALIDATION_DF["Cell 11 Status"].eq("Pass"), "Source Row Order"].tolist()
    review_source_orders = SEMANTIC_VALIDATION_DF.loc[SEMANTIC_VALIDATION_DF["Expert Review Required"].eq(True), "Source Row Order"].tolist()

    if SEMANTIC_PASS_DF["Source Row Order"].tolist() != pass_source_orders:
        raise AssertionError("SEMANTIC_PASS_DF does not exactly represent Cell 11 passes.")

    if SEMANTIC_REVIEW_DF["Source Row Order"].tolist() != review_source_orders:
        raise AssertionError("SEMANTIC_REVIEW_DF does not exactly represent final expert-review rows.")

    if len(SEMANTIC_PASS_DF) + len(SEMANTIC_REVIEW_DF) != len(SEMANTIC_VALIDATION_DF):
        raise AssertionError("Cell 11 pass and review outputs do not form a complete source-row partition.")


try:
    _CELL11_MULTI_LOOKUP = cell11_validate_inputs()

    # Deep snapshots make the non-destructive convergence verifiable.
    _CELL11_UPSTREAM_SNAPSHOTS = {
        "MNEMONIC_VALIDATION_DF": MNEMONIC_VALIDATION_DF.copy(deep=True),
        "UNIT_VALIDATION_DF": UNIT_VALIDATION_DF.copy(deep=True),
        "UNIT_MULTI_CLASS_VALIDATION_DF": UNIT_MULTI_CLASS_VALIDATION_DF.copy(deep=True),
    }

    SEMANTIC_VALIDATION_DF, SEMANTIC_PASS_DF, SEMANTIC_REVIEW_DF, SEMANTIC_VALIDATION_STATS = cell11_apply_validation(
        MNEMONIC_VALIDATION_DF,
        UNIT_VALIDATION_DF,
        UNIT_MULTI_CLASS_VALIDATION_DF,
        _CELL11_MULTI_LOOKUP,
    )

    cell11_verify_outputs(_CELL11_UPSTREAM_SNAPSHOTS)

    print("Cell 11 semantic Quantity Class validation completed successfully.")

    for statistic_name, statistic_value in SEMANTIC_VALIDATION_STATS.items():
        print(f"{statistic_name}: {statistic_value:,}")

    print("\nSemantic validation preview:")
    display(SEMANTIC_VALIDATION_DF.head(10))

    if SEMANTIC_REVIEW_DF.empty:
        print("\nAll source rows passed final semantic validation.")
    else:
        print("\nFinal expert-review queue:")
        display(SEMANTIC_REVIEW_DF.head(20))

except Exception as error:
    # Reset every output so later output cells cannot consume partial semantic results.
    SEMANTIC_VALIDATION_DF = pd.DataFrame(columns=SEMANTIC_VALIDATION_COLUMNS)
    SEMANTIC_PASS_DF = pd.DataFrame(columns=SEMANTIC_VALIDATION_COLUMNS)
    SEMANTIC_PASS_DF.index = pd.RangeIndex(start=1, stop=1, name="pass_order")
    SEMANTIC_REVIEW_DF = pd.DataFrame(columns=SEMANTIC_VALIDATION_COLUMNS)
    SEMANTIC_REVIEW_DF.index = pd.RangeIndex(start=1, stop=1, name="review_order")
    SEMANTIC_VALIDATION_STATS = {}

    print("Cell 11 failed.")
    print(f"Error: {error}")


# In[ ]:


# ============================================================
# CELL 12 — OUTPUT PREPARATION
# ============================================================

# Requires the imports from Cell 2 and the completed outputs from Cells 3–11.
# Cell 12 does not repeat validation. It converts the preserved technical results
# into a concise, human-readable Excel report and keeps the raw input unchanged.

from openpyxl.chart import BarChart, DoughnutChart, Reference


CELL12_SHEET_NAMES = [
    "1 Sources",
    "2 Treatment Statistics",
    "3 Raw Input",
    "4 Mnemonic Check",
    "5 Unit Check",
    "6 Multiple Unit Classes",
    "7 Semantic Validation",
    "8 Expert Review Queue",
    "9 Charts",
]

CELL12_TITLE_FILL = "FF17324D"
CELL12_HEADER_FILL = "FFFF773D"
CELL12_SUBHEADER_FILL = "FFF3F6F8"
CELL12_PASS_FILL = "FFC6EFCE"
CELL12_PASS_FONT = "FF006100"
CELL12_REVIEW_FILL = "FFFFC7CE"
CELL12_REVIEW_FONT = "FF9C0006"
CELL12_CANDIDATE_FILL = "FFFFEB9C"
CELL12_CANDIDATE_FONT = "FF9C6500"
CELL12_NEUTRAL_FILL = "FFE7E6E6"
CELL12_NEUTRAL_FONT = "FF595959"
CELL12_BODY_FONT = "FF1F1F1F"
CELL12_LINK_FONT = "FF0563C1"
CELL12_MAX_EXCEL_TEXT_LENGTH = 32700

CELL12_METADATA_COLUMNS = ["Service company", "mnemonic", "unit", "description"]

CELL12_MNEMONIC_REQUIRED_COLUMNS = [
    "Source Row Order", "Source Index", "Original Service Company", "Original Mnemonic",
    "Prepared Service Company", "Prepared Mnemonic", "Company Resolution Status",
    "Company Match Method", "Resolved Company Code", "Resolved Company Name",
    "Mnemonic Match Method", "Mnemonic Resolution Status", "Mnemonic Validation Status",
    "Reference Curve Mnemonic", "Candidate Property Kind", "Property Kind Title",
    "Property Kind Match Method", "Declared Quantity Class", "Expected Quantity Class",
    "Candidate Property Kind Titles", "Candidate Quantity Classes", "Cell 9 Status",
    "Review Reason", "PWLS Curve Source Files", "PWLS Curve Source Commit",
    "Property Kind Source Files", "Property Kind Source Commit",
]

CELL12_UNIT_REQUIRED_COLUMNS = [
    "Source Row Order", "Source Index", "Original Unit", "Prepared Unit",
    "Unit Recognition Status", "Unit Match Method", "Canonical Unit", "Unit Name",
    "Unit Dimension", "Unit Base Unit", "Unit Conversion Reference",
    "Actual Quantity Class Memberships", "Distinct Quantity Class Membership Count",
    "Candidate Canonical Units", "Candidate Unit Dimensions", "Unit Validation Status",
    "Cell 10 Status", "Review Reason", "UOM Unit Source Files",
    "UOM Membership Source Files", "UOM Mapping Source Files", "UOM Source Commit",
]

CELL12_MULTI_REQUIRED_COLUMNS = [
    "Source Row Order", "Source Index", "Prepared Service Company", "Prepared Mnemonic",
    "Prepared Unit", "Canonical Unit", "Unit Name", "Unit Dimension",
    "Property Kind Title", "Expected Quantity Class", "Actual Quantity Class Memberships",
    "Distinct Quantity Class Membership Count", "Expected Class Membership Found",
    "Matching Quantity Class", "Unit Compatibility Status", "Cell 10+ Status",
    "Review Required", "Review Reason",
]

CELL12_SEMANTIC_REQUIRED_COLUMNS = [
    "Source Row Order", "Source Index", "Original Service Company", "Original Mnemonic",
    "Original Unit", "Original Description", "Prepared Service Company", "Prepared Mnemonic",
    "Prepared Unit", "Prepared Description", "Resolved Company Code", "Resolved Company Name",
    "Property Kind Title", "Expected Quantity Class", "Canonical Unit", "Unit Name",
    "Unit Dimension", "Actual Quantity Class Memberships", "Cell 9 Gate Passed",
    "Cell 10 Gate Passed", "Cell 10+ Check Required", "Cell 10+ Gate Passed",
    "Semantic Comparison Eligible", "Semantic Comparison Performed", "Semantic Result Source",
    "Expected Class Membership Found", "Matching Quantity Class",
    "Semantic Compatibility Status", "Semantic Validation Status", "Cell 11 Status",
    "Final Workflow Status", "Expert Review Required", "Review Origin", "Review Reason",
]

CELL12_UOM_MEMBERSHIP_REQUIRED_COLUMNS = [
    "Quantity Class", "Quantity Class Dimension", "Base For Conversion", "Member Unit",
]

CELL12_MNEMONIC_RESULT_LABELS = {
    "Company-specific exact PWLS match": "Company-specific exact match",
    "Global exact PWLS match": "Global exact match",
    "Global case-insensitive PWLS candidate": "Case-insensitive candidate — review required",
    "Expert-approved mnemonic alias": "Expert-approved mnemonic alias match",
    "No mnemonic match": "No match",
    "Missing mnemonic": "Mnemonic not supplied",
}

CELL12_UNIT_RESULT_LABELS = {
    "Exact UOM symbol": "Raw unit exact match",
    "Exact automatic official mapping": "Official mapping match",
    "Official mapping requires review": "Official mapping candidate — review required",
    "Exact approved unit alias": "Expert-approved unit alias match",
    "Case-insensitive UOM candidate": "Case-insensitive candidate — review required",
    "No unit match": "No match",
    "No unit supplied": "Unit not supplied",
}

# Only two compact objects are exposed after execution. The cleaned report tables
# remain local to the export function so the notebook front end stays uncluttered.
CELL12_OUTPUT_PATH = None
CELL12_EXPORT_SUMMARY = {}


def cell12_is_missing(value):
    """Return True only for a scalar missing value."""

    if value is None:
        return True

    if isinstance(value, (list, tuple, set, dict)):
        return False

    try:
        missing_result = pd.isna(value)
        return bool(missing_result)
    except (TypeError, ValueError):
        return False


def cell12_stable_text(value):
    """Create readable output text without changing authoritative capitalization."""

    if cell12_is_missing(value):
        return ""

    return str(value)


CELL12_STAGE_NAME_REPLACEMENTS = [
    ("Cell 10+", "Multiple-Class Compatibility Check"),
    ("Cell 12", "Output Preparation"),
    ("Cell 11", "Semantic Quantity-Class Validation"),
    ("Cell 10", "Unit Validation"),
    ("Cell 9", "Company and Mnemonic Validation"),
    ("Cell 8", "Input Data Preprocessing"),
]


def cell12_friendly_stage_text(value):
    """Replace notebook cell numbers with stage names in exported text."""

    text_value = cell12_stable_text(value)

    for cell_label, stage_label in CELL12_STAGE_NAME_REPLACEMENTS:
        text_value = text_value.replace(cell_label, stage_label)

    return text_value


def cell12_collection(value):
    """Read upstream tuple/list values while preserving their first-seen order."""

    if cell12_is_missing(value):
        return []

    if isinstance(value, set):
        source_values = sorted(value, key=lambda item: cell12_stable_text(item))
    elif isinstance(value, (list, tuple)):
        source_values = value
    else:
        source_values = [value]

    ordered_values = []
    seen_values = set()

    for source_value in source_values:
        if isinstance(source_value, (list, tuple, set)):
            nested_values = cell12_collection(source_value)
        else:
            nested_values = [source_value]

        for nested_value in nested_values:
            text_value = cell12_stable_text(nested_value)

            if text_value and text_value not in seen_values:
                seen_values.add(text_value)
                ordered_values.append(text_value)

    return ordered_values


def cell12_ordered_text(values):
    """Combine scalar and collection values into one deterministic collection."""

    combined_values = []
    seen_values = set()

    for value in values:
        for text_value in cell12_collection(value):
            if text_value not in seen_values:
                seen_values.add(text_value)
                combined_values.append(text_value)

    return combined_values


def cell12_join(values, separator="; "):
    """Join an ordered collection for presentation in one Excel cell."""

    return separator.join(cell12_ordered_text(values))


def cell12_arrow_path(parts):
    """Create a compact, readable provenance chain."""

    path_parts = []

    for part in parts:
        text_value = cell12_stable_text(part)

        if text_value and (not path_parts or text_value != path_parts[-1]):
            path_parts.append(text_value)

    return " → ".join(path_parts)


def cell12_excel_value(value):
    """Convert pandas and collection values to safe native Excel cell values."""

    if cell12_is_missing(value):
        return None

    if isinstance(value, dict):
        value = json.dumps(value, ensure_ascii=False, default=str)
    elif isinstance(value, (list, tuple, set)):
        value = cell12_join([value])

    if hasattr(value, "item") and not isinstance(value, (str, bytes)):
        try:
            value = value.item()
        except (AttributeError, ValueError):
            pass

    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()

    if isinstance(value, str) and len(value) > CELL12_MAX_EXCEL_TEXT_LENGTH:
        value = value[:CELL12_MAX_EXCEL_TEXT_LENGTH] + " … [truncated for Excel]"

    return value


def cell12_require_dataframe(object_name, required_columns, allow_empty=False):
    """Validate one required DataFrame and report the received schema clearly."""

    if object_name not in globals() or globals()[object_name] is None:
        raise NameError(f"{object_name} is unavailable.")

    dataframe = globals()[object_name]

    if not isinstance(dataframe, pd.DataFrame):
        raise TypeError(f"{object_name} must be a pandas DataFrame.")

    actual_columns = list(dataframe.columns)
    missing_columns = [column for column in required_columns if column not in actual_columns]

    if missing_columns:
        raise ValueError(
            f"{object_name} is missing required columns. "
            f"Expected at least: {required_columns}. Received: {actual_columns}"
        )

    if not allow_empty and dataframe.empty:
        raise ValueError(f"{object_name} is empty.")


def cell12_validate_inputs():
    """Confirm that every source, statistic, and validation result needed for the report exists."""

    dependency_messages = {
        "METADATA_DF": "METADATA_DF is unavailable. Run Cell 3 before running Cell 12.",
        "METADATA_ORIGINAL_DF": "METADATA_ORIGINAL_DF is unavailable. Run Cell 8 before running Cell 12.",
        "METADATA_INPUT_PREPROCESSING_DF": "METADATA_INPUT_PREPROCESSING_DF is unavailable. Run Cell 8 before running Cell 12.",
        "MNEMONIC_VALIDATION_DF": "MNEMONIC_VALIDATION_DF is unavailable. Run Cell 9 before running Cell 12.",
        "UNIT_VALIDATION_DF": "UNIT_VALIDATION_DF is unavailable. Run Cell 10 before running Cell 12.",
        "UNIT_MULTI_CLASS_VALIDATION_DF": "UNIT_MULTI_CLASS_VALIDATION_DF is unavailable. Run Cell 10+ before running Cell 12.",
        "SEMANTIC_VALIDATION_DF": "SEMANTIC_VALIDATION_DF is unavailable. Run Cell 11 before running Cell 12.",
    }

    for object_name, error_message in dependency_messages.items():
        if object_name not in globals() or globals()[object_name] is None:
            raise NameError(error_message)

    cell12_require_dataframe("METADATA_DF", CELL12_METADATA_COLUMNS)
    cell12_require_dataframe("METADATA_ORIGINAL_DF", CELL12_METADATA_COLUMNS)
    cell12_require_dataframe("METADATA_INPUT_PREPROCESSING_DF", CELL12_METADATA_COLUMNS)
    cell12_require_dataframe("METADATA_INPUT_PREPROCESSING_LOG_DF", ["Source Row Order", "Column", "Original Value", "Prepared Value", "Applied Rules"], allow_empty=True)
    cell12_require_dataframe("MNEMONIC_VALIDATION_DF", CELL12_MNEMONIC_REQUIRED_COLUMNS)
    cell12_require_dataframe("UNIT_VALIDATION_DF", CELL12_UNIT_REQUIRED_COLUMNS)
    cell12_require_dataframe("UNIT_MULTI_CLASS_VALIDATION_DF", CELL12_MULTI_REQUIRED_COLUMNS, allow_empty=True)
    cell12_require_dataframe("SEMANTIC_VALIDATION_DF", CELL12_SEMANTIC_REQUIRED_COLUMNS)
    cell12_require_dataframe("SEMANTIC_REVIEW_DF", CELL12_SEMANTIC_REQUIRED_COLUMNS, allow_empty=True)
    cell12_require_dataframe("UOM_QUANTITY_CLASS_MEMBERS_DF", CELL12_UOM_MEMBERSHIP_REQUIRED_COLUMNS)
    cell12_require_dataframe("PWLS_CATALOG_FILES_DF", ["Source File", "Catalog Type", "Is Archive"])
    cell12_require_dataframe("UOM_SOURCE_FILES_DF", ["Source File", "File Category", "Root Element", "File Size Bytes"])

    required_dictionaries = [
        "METADATA_INPUT", "PWLS_CATALOG_SOURCE", "PWLS_IMPORT_STATS",
        "PWLS_PROPERTY_SOURCE", "PWLS_PROPERTY_STATS", "UOM_SOURCE", "UOM_STATS",
        "UOM_UNIT_MAPPING_STATE_STATS", "ALIASES_WORKBOOK_INFO", "ALIASES_WORKBOOK_STATS",
        "METADATA_INPUT_PREPROCESSING_STATS", "MNEMONIC_VALIDATION_STATS", "UNIT_VALIDATION_STATS",
        "UNIT_MULTI_CLASS_STATS", "SEMANTIC_VALIDATION_STATS",
    ]

    for object_name in required_dictionaries:
        if object_name not in globals() or not isinstance(globals()[object_name], dict):
            raise NameError(f"{object_name} is unavailable. Rerun its source cell before running Cell 12.")

    if "ALIASES_WORKBOOK_PATH" not in globals() or ALIASES_WORKBOOK_PATH is None:
        raise NameError("ALIASES_WORKBOOK_PATH is unavailable. Run Cell 7 before running Cell 12.")

    input_row_count = len(METADATA_ORIGINAL_DF)

    for dataframe_name in ["METADATA_INPUT_PREPROCESSING_DF", "MNEMONIC_VALIDATION_DF", "UNIT_VALIDATION_DF", "SEMANTIC_VALIDATION_DF"]:
        if len(globals()[dataframe_name]) != input_row_count:
            raise ValueError(f"{dataframe_name} does not contain exactly one row per original input row.")

    if not METADATA_DF.index.equals(METADATA_ORIGINAL_DF.index):
        raise ValueError("METADATA_DF and METADATA_ORIGINAL_DF indices do not agree.")

    for dataframe_name in ["METADATA_INPUT_PREPROCESSING_DF", "MNEMONIC_VALIDATION_DF", "UNIT_VALIDATION_DF", "SEMANTIC_VALIDATION_DF"]:
        if not globals()[dataframe_name].index.equals(METADATA_ORIGINAL_DF.index):
            raise ValueError(f"{dataframe_name} did not preserve the original input index.")

    expected_source_orders = list(range(1, input_row_count + 1))

    for dataframe_name in ["MNEMONIC_VALIDATION_DF", "UNIT_VALIDATION_DF", "SEMANTIC_VALIDATION_DF"]:
        if globals()[dataframe_name]["Source Row Order"].tolist() != expected_source_orders:
            raise ValueError(f"{dataframe_name} Source Row Order is not aligned with the input rows.")

    multi_source_orders = UNIT_MULTI_CLASS_VALIDATION_DF["Source Row Order"].tolist()

    if len(multi_source_orders) != len(set(multi_source_orders)):
        raise ValueError("UNIT_MULTI_CLASS_VALIDATION_DF contains repeated Source Row Order values.")

    if not set(multi_source_orders).issubset(set(expected_source_orders)):
        raise ValueError("UNIT_MULTI_CLASS_VALIDATION_DF contains an invalid Source Row Order.")

    pd.testing.assert_frame_equal(METADATA_DF, METADATA_ORIGINAL_DF, check_dtype=True, check_exact=True)


def cell12_create_snapshots():
    """Snapshot upstream tables because output preparation must never alter analytical results."""

    dataframe_names = [
        "METADATA_DF", "METADATA_ORIGINAL_DF", "METADATA_INPUT_PREPROCESSING_DF",
        "METADATA_INPUT_PREPROCESSING_LOG_DF", "MNEMONIC_VALIDATION_DF", "UNIT_VALIDATION_DF",
        "UNIT_MULTI_CLASS_VALIDATION_DF", "SEMANTIC_VALIDATION_DF", "SEMANTIC_REVIEW_DF",
        "UOM_QUANTITY_CLASS_MEMBERS_DF", "PWLS_CATALOG_FILES_DF", "UOM_SOURCE_FILES_DF",
    ]
    dictionary_names = [
        "PWLS_CATALOG_SOURCE", "PWLS_IMPORT_STATS", "PWLS_PROPERTY_SOURCE",
        "PWLS_PROPERTY_STATS", "UOM_SOURCE", "UOM_STATS", "UOM_UNIT_MAPPING_STATE_STATS",
        "ALIASES_WORKBOOK_INFO", "ALIASES_WORKBOOK_STATS", "METADATA_INPUT_PREPROCESSING_STATS",
        "MNEMONIC_VALIDATION_STATS", "UNIT_VALIDATION_STATS", "UNIT_MULTI_CLASS_STATS",
        "SEMANTIC_VALIDATION_STATS",
    ]

    dataframe_snapshots = {name: globals()[name].copy(deep=True) for name in dataframe_names}
    dictionary_snapshots = {
        name: json.dumps(globals()[name], ensure_ascii=False, sort_keys=True, default=str)
        for name in dictionary_names
    }
    scalar_snapshots = {
        "Metadata Input Path": cell12_stable_text(METADATA_INPUT.get("path")),
        "Alias Workbook Path": cell12_stable_text(ALIASES_WORKBOOK_PATH),
    }

    return dataframe_snapshots, dictionary_snapshots, scalar_snapshots


def cell12_verify_upstream_unchanged(dataframe_snapshots, dictionary_snapshots, scalar_snapshots):
    """Verify that the report export was a read-only consumer of all upstream results."""

    for object_name, snapshot in dataframe_snapshots.items():
        pd.testing.assert_frame_equal(globals()[object_name], snapshot, check_dtype=True, check_exact=True)

    for object_name, snapshot in dictionary_snapshots.items():
        current_value = json.dumps(globals()[object_name], ensure_ascii=False, sort_keys=True, default=str)

        if current_value != snapshot:
            raise AssertionError(f"Cell 12 changed upstream object {object_name}.")

    if cell12_stable_text(METADATA_INPUT.get("path")) != scalar_snapshots["Metadata Input Path"]:
        raise AssertionError("Cell 12 changed the stored metadata input path.")

    if cell12_stable_text(ALIASES_WORKBOOK_PATH) != scalar_snapshots["Alias Workbook Path"]:
        raise AssertionError("Cell 12 changed the Expert Review Workbook path.")


def cell12_choose_output_folder():
    """Ask for the output folder through Windows Explorer, with a manual fallback."""

    tkinter_ready = bool(globals().get("TKINTER_AVAILABLE", False))
    dialog_object = globals().get("filedialog")
    root_factory = globals().get("Tk")

    if root_factory is None and globals().get("tk") is not None:
        root_factory = tk.Tk

    if tkinter_ready and dialog_object is not None and root_factory is not None:
        root_window = None
        dialog_failed = False

        try:
            root_window = root_factory()
            root_window.withdraw()

            try:
                root_window.attributes("-topmost", True)
            except Exception:
                pass

            selected_folder = dialog_object.askdirectory(
                parent=root_window,
                title="Select the folder for the metadata-validation output",
                mustexist=True,
            )

        except Exception as dialog_error:
            print(f"The Windows folder window could not be opened: {dialog_error}")
            selected_folder = ""
            dialog_failed = True

        finally:
            if root_window is not None:
                try:
                    root_window.destroy()
                except Exception:
                    pass

        if selected_folder:
            output_folder = Path(selected_folder).expanduser()
        elif not dialog_failed:
            raise RuntimeError("No output folder was selected. Cell 12 did not create a workbook.")
        else:
            output_folder = None
    else:
        output_folder = None

    if output_folder is None:
        print("Paste the complete output-folder path below.")
        manual_folder = input("Output folder: ")
        cleaned_folder = str(manual_folder).strip().strip("\"'").strip()

        if not cleaned_folder:
            raise RuntimeError("No output folder was supplied. Cell 12 did not create a workbook.")

        output_folder = Path(cleaned_folder).expanduser()

    if not output_folder.exists():
        raise FileNotFoundError(f"The selected output folder does not exist: {output_folder}")

    if not output_folder.is_dir():
        raise NotADirectoryError(f"The selected output location is not a folder: {output_folder}")

    return output_folder.resolve()


def cell12_create_output_path(output_folder, export_time):
    """Create a Windows-safe, timestamped filename without overwriting an earlier report."""

    timestamp_text = export_time.strftime("%Y-%m-%d_%H-%M-%S")
    base_stem = f"Well_Metadata_Validation_Output_{timestamp_text}"
    output_path = output_folder / f"{base_stem}.xlsx"
    collision_number = 2

    while output_path.exists():
        output_path = output_folder / f"{base_stem}_{collision_number}.xlsx"
        collision_number += 1

    return output_path


def cell12_build_quantity_class_index():
    """Index UOM membership metadata once for unit enrichment in the output tables."""

    quantity_class_index = {}

    for source_position in range(len(UOM_QUANTITY_CLASS_MEMBERS_DF)):
        source_row = UOM_QUANTITY_CLASS_MEMBERS_DF.iloc[source_position]
        quantity_class = cell12_stable_text(source_row["Quantity Class"])

        if not quantity_class:
            continue

        class_details = quantity_class_index.setdefault(quantity_class, {
            "dimensions": [], "base_units": [], "member_units": [],
        })

        detail_values = {
            "dimensions": source_row["Quantity Class Dimension"],
            "base_units": source_row["Base For Conversion"],
            "member_units": source_row["Member Unit"],
        }

        for detail_name, detail_value in detail_values.items():
            text_value = cell12_stable_text(detail_value)

            if text_value and text_value not in class_details[detail_name]:
                class_details[detail_name].append(text_value)

    return quantity_class_index


def cell12_enrich_quantity_classes(quantity_classes, quantity_class_index):
    """Return dimensions, conversion bases, and member units for preserved class candidates."""

    class_names = cell12_ordered_text([quantity_classes])
    dimensions = []
    base_units = []
    member_units = []
    grouped_member_units = []

    for class_name in class_names:
        class_details = quantity_class_index.get(class_name, {})
        class_dimensions = class_details.get("dimensions", [])
        class_base_units = class_details.get("base_units", [])
        class_member_units = class_details.get("member_units", [])

        dimensions = cell12_ordered_text([dimensions, class_dimensions])
        base_units = cell12_ordered_text([base_units, class_base_units])
        member_units = cell12_ordered_text([member_units, class_member_units])

        if class_member_units:
            grouped_member_units.append(f"{class_name}: {', '.join(class_member_units)}")

    return {
        "classes": "; ".join(class_names),
        "dimensions": "; ".join(dimensions),
        "base_units": "; ".join(base_units),
        "member_units": "; ".join(member_units),
        "grouped_member_units": " | ".join(grouped_member_units),
        "member_unit_count": len(member_units),
    }


def cell12_add_source_row(rows, section, source, item, value, notes=""):
    """Add one auditable source or source-statistic row."""

    rows.append({
        "Section": section,
        "Source": source,
        "Item": item,
        "Value": value,
        "Notes": notes,
    })


def cell12_build_sources_dataframe(output_path, export_time):
    """Build the input-path, repository, commit, source-file, and import-statistics sheet."""

    source_rows = []
    metadata_path_value = METADATA_INPUT.get("path")
    metadata_path_text = cell12_stable_text(metadata_path_value)

    cell12_add_source_row(source_rows, "Export", "Output Preparation", "Generated at", export_time, "Local notebook time")
    cell12_add_source_row(source_rows, "Export", "Output Preparation", "Output workbook", str(output_path), "Timestamped; no existing file overwritten")
    cell12_add_source_row(source_rows, "Raw input", "Metadata CSV", "Input path", metadata_path_text or "Unavailable", "Path selected during Metadata CSV Import")
    cell12_add_source_row(source_rows, "Raw input", "Metadata CSV", "Rows", len(METADATA_ORIGINAL_DF))
    cell12_add_source_row(source_rows, "Raw input", "Metadata CSV", "Columns", len(METADATA_ORIGINAL_DF.columns))
    cell12_add_source_row(source_rows, "Raw input", "Metadata CSV", "Column names", cell12_join([list(METADATA_ORIGINAL_DF.columns)]))
    cell12_add_source_row(source_rows, "Raw input", "Metadata CSV", "Index name", cell12_stable_text(METADATA_ORIGINAL_DF.index.name) or "Unnamed")
    cell12_add_source_row(source_rows, "Raw input", "Metadata CSV", "Duplicate index labels", int(METADATA_ORIGINAL_DF.index.duplicated(keep=False).sum()))
    cell12_add_source_row(source_rows, "Raw input", "Metadata CSV", "Duplicate records after first", int(METADATA_ORIGINAL_DF.duplicated(keep="first").sum()))

    if metadata_path_text:
        try:
            metadata_file_path = Path(metadata_path_text).expanduser()

            if metadata_file_path.exists() and metadata_file_path.is_file():
                file_statistics = metadata_file_path.stat()
                cell12_add_source_row(source_rows, "Raw input", "Metadata CSV", "File size (bytes)", file_statistics.st_size)
                cell12_add_source_row(source_rows, "Raw input", "Metadata CSV", "Last modified", datetime.fromtimestamp(file_statistics.st_mtime))
        except (OSError, ValueError):
            cell12_add_source_row(source_rows, "Raw input", "Metadata CSV", "File-system status", "Path could not be inspected at export time")

    for column_name in CELL12_METADATA_COLUMNS:
        source_values = [METADATA_ORIGINAL_DF.iat[position, METADATA_ORIGINAL_DF.columns.get_loc(column_name)] for position in range(len(METADATA_ORIGINAL_DF))]
        populated_values = [value for value in source_values if cell12_stable_text(value) != ""]
        distinct_values = cell12_ordered_text([populated_values])
        cell12_add_source_row(source_rows, "Raw input profile", column_name, "Populated values", len(populated_values))
        cell12_add_source_row(source_rows, "Raw input profile", column_name, "Blank or missing values", len(source_values) - len(populated_values))
        cell12_add_source_row(source_rows, "Raw input profile", column_name, "Distinct populated values", len(distinct_values))

    source_definitions = [
        ("PWLS Curve Catalog", "GitLab", PWLS_CATALOG_SOURCE, PWLS_IMPORT_STATS),
        ("PWLS Property Kind Dictionary", "GitHub", PWLS_PROPERTY_SOURCE, PWLS_PROPERTY_STATS),
        ("Energistics UOM", "GitHub", UOM_SOURCE, UOM_STATS),
    ]

    for source_name, platform_name, source_dictionary, statistic_dictionary in source_definitions:
        for item_name, item_value in source_dictionary.items():
            if item_name == "files":
                continue
            cell12_add_source_row(source_rows, "Reference source", f"{source_name} ({platform_name})", str(item_name).title(), item_value)

        for statistic_name, statistic_value in statistic_dictionary.items():
            cell12_add_source_row(source_rows, "Original import statistics", source_name, statistic_name, statistic_value)

    for mapping_state, mapping_count in UOM_UNIT_MAPPING_STATE_STATS.items():
        cell12_add_source_row(source_rows, "Original import statistics", "Energistics UOM mapping states", mapping_state, mapping_count)

    cell12_add_source_row(source_rows, "Governed local source", "Expert Review Workbook", "Workbook path", str(ALIASES_WORKBOOK_PATH))

    for item_name, item_value in ALIASES_WORKBOOK_INFO.items():
        cell12_add_source_row(source_rows, "Governed local source", "Expert Review Workbook", item_name, item_value)

    for statistic_name, statistic_value in ALIASES_WORKBOOK_STATS.items():
        cell12_add_source_row(source_rows, "Original import statistics", "Expert Review Workbook", statistic_name, statistic_value)

    for source_position in range(len(PWLS_CATALOG_FILES_DF)):
        source_row = PWLS_CATALOG_FILES_DF.iloc[source_position]
        source_notes = [f"Type: {cell12_stable_text(source_row['Catalog Type'])}"]

        if "Catalog Company" in PWLS_CATALOG_FILES_DF.columns and cell12_stable_text(source_row["Catalog Company"]):
            source_notes.append(f"Catalog company: {cell12_stable_text(source_row['Catalog Company'])}")

        source_notes.append(f"Archive: {bool(source_row['Is Archive'])}")
        cell12_add_source_row(
            source_rows,
            "Imported reference file",
            "PWLS Curve Catalog",
            f"File {source_position + 1}",
            source_row["Source File"],
            "; ".join(source_notes),
        )

    for source_position in range(len(UOM_SOURCE_FILES_DF)):
        source_row = UOM_SOURCE_FILES_DF.iloc[source_position]
        notes = f"Category: {cell12_stable_text(source_row['File Category'])}; Root: {cell12_stable_text(source_row['Root Element'])}; Size: {cell12_stable_text(source_row['File Size Bytes'])} bytes"
        cell12_add_source_row(source_rows, "Imported reference file", "Energistics UOM", f"File {source_position + 1}", source_row["Source File"], notes)

    return pd.DataFrame(source_rows, columns=["Section", "Source", "Item", "Value", "Notes"])


def cell12_stat_flag(metric_name):
    """Group statistics by the treatment meaning represented by the workflow flags."""

    metric_text = metric_name.casefold()

    if any(token in metric_text for token in ["review", "incompatib", "not found", "not recognized", "unresolved", "ambiguous", "conflict", "missing"]):
        return "Review / Exception"

    if any(token in metric_text for token in ["not provided", "not assessed", "deferred", "unavailable"]):
        return "Deferred / Not assessed"

    if any(token in metric_text for token in ["pass", "recognized", "resolved", "exact mnemonic match", "alias match", "compatible"]):
        return "Resolved / Pass"

    if any(token in metric_text for token in ["input", "rows", "checked", "performed", "eligible"]):
        return "Population / Flow"

    return "Information"


def cell12_stat_meaning(phase_name, metric_name):
    """Provide concise interpretations for the most important report metrics."""

    meanings = {
        "Rows containing changes": "Input rows changed by at least one permitted Input Data Preprocessing rule.",
        "Mnemonic passes": "Rows with one supported, non-abstract Property Kind and expected Quantity Class.",
        "Mnemonic reviews required": "Rows that could not receive an automatic mnemonic-side pass.",
        "Rows with multiple Quantity Class memberships": "Recognized units that require contextual Quantity Class selection.",
        "Unit Validation passes": "Rows whose prepared units resolved automatically to one canonical UOM unit.",
        "Semantic comparisons performed": "Rows that passed every required upstream gate and were compared.",
        "Semantic compatibility passes": "Rows whose expected Quantity Class occurs among the canonical unit memberships.",
        "Semantic incompatibilities": "Rows whose expected class is absent from the canonical unit memberships.",
        "Semantic Quantity-Class Validation expert-review rows": "Final rows requiring expert action for any upstream or semantic issue.",
        "Final workflow pass rate (%)": "Final automatic semantic passes divided by all input rows.",
    }

    return meanings.get(metric_name, f"Statistic reported by {phase_name}.")


def cell12_add_stat_row(rows, phase, metric, value, input_rows, meaning=None):
    """Add one statistic with a consistent type and optional share of input rows."""

    metric_text = cell12_friendly_stage_text(metric)
    is_percentage = metric_text.endswith("(%)")
    metric_type = "Percent" if is_percentage else "Count" if isinstance(value, (int, float)) and not isinstance(value, bool) else "Value"

    if is_percentage and isinstance(value, (int, float)):
        stored_value = float(value) / 100.0
        share_of_input = stored_value
    else:
        stored_value = value
        share_of_input = None

        if isinstance(value, (int, float)) and not isinstance(value, bool) and input_rows and 0 <= value <= input_rows:
            share_of_input = float(value) / float(input_rows)

    rows.append({
        "Phase": phase,
        "Flag Group": cell12_stat_flag(metric_text),
        "Metric": metric_text,
        "Value": stored_value,
        "Metric Type": metric_type,
        "Share of Input": share_of_input,
        "Meaning": meaning or cell12_stat_meaning(phase, metric_text),
    })


def cell12_build_statistics_dataframe():
    """Combine the treatment flags and statistics from Input Data Preprocessing through final validation."""

    statistics_rows = []
    input_rows = len(METADATA_ORIGINAL_DF)
    populated_companies = [cell12_stable_text(value) for value in METADATA_ORIGINAL_DF["Service company"].tolist() if cell12_stable_text(value)]
    populated_mnemonics = [cell12_stable_text(value) for value in METADATA_ORIGINAL_DF["mnemonic"].tolist() if cell12_stable_text(value)]
    populated_units = [cell12_stable_text(value) for value in METADATA_ORIGINAL_DF["unit"].tolist() if cell12_stable_text(value)]
    final_passes = int(SEMANTIC_VALIDATION_DF["Cell 11 Status"].eq("Pass").sum())
    final_reviews = int(SEMANTIC_VALIDATION_DF["Expert Review Required"].eq(True).sum())

    overview_statistics = [
        ("Input rows", input_rows, "Complete positional input population."),
        ("Distinct populated service-company values", len(cell12_ordered_text([populated_companies])), "Exact raw service-company values before Input Data Preprocessing."),
        ("Distinct populated mnemonic values", len(cell12_ordered_text([populated_mnemonics])), "Exact raw mnemonic values before Input Data Preprocessing."),
        ("Distinct populated unit values", len(cell12_ordered_text([populated_units])), "Exact raw unit values before Input Data Preprocessing."),
        ("Duplicate input records after first", int(METADATA_ORIGINAL_DF.duplicated(keep="first").sum()), "Repeated complete source records retained in the workflow."),
        ("Rows with duplicate input index labels", int(METADATA_ORIGINAL_DF.index.duplicated(keep=False).sum()), "Rows whose original DataFrame index label occurs more than once."),
        ("Final semantic passes", final_passes, "Rows passing the complete mnemonic–unit semantic workflow."),
        ("Final expert-review rows", final_reviews, "Rows preserved for expert action after all automated checks."),
        ("Final workflow pass rate (%)", round((final_passes / input_rows) * 100, 2) if input_rows else 0.0, "Final semantic passes divided by all input rows."),
        ("Final expert-review rate (%)", round((final_reviews / input_rows) * 100, 2) if input_rows else 0.0, "Final expert-review rows divided by all input rows."),
    ]

    for metric_name, metric_value, metric_meaning in overview_statistics:
        cell12_add_stat_row(statistics_rows, "Overall", metric_name, metric_value, input_rows, metric_meaning)

    phase_statistics = [
        ("Input Data Preprocessing", METADATA_INPUT_PREPROCESSING_STATS),
        ("Company and Mnemonic Validation", MNEMONIC_VALIDATION_STATS),
        ("Unit Validation", UNIT_VALIDATION_STATS),
        ("Multiple-Class Compatibility Check", UNIT_MULTI_CLASS_STATS),
        ("Semantic Quantity-Class Validation", SEMANTIC_VALIDATION_STATS),
    ]

    for phase_name, phase_dictionary in phase_statistics:
        for statistic_name, statistic_value in phase_dictionary.items():
            cell12_add_stat_row(statistics_rows, phase_name, statistic_name, statistic_value, input_rows)

    return pd.DataFrame(statistics_rows, columns=["Phase", "Flag Group", "Metric", "Value", "Metric Type", "Share of Input", "Meaning"])


def cell12_build_raw_dataframe():
    """Expose the untouched input with both positional order and original index label."""

    raw_dataframe = METADATA_ORIGINAL_DF.copy(deep=True)
    raw_dataframe.insert(0, "Input Index", list(METADATA_ORIGINAL_DF.index))
    raw_dataframe.insert(0, "Source Row Order", list(range(1, len(raw_dataframe) + 1)))
    return raw_dataframe


def cell12_build_mnemonic_dataframe():
    """Create one readable mnemonic-side summary row for every source record."""

    output_rows = []

    for source_position in range(len(MNEMONIC_VALIDATION_DF)):
        source_row = MNEMONIC_VALIDATION_DF.iloc[source_position]
        match_method = cell12_stable_text(source_row["Mnemonic Match Method"])
        result_label = CELL12_MNEMONIC_RESULT_LABELS.get(match_method, match_method or "Not assessed")
        mnemonic_pass = cell12_stable_text(source_row["Mnemonic Validation Status"]) == "Pass"

        if mnemonic_pass and cell12_stable_text(source_row["Property Kind Title"]):
            property_kinds = cell12_ordered_text([source_row["Property Kind Title"]])
        else:
            property_kinds = cell12_ordered_text([
                source_row["Property Kind Title"],
                source_row["Candidate Property Kind"],
                source_row["Candidate Property Kind Titles"],
            ])

        if mnemonic_pass and cell12_stable_text(source_row["Expected Quantity Class"]):
            quantity_classes = cell12_ordered_text([source_row["Expected Quantity Class"]])
        else:
            quantity_classes = cell12_ordered_text([
                source_row["Expected Quantity Class"],
                source_row["Declared Quantity Class"],
                source_row["Candidate Quantity Classes"],
            ])

        raw_mnemonic = cell12_stable_text(source_row["Original Mnemonic"])
        prepared_mnemonic = cell12_stable_text(source_row["Prepared Mnemonic"])
        reference_mnemonic = cell12_join([source_row["Reference Curve Mnemonic"]])
        provenance = cell12_arrow_path([
            raw_mnemonic or "[mnemonic not supplied]",
            f"Input Data Preprocessing: {prepared_mnemonic}" if prepared_mnemonic and prepared_mnemonic != raw_mnemonic else prepared_mnemonic,
            result_label,
            f"Reference mnemonic: {reference_mnemonic}" if reference_mnemonic else "",
            f"Property Kind: {'; '.join(property_kinds)}" if property_kinds else "",
            f"Quantity Class: {'; '.join(quantity_classes)}" if quantity_classes else "",
        ])

        output_rows.append({
            "Source Row Order": source_row["Source Row Order"],
            "Input Index": source_row["Source Index"],
            "Raw Service Company": source_row["Original Service Company"],
            "Prepared Service Company": source_row["Prepared Service Company"],
            "Company Check Result": source_row["Company Match Method"],
            "Resolved Company Code": source_row["Resolved Company Code"],
            "Resolved Company Name": source_row["Resolved Company Name"],
            "Raw Mnemonic": source_row["Original Mnemonic"],
            "Prepared Mnemonic": source_row["Prepared Mnemonic"],
            "Mnemonic Check Result": result_label,
            "Reference Mnemonic": reference_mnemonic,
            "Found Property Kind(s)": "; ".join(property_kinds),
            "Found Quantity Class(es)": "; ".join(quantity_classes),
            "Mnemonic Pass": mnemonic_pass,
            "Resolution Status": source_row["Mnemonic Resolution Status"],
            "Review Reason": cell12_friendly_stage_text(source_row["Review Reason"]),
            "Provenance": provenance,
        })

    columns = [
        "Source Row Order", "Input Index", "Raw Service Company", "Prepared Service Company",
        "Company Check Result", "Resolved Company Code", "Resolved Company Name", "Raw Mnemonic",
        "Prepared Mnemonic", "Mnemonic Check Result", "Reference Mnemonic",
        "Found Property Kind(s)", "Found Quantity Class(es)", "Mnemonic Pass",
        "Resolution Status", "Review Reason", "Provenance",
    ]
    return pd.DataFrame(output_rows, columns=columns)


def cell12_build_unit_dataframe(quantity_class_index):
    """Create the recognized-unit view and enrich only automatically passing units."""

    output_rows = []

    for source_position in range(len(UNIT_VALIDATION_DF)):
        source_row = UNIT_VALIDATION_DF.iloc[source_position]
        match_method = cell12_stable_text(source_row["Unit Match Method"])
        result_label = CELL12_UNIT_RESULT_LABELS.get(match_method, match_method or "Not assessed")
        unit_pass = cell12_stable_text(source_row["Cell 10 Status"]) == "Pass"
        memberships = cell12_ordered_text([source_row["Actual Quantity Class Memberships"]]) if unit_pass else []
        enrichment = cell12_enrich_quantity_classes(memberships, quantity_class_index) if unit_pass else {
            "classes": "", "dimensions": "", "base_units": "", "member_units": "",
            "grouped_member_units": "", "member_unit_count": 0,
        }
        raw_unit = cell12_stable_text(source_row["Original Unit"])
        prepared_unit = cell12_stable_text(source_row["Prepared Unit"])
        canonical_unit = cell12_stable_text(source_row["Canonical Unit"])
        unit_dimensions = cell12_ordered_text([source_row["Unit Dimension"], enrichment["dimensions"]]) if unit_pass else []
        provenance = cell12_arrow_path([
            raw_unit or "[unit not supplied]",
            f"Input Data Preprocessing: {prepared_unit}" if prepared_unit and prepared_unit != raw_unit else prepared_unit,
            result_label,
            f"Canonical unit: {canonical_unit}" if canonical_unit else "",
            f"Possible Quantity Class: {enrichment['classes']}" if enrichment["classes"] else "",
            f"Dimension: {'; '.join(unit_dimensions)}" if unit_dimensions else "",
        ])

        output_rows.append({
            "Source Row Order": source_row["Source Row Order"],
            "Input Index": source_row["Source Index"],
            "Raw Unit": source_row["Original Unit"],
            "Prepared Unit": source_row["Prepared Unit"],
            "Unit Check Result": result_label,
            "Canonical Unit": source_row["Canonical Unit"],
            "Unit Name": source_row["Unit Name"],
            "Unit Pass": unit_pass,
            "Possible Quantity Class(es)": enrichment["classes"],
            "Possible Dimension(s)": "; ".join(unit_dimensions),
            "Base Unit(s) for Conversion": enrichment["base_units"],
            "Unit Base Unit": source_row["Unit Base Unit"] if unit_pass else "",
            "Conversion Reference": source_row["Unit Conversion Reference"] if unit_pass else "",
            "Member Unit Count": enrichment["member_unit_count"],
            "Member Units in Possible Class(es)": enrichment["grouped_member_units"],
            "Recognition Status": source_row["Unit Recognition Status"],
            "Review Reason": cell12_friendly_stage_text(source_row["Review Reason"]),
            "Provenance": provenance,
        })

    columns = [
        "Source Row Order", "Input Index", "Raw Unit", "Prepared Unit", "Unit Check Result",
        "Canonical Unit", "Unit Name", "Unit Pass", "Possible Quantity Class(es)",
        "Possible Dimension(s)", "Base Unit(s) for Conversion", "Unit Base Unit",
        "Conversion Reference", "Member Unit Count", "Member Units in Possible Class(es)",
        "Recognition Status", "Review Reason", "Provenance",
    ]
    return pd.DataFrame(output_rows, columns=columns)


def cell12_build_multi_class_dataframe(quantity_class_index):
    """Present every multiple-membership unit, its contextual result, and its enrichment."""

    output_rows = []

    for source_position in range(len(UNIT_MULTI_CLASS_VALIDATION_DF)):
        source_row = UNIT_MULTI_CLASS_VALIDATION_DF.iloc[source_position]
        source_row_order = int(source_row["Source Row Order"])
        unit_source_row = UNIT_VALIDATION_DF.iloc[source_row_order - 1]
        memberships = cell12_ordered_text([source_row["Actual Quantity Class Memberships"]])
        all_enrichment = cell12_enrich_quantity_classes(memberships, quantity_class_index)
        matching_class = cell12_stable_text(source_row["Matching Quantity Class"])
        matched_enrichment = cell12_enrich_quantity_classes([matching_class], quantity_class_index) if matching_class else {
            "classes": "", "dimensions": "", "base_units": "", "member_units": "",
            "grouped_member_units": "", "member_unit_count": 0,
        }
        compatibility_status = cell12_stable_text(source_row["Unit Compatibility Status"])
        result_label = {
            "Compatible": "Expected class found among unit memberships",
            "Incompatible": "Expected class absent from unit memberships",
            "Not assessed": "Not assessed — upstream review required",
        }.get(compatibility_status, compatibility_status or "Not assessed")
        provenance = cell12_arrow_path([
            cell12_stable_text(unit_source_row["Original Unit"]) or "[unit not supplied]",
            f"Canonical unit: {cell12_stable_text(source_row['Canonical Unit'])}",
            f"Possible classes: {'; '.join(memberships)}",
            f"Mnemonic expected class: {cell12_stable_text(source_row['Expected Quantity Class'])}",
            result_label,
            f"Matching class: {matching_class}" if matching_class else "",
        ])

        output_rows.append({
            "Source Row Order": source_row["Source Row Order"],
            "Input Index": source_row["Source Index"],
            "Raw Unit": unit_source_row["Original Unit"],
            "Prepared Unit": source_row["Prepared Unit"],
            "Canonical Unit": source_row["Canonical Unit"],
            "Prepared Mnemonic": source_row["Prepared Mnemonic"],
            "Property Kind": source_row["Property Kind Title"],
            "Mnemonic Expected Quantity Class": source_row["Expected Quantity Class"],
            "Unit Possible Quantity Classes": "; ".join(memberships),
            "Multiple-Class Result": result_label,
            "Matching Quantity Class": source_row["Matching Quantity Class"],
            "Multiple-Class Check Passed": cell12_stable_text(source_row["Cell 10+ Status"]) == "Pass",
            "Possible Dimension(s)": cell12_join([[source_row["Unit Dimension"]], [all_enrichment["dimensions"]]]),
            "Possible Base Unit(s) for Conversion": all_enrichment["base_units"],
            "Matching-Class Base Unit for Conversion": matched_enrichment["base_units"],
            "Matching-Class Member Units": matched_enrichment["member_units"],
            "All Possible-Class Member Units": all_enrichment["grouped_member_units"],
            "Review Reason": cell12_friendly_stage_text(source_row["Review Reason"]),
            "Provenance": provenance,
        })

    columns = [
        "Source Row Order", "Input Index", "Raw Unit", "Prepared Unit", "Canonical Unit",
        "Prepared Mnemonic", "Property Kind", "Mnemonic Expected Quantity Class",
        "Unit Possible Quantity Classes", "Multiple-Class Result", "Matching Quantity Class",
        "Multiple-Class Check Passed", "Possible Dimension(s)", "Possible Base Unit(s) for Conversion",
        "Matching-Class Base Unit for Conversion", "Matching-Class Member Units",
        "All Possible-Class Member Units", "Review Reason", "Provenance",
    ]
    return pd.DataFrame(output_rows, columns=columns)


def cell12_build_semantic_dataframe():
    """Create the final mnemonic–unit comparison view, including every gated review row."""

    output_rows = []

    for source_position in range(len(SEMANTIC_VALIDATION_DF)):
        source_row = SEMANTIC_VALIDATION_DF.iloc[source_position]
        semantic_pass = cell12_stable_text(source_row["Cell 11 Status"]) == "Pass"
        compatibility_status = cell12_stable_text(source_row["Semantic Compatibility Status"])
        multi_class_check_required = bool(source_row["Cell 10+ Check Required"])
        multi_class_check_result = bool(source_row["Cell 10+ Gate Passed"]) if multi_class_check_required else "Not required"

        if semantic_pass:
            final_result = "Semantic match confirmed"
        elif compatibility_status == "Incompatible":
            final_result = "Semantically incompatible — expert review required"
        elif cell12_stable_text(source_row["Semantic Validation Status"]) == "Not assessed":
            final_result = "Not assessed — upstream expert review required"
        else:
            final_result = "Expert review required"

        provenance = cell12_arrow_path([
            cell12_stable_text(source_row["Original Mnemonic"]) or "[mnemonic not supplied]",
            f"Property Kind: {cell12_stable_text(source_row['Property Kind Title'])}" if cell12_stable_text(source_row["Property Kind Title"]) else "",
            f"Expected class: {cell12_stable_text(source_row['Expected Quantity Class'])}" if cell12_stable_text(source_row["Expected Quantity Class"]) else "",
            "↔",
            cell12_stable_text(source_row["Original Unit"]) or "[unit not supplied]",
            f"Canonical unit: {cell12_stable_text(source_row['Canonical Unit'])}" if cell12_stable_text(source_row["Canonical Unit"]) else "",
            f"Unit classes: {cell12_join([source_row['Actual Quantity Class Memberships']])}" if cell12_join([source_row["Actual Quantity Class Memberships"]]) else "",
            final_result,
        ])

        output_rows.append({
            "Source Row Order": source_row["Source Row Order"],
            "Input Index": source_row["Source Index"],
            "Service Company": source_row["Prepared Service Company"],
            "Raw Mnemonic": source_row["Original Mnemonic"],
            "Prepared Mnemonic": source_row["Prepared Mnemonic"],
            "Property Kind": source_row["Property Kind Title"],
            "Expected Quantity Class": source_row["Expected Quantity Class"],
            "Raw Unit": source_row["Original Unit"],
            "Prepared Unit": source_row["Prepared Unit"],
            "Canonical Unit": source_row["Canonical Unit"],
            "Unit Quantity Class Memberships": cell12_join([source_row["Actual Quantity Class Memberships"]]),
            "Company and Mnemonic Validation Passed": bool(source_row["Cell 9 Gate Passed"]),
            "Unit Validation Passed": bool(source_row["Cell 10 Gate Passed"]),
            "Multiple-Class Check Required": multi_class_check_required,
            "Multiple-Class Check Passed": multi_class_check_result,
            "Semantic Comparison Performed": bool(source_row["Semantic Comparison Performed"]),
            "Matching Quantity Class": source_row["Matching Quantity Class"],
            "Semantic Compatibility": source_row["Semantic Compatibility Status"],
            "Semantic Pass": semantic_pass,
            "Final Result": final_result,
            "Expert Review Required": bool(source_row["Expert Review Required"]),
            "Review Origin": cell12_friendly_stage_text(source_row["Review Origin"]),
            "Review Reason": cell12_friendly_stage_text(source_row["Review Reason"]),
            "Provenance": provenance,
        })

    columns = [
        "Source Row Order", "Input Index", "Service Company", "Raw Mnemonic",
        "Prepared Mnemonic", "Property Kind", "Expected Quantity Class", "Raw Unit",
        "Prepared Unit", "Canonical Unit", "Unit Quantity Class Memberships",
        "Company and Mnemonic Validation Passed", "Unit Validation Passed", "Multiple-Class Check Required",
        "Multiple-Class Check Passed", "Semantic Comparison Performed",
        "Matching Quantity Class", "Semantic Compatibility", "Semantic Pass",
        "Final Result", "Expert Review Required", "Review Origin", "Review Reason", "Provenance",
    ]
    return pd.DataFrame(output_rows, columns=columns)


def cell12_review_action(review_origin, review_reason):
    """Translate final review provenance into a practical expert action."""

    combined_text = f"{cell12_stable_text(review_origin)} {cell12_stable_text(review_reason)}".casefold()
    has_cell9 = "cell 9" in combined_text
    has_cell10 = "cell 10" in combined_text and "cell 10+" not in combined_text
    has_cell10plus = "cell 10+" in combined_text or "multiple" in combined_text
    has_cell11 = "cell 11" in combined_text or "incompatib" in combined_text

    if sum([has_cell9, has_cell10, has_cell10plus, has_cell11]) > 1:
        return "Resolve every listed upstream issue, record governed aliases where appropriate, then rerun the validation and output stages."

    if has_cell9:
        return "Verify the service company and mnemonic, then confirm the canonical Property Kind and expected Quantity Class."

    if has_cell10plus:
        return "Confirm which Quantity Class applies to this recognized multi-membership unit in the mnemonic context."

    if has_cell10:
        return "Verify the source unit and canonical UOM symbol; add a governed unit alias only after expert approval."

    if has_cell11:
        return "Resolve the mismatch between the mnemonic-expected class and the canonical unit memberships."

    return "Review the preserved evidence and update the governed Expert Review Workbook when a decision is approved."


def cell12_build_review_dataframe():
    """Create one action-focused row for every final expert-review item."""

    output_rows = []

    for review_position in range(len(SEMANTIC_REVIEW_DF)):
        source_row = SEMANTIC_REVIEW_DF.iloc[review_position]
        source_row_order = int(source_row["Source Row Order"])
        mnemonic_row = MNEMONIC_VALIDATION_DF.iloc[source_row_order - 1]
        unit_row = UNIT_VALIDATION_DF.iloc[source_row_order - 1]
        candidate_properties = cell12_ordered_text([
            mnemonic_row["Property Kind Title"],
            mnemonic_row["Candidate Property Kind"],
            mnemonic_row["Candidate Property Kind Titles"],
        ])
        candidate_classes = cell12_ordered_text([
            mnemonic_row["Expected Quantity Class"],
            mnemonic_row["Candidate Quantity Classes"],
        ])

        output_rows.append({
            "Source Row Order": source_row["Source Row Order"],
            "Input Index": source_row["Source Index"],
            "Raw Service Company": source_row["Original Service Company"],
            "Raw Mnemonic": source_row["Original Mnemonic"],
            "Raw Unit": source_row["Original Unit"],
            "Raw Description": source_row["Original Description"],
            "Prepared Service Company": source_row["Prepared Service Company"],
            "Prepared Mnemonic": source_row["Prepared Mnemonic"],
            "Prepared Unit": source_row["Prepared Unit"],
            "Mnemonic Result": CELL12_MNEMONIC_RESULT_LABELS.get(cell12_stable_text(mnemonic_row["Mnemonic Match Method"]), mnemonic_row["Mnemonic Match Method"]),
            "Candidate Property Kind(s)": "; ".join(candidate_properties),
            "Candidate Expected Quantity Class(es)": "; ".join(candidate_classes),
            "Unit Result": CELL12_UNIT_RESULT_LABELS.get(cell12_stable_text(unit_row["Unit Match Method"]), unit_row["Unit Match Method"]),
            "Candidate Canonical Unit(s)": cell12_join([[unit_row["Canonical Unit"]], [unit_row["Candidate Canonical Units"]]]),
            "Unit Quantity Class Memberships": cell12_join([unit_row["Actual Quantity Class Memberships"]]),
            "Semantic Compatibility": source_row["Semantic Compatibility Status"],
            "Review Origin": cell12_friendly_stage_text(source_row["Review Origin"]),
            "Review Reason": cell12_friendly_stage_text(source_row["Review Reason"]),
            "Suggested Expert Action": cell12_review_action(source_row["Review Origin"], source_row["Review Reason"]),
        })

    columns = [
        "Source Row Order", "Input Index", "Raw Service Company", "Raw Mnemonic",
        "Raw Unit", "Raw Description", "Prepared Service Company", "Prepared Mnemonic",
        "Prepared Unit", "Mnemonic Result", "Candidate Property Kind(s)",
        "Candidate Expected Quantity Class(es)", "Unit Result", "Candidate Canonical Unit(s)",
        "Unit Quantity Class Memberships", "Semantic Compatibility", "Review Origin",
        "Review Reason", "Suggested Expert Action",
    ]
    return pd.DataFrame(output_rows, columns=columns)


def cell12_status_fill(cell, column_name):
    """Apply restrained flag colors without making color the only source of meaning."""

    value = cell.value
    column_text = column_name.casefold()

    if isinstance(value, bool):
        if "review" in column_text:
            use_pass_fill = not value
        elif "check required" in column_text:
            if value:
                cell.fill = PatternFill(fill_type="solid", fgColor=CELL12_CANDIDATE_FILL)
                cell.font = Font(color=CELL12_CANDIDATE_FONT)
            else:
                cell.fill = PatternFill(fill_type="solid", fgColor=CELL12_NEUTRAL_FILL)
                cell.font = Font(color=CELL12_NEUTRAL_FONT)
            return
        else:
            use_pass_fill = value

        if use_pass_fill:
            cell.fill = PatternFill(fill_type="solid", fgColor=CELL12_PASS_FILL)
            cell.font = Font(color=CELL12_PASS_FONT, bold=True)
        else:
            cell.fill = PatternFill(fill_type="solid", fgColor=CELL12_REVIEW_FILL)
            cell.font = Font(color=CELL12_REVIEW_FONT, bold=True)
        return

    text_value = cell12_stable_text(value).casefold()

    if not text_value:
        return

    pass_values = ["pass", "compatible", "recognized", "resolved", "agree", "semantic match confirmed", "exact match", "alias match"]
    review_values = ["review required", "incompatible", "unresolved", "not recognized", "no match", "ambiguous", "conflict", "absent"]
    candidate_values = ["candidate", "not assessed", "deferred", "not provided", "not supplied", "missing"]

    if any(token in text_value for token in pass_values) and not any(token in text_value for token in review_values):
        cell.fill = PatternFill(fill_type="solid", fgColor=CELL12_PASS_FILL)
        cell.font = Font(color=CELL12_PASS_FONT)
    elif any(token in text_value for token in review_values):
        cell.fill = PatternFill(fill_type="solid", fgColor=CELL12_REVIEW_FILL)
        cell.font = Font(color=CELL12_REVIEW_FONT)
    elif any(token in text_value for token in candidate_values):
        cell.fill = PatternFill(fill_type="solid", fgColor=CELL12_CANDIDATE_FILL)
        cell.font = Font(color=CELL12_CANDIDATE_FONT)


def cell12_write_dataframe_sheet(workbook, sheet_name, title, subtitle, dataframe, wrap_columns=None, width_overrides=None):
    """Write one clean, filterable output table with consistent report styling."""

    wrap_columns = set(wrap_columns or [])
    width_overrides = width_overrides or {}
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 90
    worksheet.freeze_panes = "A5"
    worksheet.sheet_properties.tabColor = CELL12_HEADER_FILL[-6:]

    column_count = max(1, len(dataframe.columns))
    last_column_letter = get_column_letter(column_count)

    if column_count > 1:
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=column_count)

    title_cell = worksheet.cell(row=1, column=1, value=title)
    title_cell.fill = PatternFill(fill_type="solid", fgColor=CELL12_TITLE_FILL)
    title_cell.font = Font(color="FFFFFFFF", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 26

    subtitle_cell = worksheet.cell(row=2, column=1, value=subtitle)
    subtitle_cell.fill = PatternFill(fill_type="solid", fgColor=CELL12_SUBHEADER_FILL)
    subtitle_cell.font = Font(color=CELL12_BODY_FONT, italic=True, size=10)
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    worksheet.row_dimensions[2].height = 32

    header_row = 4

    for column_position, column_name in enumerate(dataframe.columns, start=1):
        header_cell = worksheet.cell(row=header_row, column=column_position, value=column_name)
        header_cell.fill = PatternFill(fill_type="solid", fgColor=CELL12_HEADER_FILL)
        header_cell.font = Font(color="FF000000", bold=True)
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    worksheet.row_dimensions[header_row].height = 38

    for source_position in range(len(dataframe)):
        excel_row = header_row + source_position + 1

        for column_position, column_name in enumerate(dataframe.columns, start=1):
            source_value = dataframe.iat[source_position, column_position - 1]
            excel_value = cell12_excel_value(source_value)
            output_cell = worksheet.cell(row=excel_row, column=column_position, value=excel_value)

            # Preserve values beginning with '=' as literal source text, not formulas.
            if isinstance(excel_value, str) and excel_value.startswith("="):
                output_cell.data_type = "s"

            output_cell.font = Font(color=CELL12_BODY_FONT, size=10)
            output_cell.alignment = Alignment(
                horizontal="center" if isinstance(excel_value, bool) else "left",
                vertical="top",
                wrap_text=column_name in wrap_columns,
            )

            if isinstance(excel_value, str) and excel_value.startswith(("https://", "http://")):
                output_cell.hyperlink = excel_value
                output_cell.font = Font(color=CELL12_LINK_FONT, underline="single", size=10)

            if any(token in column_name.casefold() for token in ["pass", "status", "result", "required", "compatibility", "resolution", "flag group"]):
                cell12_status_fill(output_cell, column_name)

        if wrap_columns:
            worksheet.row_dimensions[excel_row].height = 42

    if dataframe.empty:
        note_cell = worksheet.cell(row=6, column=1, value="No records were produced for this section.")
        note_cell.font = Font(color=CELL12_NEUTRAL_FONT, italic=True)

    worksheet.auto_filter.ref = f"A{header_row}:{last_column_letter}{max(header_row, header_row + len(dataframe))}"

    sample_size = min(len(dataframe), 1000)

    for column_position, column_name in enumerate(dataframe.columns, start=1):
        if column_name in width_overrides:
            column_width = width_overrides[column_name]
        else:
            content_lengths = [len(str(column_name))]

            for source_position in range(sample_size):
                source_value = cell12_excel_value(dataframe.iat[source_position, column_position - 1])

                if source_value is not None:
                    content_lengths.append(min(len(str(source_value)), 70))

            column_width = min(max(content_lengths) + 2, 48)
            column_width = max(column_width, 12)

        worksheet.column_dimensions[get_column_letter(column_position)].width = column_width

    if sheet_name == "2 Treatment Statistics":
        value_column = list(dataframe.columns).index("Value") + 1
        type_column = list(dataframe.columns).index("Metric Type") + 1
        share_column = list(dataframe.columns).index("Share of Input") + 1

        for source_position in range(len(dataframe)):
            excel_row = header_row + source_position + 1

            if worksheet.cell(excel_row, type_column).value == "Percent":
                worksheet.cell(excel_row, value_column).number_format = "0.0%"
            elif isinstance(worksheet.cell(excel_row, value_column).value, (int, float)):
                worksheet.cell(excel_row, value_column).number_format = "#,##0.00" if isinstance(worksheet.cell(excel_row, value_column).value, float) else "#,##0"

            worksheet.cell(excel_row, share_column).number_format = "0.0%"

    return worksheet


def cell12_ordered_counts(values, label_map=None):
    """Count values in first-seen order for chart-driving tables."""

    label_map = label_map or {}
    ordered_labels = []
    counts = {}

    for value in values:
        raw_label = cell12_stable_text(value) or "Not available"
        display_label = label_map.get(raw_label, raw_label)

        if display_label not in counts:
            ordered_labels.append(display_label)
            counts[display_label] = 0

        counts[display_label] += 1

    return [(label, counts[label]) for label in ordered_labels]


def cell12_write_chart_source_table(worksheet, start_row, table_title, rows):
    """Write a visible, auditable two-column table that drives one native Excel chart."""

    worksheet.cell(start_row, 1, table_title)
    worksheet.cell(start_row, 1).fill = PatternFill(fill_type="solid", fgColor=CELL12_TITLE_FILL)
    worksheet.cell(start_row, 1).font = Font(color="FFFFFFFF", bold=True)
    worksheet.cell(start_row + 1, 1, "Category")
    worksheet.cell(start_row + 1, 2, "Rows")

    for column_position in [1, 2]:
        header_cell = worksheet.cell(start_row + 1, column_position)
        header_cell.fill = PatternFill(fill_type="solid", fgColor=CELL12_HEADER_FILL)
        header_cell.font = Font(color="FF000000", bold=True)

    chart_rows = rows or [("No records", 0)]

    for row_offset, (category_name, category_value) in enumerate(chart_rows, start=2):
        worksheet.cell(start_row + row_offset, 1, category_name)
        worksheet.cell(start_row + row_offset, 2, category_value)
        worksheet.cell(start_row + row_offset, 2).number_format = "#,##0"

    return start_row + 2, start_row + len(chart_rows) + 1


def cell12_build_chart_sheet(workbook):
    """Create four focused charts from visible source tables in the workbook."""

    worksheet = workbook.create_sheet("9 Charts")
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 85
    worksheet.sheet_properties.tabColor = CELL12_HEADER_FILL[-6:]
    worksheet.merge_cells("A1:R1")
    worksheet.merge_cells("A2:R2")
    worksheet["A1"] = "Validation Results — Charts"
    worksheet["A1"].fill = PatternFill(fill_type="solid", fgColor=CELL12_TITLE_FILL)
    worksheet["A1"].font = Font(color="FFFFFFFF", bold=True, size=14)
    worksheet["A2"] = "Charts summarize the same row-level flags shown in the treatment and validation sheets."
    worksheet["A2"].fill = PatternFill(fill_type="solid", fgColor=CELL12_SUBHEADER_FILL)
    worksheet["A2"].font = Font(color=CELL12_BODY_FONT, italic=True)
    worksheet.column_dimensions["A"].width = 46
    worksheet.column_dimensions["B"].width = 14

    input_rows = len(SEMANTIC_VALIDATION_DF)
    final_passes = int(SEMANTIC_VALIDATION_DF["Cell 11 Status"].eq("Pass").sum())
    final_reviews = int(SEMANTIC_VALIDATION_DF["Expert Review Required"].eq(True).sum())
    final_rows = [("Pass", final_passes), ("Expert review", final_reviews)]
    mnemonic_rows = cell12_ordered_counts(MNEMONIC_VALIDATION_DF["Mnemonic Match Method"].tolist(), CELL12_MNEMONIC_RESULT_LABELS)
    unit_rows = cell12_ordered_counts(UNIT_VALIDATION_DF["Unit Match Method"].tolist(), CELL12_UNIT_RESULT_LABELS)
    pipeline_rows = [
        ("Input rows", input_rows),
        ("Company and mnemonic validation passes", int(SEMANTIC_VALIDATION_DF["Cell 9 Gate Passed"].eq(True).sum())),
        ("Unit validation passes", int(SEMANTIC_VALIDATION_DF["Cell 10 Gate Passed"].eq(True).sum())),
        ("Semantic comparisons", int(SEMANTIC_VALIDATION_DF["Semantic Comparison Performed"].eq(True).sum())),
        ("Final semantic passes", final_passes),
    ]

    final_first, final_last = cell12_write_chart_source_table(worksheet, 4, "Final workflow outcome", final_rows)
    mnemonic_first, mnemonic_last = cell12_write_chart_source_table(worksheet, 10, "Mnemonic match routes", mnemonic_rows)
    unit_start = max(20, mnemonic_last + 3)
    unit_first, unit_last = cell12_write_chart_source_table(worksheet, unit_start, "Unit match routes", unit_rows)
    pipeline_start = max(32, unit_last + 3)
    pipeline_first, pipeline_last = cell12_write_chart_source_table(worksheet, pipeline_start, "Validation pipeline", pipeline_rows)

    outcome_chart = DoughnutChart()
    outcome_chart.title = "Final workflow outcome"
    outcome_chart.add_data(Reference(worksheet, min_col=2, min_row=5, max_row=final_last), titles_from_data=True)
    outcome_chart.set_categories(Reference(worksheet, min_col=1, min_row=final_first, max_row=final_last))
    outcome_chart.holeSize = 55
    outcome_chart.varyColors = True
    outcome_chart.height = 9
    outcome_chart.width = 13
    outcome_chart.legend.position = "r"
    worksheet.add_chart(outcome_chart, "D4")

    pipeline_chart = BarChart()
    pipeline_chart.type = "bar"
    pipeline_chart.style = 10
    pipeline_chart.title = "Rows retained through validation gates"
    pipeline_chart.add_data(Reference(worksheet, min_col=2, min_row=pipeline_start + 1, max_row=pipeline_last), titles_from_data=True)
    pipeline_chart.set_categories(Reference(worksheet, min_col=1, min_row=pipeline_first, max_row=pipeline_last))
    pipeline_chart.height = 9
    pipeline_chart.width = 14
    pipeline_chart.legend = None
    pipeline_chart.x_axis.title = "Rows"
    pipeline_chart.x_axis.numFmt = "0"
    pipeline_chart.x_axis.majorUnit = 1
    pipeline_chart.y_axis.scaling.orientation = "maxMin"
    worksheet.add_chart(pipeline_chart, "L4")

    mnemonic_chart = BarChart()
    mnemonic_chart.type = "bar"
    mnemonic_chart.style = 10
    mnemonic_chart.title = "Mnemonic validation routes"
    mnemonic_chart.add_data(Reference(worksheet, min_col=2, min_row=11, max_row=mnemonic_last), titles_from_data=True)
    mnemonic_chart.set_categories(Reference(worksheet, min_col=1, min_row=mnemonic_first, max_row=mnemonic_last))
    mnemonic_chart.height = 10
    mnemonic_chart.width = 14
    mnemonic_chart.legend = None
    mnemonic_chart.x_axis.title = "Rows"
    mnemonic_chart.x_axis.numFmt = "0"
    mnemonic_chart.x_axis.majorUnit = 1
    mnemonic_chart.y_axis.scaling.orientation = "maxMin"
    worksheet.add_chart(mnemonic_chart, "D23")

    unit_chart = BarChart()
    unit_chart.type = "bar"
    unit_chart.style = 10
    unit_chart.title = "Unit validation routes"
    unit_chart.add_data(Reference(worksheet, min_col=2, min_row=unit_start + 1, max_row=unit_last), titles_from_data=True)
    unit_chart.set_categories(Reference(worksheet, min_col=1, min_row=unit_first, max_row=unit_last))
    unit_chart.height = 10
    unit_chart.width = 14
    unit_chart.legend = None
    unit_chart.x_axis.title = "Rows"
    unit_chart.x_axis.numFmt = "0"
    unit_chart.x_axis.majorUnit = 1
    unit_chart.y_axis.scaling.orientation = "maxMin"
    worksheet.add_chart(unit_chart, "L23")

    return worksheet


def cell12_create_workbook(output_path, export_time, tables):
    """Create the complete nine-sheet workbook from the cleaned report tables."""

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "Younes Belouettar"
    workbook.properties.title = "Well Metadata Semantic Validation Output"
    workbook.properties.subject = "PWLS mnemonic, UOM unit, and semantic Quantity Class validation"
    workbook.properties.description = "Output Preparation export; raw input preserved and technical results translated for review."
    workbook.properties.created = export_time

    cell12_write_dataframe_sheet(
        workbook,
        "1 Sources",
        "Input and Reference Sources",
        "Original input path, GitLab/GitHub provenance, immutable commits, imported files, and source statistics.",
        tables["sources"],
        wrap_columns={"Value", "Notes"},
        width_overrides={"Section": 24, "Source": 34, "Item": 36, "Value": 64, "Notes": 58},
    )
    cell12_write_dataframe_sheet(
        workbook,
        "2 Treatment Statistics",
        "Data Treatment Statistics",
        "Counts and rates follow the company, mnemonic, unit, multi-membership, and semantic-validation flags.",
        tables["statistics"],
        wrap_columns={"Meaning"},
        width_overrides={"Phase": 36, "Flag Group": 25, "Metric": 50, "Value": 16, "Metric Type": 14, "Share of Input": 16, "Meaning": 70},
    )
    cell12_write_dataframe_sheet(
        workbook,
        "3 Raw Input",
        "Original Raw Input",
        "Untouched source metadata. Source Row Order is positional; Input Index preserves the original DataFrame index, including duplicates.",
        tables["raw"],
        wrap_columns={"description"},
        width_overrides={"Source Row Order": 18, "Input Index": 18, "Service company": 28, "mnemonic": 22, "unit": 18, "description": 70},
    )
    cell12_write_dataframe_sheet(
        workbook,
        "4 Mnemonic Check",
        "Mnemonic and Property Kind Check",
        "Human-readable company and mnemonic validation results. Case-insensitive PWLS results remain review candidates and never receive an automatic pass.",
        tables["mnemonic"],
        wrap_columns={"Found Property Kind(s)", "Found Quantity Class(es)", "Review Reason", "Provenance"},
        width_overrides={"Source Row Order": 18, "Input Index": 16, "Raw Service Company": 26, "Prepared Service Company": 26, "Mnemonic Check Result": 40, "Found Property Kind(s)": 42, "Found Quantity Class(es)": 34, "Review Reason": 64, "Provenance": 78},
    )
    cell12_write_dataframe_sheet(
        workbook,
        "5 Unit Check",
        "Unit Recognition and Enrichment",
        "Human-readable unit-validation results. Quantity Class, dimension, conversion-base, and member-unit enrichment is shown only for passing recognized units.",
        tables["unit"],
        wrap_columns={"Possible Quantity Class(es)", "Possible Dimension(s)", "Member Units in Possible Class(es)", "Review Reason", "Provenance"},
        width_overrides={"Source Row Order": 18, "Input Index": 16, "Unit Check Result": 42, "Possible Quantity Class(es)": 38, "Possible Dimension(s)": 28, "Base Unit(s) for Conversion": 30, "Member Units in Possible Class(es)": 85, "Review Reason": 62, "Provenance": 78},
    )
    cell12_write_dataframe_sheet(
        workbook,
        "6 Multiple Unit Classes",
        "Units with Multiple Quantity Classes",
        "Every recognized unit having more than one Quantity Class membership, its multiple-class compatibility result, and relevant UOM enrichment.",
        tables["multi"],
        wrap_columns={"Unit Possible Quantity Classes", "Multiple-Class Result", "Matching-Class Member Units", "All Possible-Class Member Units", "Review Reason", "Provenance"},
        width_overrides={"Source Row Order": 18, "Input Index": 16, "Property Kind": 36, "Mnemonic Expected Quantity Class": 34, "Unit Possible Quantity Classes": 52, "Multiple-Class Result": 46, "Matching-Class Member Units": 75, "All Possible-Class Member Units": 90, "Review Reason": 62, "Provenance": 78},
    )
    cell12_write_dataframe_sheet(
        workbook,
        "7 Semantic Validation",
        "Mnemonic–Unit Semantic Validation",
        "Final semantic-validation gate and exact Quantity Class membership result for every original input row.",
        tables["semantic"],
        wrap_columns={"Unit Quantity Class Memberships", "Final Result", "Review Origin", "Review Reason", "Provenance"},
        width_overrides={"Source Row Order": 18, "Input Index": 16, "Property Kind": 36, "Expected Quantity Class": 32, "Unit Quantity Class Memberships": 48, "Final Result": 48, "Review Origin": 30, "Review Reason": 68, "Provenance": 86},
    )
    cell12_write_dataframe_sheet(
        workbook,
        "8 Expert Review Queue",
        "Final Expert Review Queue",
        "Action-focused subset of final review rows. No alias or validation decision is written automatically.",
        tables["review"],
        wrap_columns={"Raw Description", "Candidate Property Kind(s)", "Candidate Expected Quantity Class(es)", "Unit Quantity Class Memberships", "Review Reason", "Suggested Expert Action"},
        width_overrides={"Source Row Order": 18, "Input Index": 16, "Raw Description": 60, "Candidate Property Kind(s)": 44, "Candidate Expected Quantity Class(es)": 38, "Unit Quantity Class Memberships": 46, "Review Origin": 30, "Review Reason": 72, "Suggested Expert Action": 78},
    )
    cell12_build_chart_sheet(workbook)
    workbook.save(output_path)
    return workbook


def cell12_verify_workbook(output_path, tables):
    """Reload the saved workbook and verify sheets, row counts, headers, charts, and cell errors."""

    if not output_path.exists() or output_path.stat().st_size == 0:
        raise AssertionError("The Cell 12 output workbook was not saved correctly.")

    verification_workbook = load_workbook(output_path, data_only=False)

    try:
        if verification_workbook.sheetnames != CELL12_SHEET_NAMES:
            raise AssertionError(
                f"Unexpected output sheets. Expected: {CELL12_SHEET_NAMES}. "
                f"Received: {verification_workbook.sheetnames}"
            )

        table_sheet_map = {
            "1 Sources": "sources",
            "2 Treatment Statistics": "statistics",
            "3 Raw Input": "raw",
            "4 Mnemonic Check": "mnemonic",
            "5 Unit Check": "unit",
            "6 Multiple Unit Classes": "multi",
            "7 Semantic Validation": "semantic",
            "8 Expert Review Queue": "review",
        }

        for sheet_name, table_name in table_sheet_map.items():
            worksheet = verification_workbook[sheet_name]
            expected_headers = list(tables[table_name].columns)
            actual_headers = [worksheet.cell(4, column_position).value for column_position in range(1, len(expected_headers) + 1)]

            if actual_headers != expected_headers:
                raise AssertionError(f"{sheet_name} headers do not match the cleaned output schema.")

            expected_max_row = 4 + len(tables[table_name])

            if not tables[table_name].empty and worksheet.max_row != expected_max_row:
                raise AssertionError(f"{sheet_name} does not contain the expected number of output rows.")

        if len(verification_workbook["9 Charts"]._charts) != 4:
            raise AssertionError("The chart sheet does not contain all four requested charts.")

        for worksheet in verification_workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type == "e":
                        raise AssertionError(f"Excel error detected in {worksheet.title}!{cell.coordinate}: {cell.value}")

    finally:
        verification_workbook.close()


try:
    _cell12_partial_output_path = None
    cell12_validate_inputs()
    _cell12_dataframe_snapshots, _cell12_dictionary_snapshots, _cell12_scalar_snapshots = cell12_create_snapshots()

    # The folder is selected only after all dependencies have been validated, so the
    # user is not interrupted by a file dialog when an upstream cell must be rerun.
    _cell12_export_time = datetime.now()
    _cell12_output_folder = cell12_choose_output_folder()
    _cell12_partial_output_path = cell12_create_output_path(_cell12_output_folder, _cell12_export_time)
    _cell12_quantity_class_index = cell12_build_quantity_class_index()

    _cell12_tables = {
        "sources": cell12_build_sources_dataframe(_cell12_partial_output_path, _cell12_export_time),
        "statistics": cell12_build_statistics_dataframe(),
        "raw": cell12_build_raw_dataframe(),
        "mnemonic": cell12_build_mnemonic_dataframe(),
        "unit": cell12_build_unit_dataframe(_cell12_quantity_class_index),
        "multi": cell12_build_multi_class_dataframe(_cell12_quantity_class_index),
        "semantic": cell12_build_semantic_dataframe(),
        "review": cell12_build_review_dataframe(),
    }

    _cell12_workbook = cell12_create_workbook(
        _cell12_partial_output_path,
        _cell12_export_time,
        _cell12_tables,
    )
    _cell12_workbook.close()

    cell12_verify_workbook(_cell12_partial_output_path, _cell12_tables)
    cell12_verify_upstream_unchanged(
        _cell12_dataframe_snapshots,
        _cell12_dictionary_snapshots,
        _cell12_scalar_snapshots,
    )

    CELL12_OUTPUT_PATH = _cell12_partial_output_path
    CELL12_EXPORT_SUMMARY = {
        "Output workbook": str(CELL12_OUTPUT_PATH),
        "Generated at": _cell12_export_time.strftime("%Y-%m-%d %H:%M:%S"),
        "Sheets created": len(CELL12_SHEET_NAMES),
        "Input rows": len(METADATA_ORIGINAL_DF),
        "Final semantic passes": int(SEMANTIC_VALIDATION_DF["Cell 11 Status"].eq("Pass").sum()),
        "Final expert-review rows": int(SEMANTIC_VALIDATION_DF["Expert Review Required"].eq(True).sum()),
    }

    print("Output Preparation completed successfully.")

    for summary_name, summary_value in CELL12_EXPORT_SUMMARY.items():
        print(f"{summary_name}: {summary_value}")

except Exception as error:
    # A failed export must not leave a partial workbook that could be mistaken for a
    # verified result. Only the unique file created during this Cell 12 run is removed.
    if "_cell12_partial_output_path" in locals() and _cell12_partial_output_path is not None:
        try:
            if Path(_cell12_partial_output_path).exists():
                Path(_cell12_partial_output_path).unlink()
        except OSError:
            pass

    CELL12_OUTPUT_PATH = None
    CELL12_EXPORT_SUMMARY = {}

    print("Cell 12 failed.")
    print(f"Error: {error}")

